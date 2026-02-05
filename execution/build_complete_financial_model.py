#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Complete 14-Sheet Financial Model Builder

Creates a comprehensive financial model with all 14 sheets:
1. Sources & References - TAM/SAM/SOM with research links
2. Assumptions - All input parameters
3. Headcount Plan - Team growth by department
4. Revenue - Multi-stream revenue calculations
5. Operating Costs - COGS, Fixed, S&M
6. P&L - Profit & Loss with margins
7. Cash Flow - Operating, Investing, Financing
8. Balance Sheet - Assets, Liabilities, Equity
9. Summary - KPI dashboard
10. Sensitivity Analysis - Scenario modeling
11. Valuation - DCF and multiples
12. Break-even Analysis - Contribution margin
13. Funding Cap Table - Equity rounds
14. Charts Data - Data for visualizations

Usage:
    python build_complete_financial_model.py --config .tmp/rapidtools/config/rapidtools_config.json
    python build_complete_financial_model.py --company "RapidTools" --years 8
"""

import argparse
import json
import os
import sys
from datetime import datetime
from typing import Any, Dict, List, Optional

try:
    from openpyxl import Workbook
    from openpyxl.formatting.rule import FormulaRule
    from openpyxl.styles import Alignment, Border, Font, NamedStyle, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


# =============================================================================
# COLOR PALETTE
# =============================================================================
class Colors:
    """Standard color palette matching the template."""

    TITLE_BLUE = "335080"  # Main titles - dark blue
    DARK_BLUE = "336699"  # Section headers
    MEDIUM_BLUE = "6699CC"  # Category headers
    SECTION_A_CAT = "4D80B3"  # Section A categories
    LIGHT_BLUE = "D8EAF9"  # Zebra stripe / alternating rows
    LIGHT_GRAY = "F2F2F2"  # Column headers
    GREEN = "E5F8E5"  # Total/summary rows
    LIGHT_GREEN = "C6EFCE"  # Positive values
    LIGHT_RED = "FFC7CE"  # Negative values
    YELLOW = "FFEB9C"  # Warning/attention
    WHITE = "FFFFFF"
    BLACK = "000000"
    URL_BLUE = "1A4CB3"  # Hyperlinks
    GRAY = "808080"  # Notes


# =============================================================================
# STYLING HELPERS
# =============================================================================
def create_header_style(
    bg_color=Colors.DARK_BLUE, font_color=Colors.WHITE, size=12, bold=True
):
    """Create header cell styling."""
    return {
        "font": Font(name="Calibri", size=size, bold=bold, color=font_color),
        "fill": PatternFill(
            start_color=bg_color, end_color=bg_color, fill_type="solid"
        ),
        "alignment": Alignment(horizontal="center", vertical="center", wrap_text=True),
        "border": Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        ),
    }


def apply_style(cell, style_dict):
    """Apply a style dictionary to a cell."""
    for attr, value in style_dict.items():
        setattr(cell, attr, value)


def style_header(cell, bg=Colors.DARK_BLUE, fg=Colors.WHITE, size=12, bold=True):
    """Apply header styling to a cell."""
    cell.font = Font(name="Calibri", size=size, bold=bold, color=fg)
    cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )


def style_title(cell, text, bg=Colors.TITLE_BLUE):
    """Apply title styling."""
    cell.value = text
    cell.font = Font(name="Calibri", size=14, bold=True, color=Colors.WHITE)
    cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")


def style_section_header(cell, text, bg=Colors.DARK_BLUE):
    """Apply section header styling."""
    cell.value = text
    style_header(cell, bg=bg, size=11)


def style_data_row(ws, row, start_col, end_col, is_total=False, zebra=False):
    """Apply styling to a data row."""
    bg = None
    if is_total:
        bg = Colors.GREEN
    elif zebra:
        bg = Colors.LIGHT_BLUE

    if bg:
        for col in range(start_col, end_col + 1):
            ws.cell(row, col).fill = PatternFill(
                start_color=bg, end_color=bg, fill_type="solid"
            )


def style_category_header(ws, row, text, end_col=10):
    """Apply category header styling with light blue fill across all columns."""
    ws.cell(row, 1).value = text
    ws.cell(row, 1).font = Font(name="Calibri", size=10, bold=True, color=Colors.BLACK)
    # Fill all columns with light blue background
    for col in range(1, end_col + 1):
        ws.cell(row, col).fill = PatternFill(
            start_color=Colors.LIGHT_BLUE,
            end_color=Colors.LIGHT_BLUE,
            fill_type="solid",
        )


def style_column_headers(ws, row, headers, bg=Colors.MEDIUM_BLUE, end_col=10):
    """Apply column header styling with background fill extending to end_col."""
    for col, header in enumerate(headers, 1):
        style_header(ws.cell(row, col), bg=bg, size=10)
        ws.cell(row, col).value = header
    # Fill remaining columns with same background
    for col in range(len(headers) + 1, end_col + 1):
        ws.cell(row, col).fill = PatternFill(
            start_color=bg, end_color=bg, fill_type="solid"
        )


# =============================================================================
# YEAR HELPERS
# =============================================================================
def get_year_headers(num_years: int = 11) -> List[str]:
    """Get year headers from Year 0 to Year N."""
    return [f"Year {i}" for i in range(num_years)]


def get_year_columns(start_col: int = 3, num_years: int = 11) -> Dict[int, str]:
    """Get mapping of year index to column letter."""
    return {yr: get_column_letter(start_col + yr) for yr in range(num_years)}


# =============================================================================
# SHEET BUILDERS
# =============================================================================


class FinancialModelBuilder:
    """Builds a complete 14-sheet financial model."""

    def __init__(self, config: Dict[str, Any], num_years: int = 11):
        self.config = config
        self.num_years = num_years
        self.wb = Workbook()

        # Remove default sheet
        if "Sheet" in self.wb.sheetnames:
            del self.wb["Sheet"]

        # Track row positions for cross-sheet references
        self.row_refs = {}

    def build(self) -> Workbook:
        """Build all 14 sheets."""
        print("Building 14-sheet financial model...")
        print("=" * 60)

        # Build sheets in order (order matters for cross-references)
        self._build_sources_references()
        self._build_assumptions()
        self._build_headcount()
        self._build_revenue()
        self._build_operating_costs()
        self._build_pnl()
        self._build_cash_flow()
        self._build_balance_sheet()
        self._build_summary()
        self._build_sensitivity()
        self._build_valuation()
        self._build_breakeven()
        self._build_cap_table()
        self._build_charts_data()

        print("=" * 60)
        print(f"âœ… Created {len(self.wb.sheetnames)} sheets")

        return self.wb

    def _build_sources_references(self):
        """Sheet 1: Sources & References - Comprehensive TAM/SAM/SOM with FORMULAS for all calculated values."""
        print("  Building Sources & References...")
        ws = self.wb.create_sheet("Sources & References", 0)

        # Column widths for comprehensive layout
        ws.column_dimensions["A"].width = 48
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 15
        ws.column_dimensions["F"].width = 15
        ws.column_dimensions["G"].width = 15
        ws.column_dimensions["H"].width = 15
        ws.column_dimensions["I"].width = 15
        ws.column_dimensions["J"].width = 50

        # Title
        style_title(ws["A1"], "SOURCES & REFERENCES")
        ws.merge_cells("A1:J1")

        row = 3
        # ================================================================
        # SECTION A: KEY METRICS (Linkable Values)
        # ================================================================
        style_section_header(
            ws.cell(row, 1), "SECTION A: KEY METRICS (Linkable Values)"
        )
        ws.merge_cells(f"A{row}:J{row}")
        row += 2

        # ----------------------------------------------------------------
        # TAM - TOTAL ADDRESSABLE MARKET
        # ----------------------------------------------------------------
        style_section_header(
            ws.cell(row, 1),
            "TAM - TOTAL ADDRESSABLE MARKET (2025)",
            bg=Colors.SECTION_A_CAT,
        )
        ws.merge_cells(f"A{row}:J{row}")
        row += 1

        # TAM headers
        tam_headers = [
            "Revenue Stream",
            "Value ($M)",
            "Calculation",
            "Source [Ref#]",
            "Confidence",
        ]
        style_column_headers(ws, row, tam_headers, bg=Colors.MEDIUM_BLUE)
        row += 1

        # Track row numbers for formulas
        tam_data_start = row

        # SOFTWARE TAM section
        style_category_header(ws, row, "SOFTWARE TAM")
        row += 1

        # CAD Software - row for formula reference
        cad_row = row
        ws.cell(row, 1).value = "  CAD Software (Tooling Subset)"
        ws.cell(row, 1).font = Font(size=10)
        ws.cell(row, 2).value = 6100
        ws.cell(row, 2).number_format = "[$-409]#,##0"
        ws.cell(row, 3).value = "Global CAD $12.2B x 50% tooling"
        ws.cell(row, 4).value = "[1] Future Market Insights"
        ws.cell(row, 5).value = "HIGH"
        row += 1

        # CAM Software
        cam_row = row
        ws.cell(row, 1).value = "  CAM Software (Tooling Subset)"
        ws.cell(row, 1).font = Font(size=10)
        ws.cell(row, 2).value = 2800
        ws.cell(row, 2).number_format = "[$-409]#,##0"
        ws.cell(row, 3).value = "Global CAM $3.45B x 80% tooling"
        ws.cell(row, 4).value = "[2] Mordor Intelligence"
        ws.cell(row, 5).value = "HIGH"
        row += 1

        # Other Mfg Software
        other_sw_row = row
        ws.cell(row, 1).value = "  Other Mfg Software (PLM/Simulation)"
        ws.cell(row, 1).font = Font(size=10)
        ws.cell(row, 2).value = 1100
        ws.cell(row, 2).number_format = "[$-409]#,##0"
        ws.cell(row, 3).value = "PLM $0.5B + Simulation $0.4B + Tolerance $0.2B"
        ws.cell(row, 4).value = "[9] MarketsandMarkets"
        ws.cell(row, 5).value = "MEDIUM"
        row += 1

        # Total Software TAM - FORMULA
        sw_total_row = row
        ws.cell(row, 1).value = "  Total Software TAM"
        ws.cell(row, 1).font = Font(bold=True, size=10)
        ws.cell(row, 2).value = f"=B{cad_row}+B{cam_row}+B{other_sw_row}"
        ws.cell(row, 2).number_format = "[$-409]#,##0"
        ws.cell(row, 3).value = "Sum of CAD + CAM + Other"
        ws.cell(row, 5).value = "HIGH"
        style_data_row(ws, row, 1, 5, is_total=True)
        row += 1

        # HARDWARE TAM section
        style_category_header(ws, row, "HARDWARE TAM")
        row += 1

        # Industrial 3D Printing Market
        print_mkt_row = row
        ws.cell(row, 1).value = "  Industrial 3D Printing Market"
        ws.cell(row, 1).font = Font(size=10)
        ws.cell(row, 2).value = 18300
        ws.cell(row, 2).number_format = "[$-409]#,##0"
        ws.cell(row, 3).value = "Global industrial 3D printing 2025"
        ws.cell(row, 4).value = "[6] GM Insights"
        ws.cell(row, 5).value = "HIGH"
        row += 1

        # Tooling Applications %
        tooling_pct_row = row
        ws.cell(row, 1).value = "  Tooling Applications %"
        ws.cell(row, 1).font = Font(size=10)
        ws.cell(row, 2).value = 0.22
        ws.cell(row, 2).number_format = "0%"
        ws.cell(row, 3).value = "Jigs 12% + Molds 7% + Assembly 3%"
        ws.cell(row, 4).value = "[10] Stratasys Annual Report"
        ws.cell(row, 5).value = "HIGH"
        row += 1

        # Total Hardware TAM - FORMULA
        hw_total_row = row
        ws.cell(row, 1).value = "  Total Hardware TAM"
        ws.cell(row, 1).font = Font(bold=True, size=10)
        ws.cell(row, 2).value = f"=B{print_mkt_row}*B{tooling_pct_row}"
        ws.cell(row, 2).number_format = "[$-409]#,##0"
        ws.cell(row, 3).value = "3D Print Market x Tooling %"
        ws.cell(row, 5).value = "HIGH"
        style_data_row(ws, row, 1, 5, is_total=True)
        row += 1

        # CONSUMABLES TAM section
        style_category_header(ws, row, "CONSUMABLES TAM")
        row += 1

        # 3D Printing Materials
        materials_row = row
        ws.cell(row, 1).value = "  3D Printing Materials Market"
        ws.cell(row, 1).font = Font(size=10)
        ws.cell(row, 2).value = 3880
        ws.cell(row, 2).number_format = "[$-409]#,##0"
        ws.cell(row, 3).value = "Global 3D printing materials 2025"
        ws.cell(row, 4).value = "[8] Grand View Research"
        ws.cell(row, 5).value = "HIGH"
        row += 1

        # Tooling Material Intensity
        intensity_row = row
        ws.cell(row, 1).value = "  Tooling Material Subset (22% x 1.33x)"
        ws.cell(row, 1).font = Font(size=10)
        ws.cell(row, 2).value = f"=B{materials_row}*0.22*1.33"
        ws.cell(row, 2).number_format = "[$-409]#,##0"
        ws.cell(row, 3).value = "Materials x 22% x 1.33 intensity"
        ws.cell(row, 4).value = "Calculated"
        ws.cell(row, 5).value = "MEDIUM"
        row += 1

        # Traditional Tooling Materials
        trad_materials_row = row
        ws.cell(row, 1).value = "  Traditional Tooling Materials"
        ws.cell(row, 1).font = Font(size=10)
        ws.cell(row, 2).value = 6870
        ws.cell(row, 2).number_format = "[$-409]#,##0"
        ws.cell(row, 3).value = "CNC $4.2B + Composite $1.5B + Other $1.17B"
        ws.cell(row, 4).value = "Industry Estimates"
        ws.cell(row, 5).value = "MEDIUM"
        row += 1

        # Total Consumables TAM - FORMULA
        cons_total_row = row
        ws.cell(row, 1).value = "  Total Consumables TAM"
        ws.cell(row, 1).font = Font(bold=True, size=10)
        ws.cell(row, 2).value = f"=B{intensity_row}+B{trad_materials_row}"
        ws.cell(row, 2).number_format = "[$-409]#,##0"
        ws.cell(row, 3).value = "3DP Materials + Traditional"
        ws.cell(row, 5).value = "MEDIUM-HIGH"
        style_data_row(ws, row, 1, 5, is_total=True)
        row += 1

        # SERVICES TAM section
        style_category_header(ws, row, "SERVICES TAM (Triple-Validated)")
        row += 1

        # Engineering Services Outsourcing
        eso_row = row
        ws.cell(row, 1).value = "  Engineering Services Outsourcing"
        ws.cell(row, 1).font = Font(size=10)
        ws.cell(row, 2).value = 315610
        ws.cell(row, 2).number_format = "[$-409]#,##0"
        ws.cell(row, 3).value = "Global ESO market 2025"
        ws.cell(row, 4).value = "[32] Mordor Intelligence"
        ws.cell(row, 5).value = "HIGH"
        row += 1

        # Mechanical Engineering Subset
        mech_row = row
        ws.cell(row, 1).value = "  Mechanical Engineering Subset"
        ws.cell(row, 1).font = Font(size=10)
        ws.cell(row, 2).value = 50000
        ws.cell(row, 2).number_format = "[$-409]#,##0"
        ws.cell(row, 3).value = "Mechanical engineering outsourcing"
        ws.cell(row, 4).value = "[33] Market Report Analytics"
        ws.cell(row, 5).value = "HIGH"
        row += 1

        # Tooling Design %
        tooling_design_pct_row = row
        ws.cell(row, 1).value = "  Tooling Design % of Mech Eng"
        ws.cell(row, 1).font = Font(size=10)
        ws.cell(row, 2).value = 0.40
        ws.cell(row, 2).number_format = "0%"
        ws.cell(row, 3).value = "Jigs/fixtures/molds/dies design work"
        ws.cell(row, 4).value = "Industry Benchmark"
        ws.cell(row, 5).value = "MEDIUM"
        row += 1

        # Total Services TAM - FORMULA
        svc_total_row = row
        ws.cell(row, 1).value = "  Total Services TAM"
        ws.cell(row, 1).font = Font(bold=True, size=10)
        ws.cell(row, 2).value = f"=B{mech_row}*B{tooling_design_pct_row}"
        ws.cell(row, 2).number_format = "[$-409]#,##0"
        ws.cell(row, 3).value = "Mech Eng x Tooling %"
        ws.cell(row, 4).value = "[32][33][34]"
        ws.cell(row, 5).value = "MEDIUM-HIGH"
        style_data_row(ws, row, 1, 5, is_total=True)
        row += 2

        # TOTAL TAM row - FORMULA summing all subtotals
        ws.cell(row, 1).value = "TOTAL TAM"
        ws.cell(row, 1).font = Font(bold=True, size=12)
        ws.cell(row, 2).value = (
            f"=B{sw_total_row}+B{hw_total_row}+B{cons_total_row}+B{svc_total_row}"
        )
        ws.cell(row, 2).number_format = "[$-409]#,##0"
        ws.cell(row, 2).font = Font(bold=True, size=12)
        ws.cell(row, 3).value = f"Software + Hardware + Consumables + Services"
        ws.cell(row, 5).value = "HIGH"
        style_data_row(ws, row, 1, 5, is_total=True)
        self.row_refs["tam_total"] = row
        row += 3

        # ----------------------------------------------------------------
        # SAM - SERVICEABLE ADDRESSABLE MARKET (Phased)
        # ----------------------------------------------------------------
        style_section_header(
            ws.cell(row, 1),
            "SAM - SERVICEABLE ADDRESSABLE MARKET (Phased Geographic Expansion)",
            bg=Colors.SECTION_A_CAT,
        )
        ws.merge_cells(f"A{row}:J{row}")
        row += 1

        sam_headers = [
            "Phase / Region",
            "SAM ($M)",
            "Years",
            "Calculation",
            "Source [Ref#]",
        ]
        style_column_headers(ws, row, sam_headers, bg=Colors.MEDIUM_BLUE)
        row += 1

        self.row_refs["sam_start"] = row

        # PHASE 1: INDIA
        style_category_header(ws, row, "PHASE 1: INDIA BEACHHEAD")
        row += 1

        india_sw_row = row
        ws.cell(row, 1).value = "  Software SAM (India)"
        ws.cell(row, 1).font = Font(size=10)
        ws.cell(row, 2).value = 206
        ws.cell(row, 2).number_format = "[$-409]#,##0"
        ws.cell(row, 3).value = "Y1-Y2"
        ws.cell(row, 4).value = "India CAD $620M x 40.9% x 50%"
        ws.cell(row, 5).value = "[3] Mordor Intelligence"
        row += 1

        india_hw_row = row
        ws.cell(row, 1).value = "  Hardware SAM (India)"
        ws.cell(row, 1).font = Font(size=10)
        ws.cell(row, 2).value = 88
        ws.cell(row, 2).number_format = "[$-409]#,##0"
        ws.cell(row, 3).value = "Y1-Y2"
        ws.cell(row, 4).value = "India 3DP $1.24B x 24% x 22%"
        ws.cell(row, 5).value = "[7] GM Insights"
        row += 1

        india_cons_row = row
        ws.cell(row, 1).value = "  Consumables SAM (India)"
        ws.cell(row, 1).font = Font(size=10)
        ws.cell(row, 2).value = 170
        ws.cell(row, 2).number_format = "[$-409]#,##0"
        ws.cell(row, 3).value = "Y1-Y2"
        ws.cell(row, 4).value = "3DP materials + traditional"
        ws.cell(row, 5).value = "[8] Grand View Research"
        row += 1

        india_svc_row = row
        ws.cell(row, 1).value = "  Services SAM (India)"
        ws.cell(row, 1).font = Font(size=10)
        ws.cell(row, 2).value = 1336
        ws.cell(row, 2).number_format = "[$-409]#,##0"
        ws.cell(row, 3).value = "Y1-Y2"
        ws.cell(row, 4).value = "500K companies x tier-weighted spend"
        ws.cell(row, 5).value = "Bottoms-up"
        row += 1

        # Phase 1 Total - FORMULA
        phase1_total_row = row
        ws.cell(row, 1).value = "  Phase 1 Total SAM"
        ws.cell(row, 1).font = Font(bold=True, size=10)
        ws.cell(row, 2).value = f"=SUM(B{india_sw_row}:B{india_svc_row})"
        ws.cell(row, 2).number_format = "[$-409]#,##0"
        ws.cell(row, 3).value = "Y1-Y2"
        ws.cell(row, 4).value = "Sum of India streams"
        ws.cell(row, 5).value = "India-only"
        style_data_row(ws, row, 1, 5, is_total=True)
        row += 2

        # PHASE 2: SE ASIA
        style_category_header(ws, row, "PHASE 2: + SOUTHEAST ASIA")
        row += 1

        sea_sw_row = row
        ws.cell(row, 1).value = "  SE Asia Software SAM"
        ws.cell(row, 1).font = Font(size=10)
        ws.cell(row, 2).value = 130
        ws.cell(row, 2).number_format = "[$-409]#,##0"
        ws.cell(row, 3).value = "Y3-Y4"
        ws.cell(row, 4).value = "SE Asia CAD $390M x 40.9% x 50%"
        ws.cell(row, 5).value = "[14] ASEAN Statistics"
        row += 1

        sea_hw_row = row
        ws.cell(row, 1).value = "  SE Asia Hardware SAM"
        ws.cell(row, 1).font = Font(size=10)
        ws.cell(row, 2).value = 90
        ws.cell(row, 2).number_format = "[$-409]#,##0"
        ws.cell(row, 3).value = "Y3-Y4"
        ws.cell(row, 4).value = "SE Asia 3DP $420M x 24% x 22%"
        ws.cell(row, 5).value = "Estimated"
        row += 1

        sea_cons_row = row
        ws.cell(row, 1).value = "  SE Asia Consumables SAM"
        ws.cell(row, 1).font = Font(size=10)
        ws.cell(row, 2).value = 156
        ws.cell(row, 2).number_format = "[$-409]#,##0"
        ws.cell(row, 3).value = "Y3-Y4"
        ws.cell(row, 4).value = "3DP materials + traditional"
        ws.cell(row, 5).value = "Estimated"
        row += 1

        sea_svc_row = row
        ws.cell(row, 1).value = "  SE Asia Services SAM"
        ws.cell(row, 1).font = Font(size=10)
        ws.cell(row, 2).value = 704
        ws.cell(row, 2).number_format = "[$-409]#,##0"
        ws.cell(row, 3).value = "Y3-Y4"
        ws.cell(row, 4).value = "150K companies x tier-weighted spend"
        ws.cell(row, 5).value = "Bottoms-up"
        row += 1

        # Phase 2 Incremental - FORMULA
        phase2_incr_row = row
        ws.cell(row, 1).value = "  Phase 2 Incremental SAM"
        ws.cell(row, 1).font = Font(bold=True, size=10)
        ws.cell(row, 2).value = f"=SUM(B{sea_sw_row}:B{sea_svc_row})"
        ws.cell(row, 2).number_format = "[$-409]#,##0"
        ws.cell(row, 3).value = "Y3-Y4"
        ws.cell(row, 4).value = "Sum of SE Asia streams"
        ws.cell(row, 5).value = "SE Asia addition"
        style_data_row(ws, row, 1, 5, is_total=True)
        row += 1

        # Phase 2 Cumulative - FORMULA
        phase2_cum_row = row
        ws.cell(row, 1).value = "  Phase 2 Cumulative SAM"
        ws.cell(row, 1).font = Font(bold=True, size=10)
        ws.cell(row, 2).value = f"=B{phase1_total_row}+B{phase2_incr_row}"
        ws.cell(row, 2).number_format = "[$-409]#,##0"
        ws.cell(row, 3).value = "Y3-Y4"
        ws.cell(row, 4).value = "Phase 1 + Phase 2 Incr"
        ws.cell(row, 5).value = "India + SE Asia"
        style_data_row(ws, row, 1, 5, is_total=True)
        row += 2

        # PHASE 3: GLOBAL
        style_category_header(ws, row, "PHASE 3: + GLOBAL VAR NETWORK")
        row += 1

        eu_row = row
        ws.cell(row, 1).value = "  EU Markets (Germany, Italy, etc)"
        ws.cell(row, 1).font = Font(size=10)
        ws.cell(row, 2).value = 1440
        ws.cell(row, 2).number_format = "[$-409]#,##0"
        ws.cell(row, 3).value = "Y5-Y8"
        ws.cell(row, 4).value = "Germany $800M + Italy $400M + Other $240M"
        ws.cell(row, 5).value = "VAR channel"
        row += 1

        americas_row = row
        ws.cell(row, 1).value = "  Americas (US, Mexico, etc)"
        ws.cell(row, 1).font = Font(size=10)
        ws.cell(row, 2).value = 1440
        ws.cell(row, 2).number_format = "[$-409]#,##0"
        ws.cell(row, 3).value = "Y5-Y8"
        ws.cell(row, 4).value = "US $1B + Mexico $300M + Other $140M"
        ws.cell(row, 5).value = "VAR channel"
        row += 1

        # Phase 3 Incremental - FORMULA
        phase3_incr_row = row
        ws.cell(row, 1).value = "  Phase 3 Incremental SAM"
        ws.cell(row, 1).font = Font(bold=True, size=10)
        ws.cell(row, 2).value = f"=B{eu_row}+B{americas_row}"
        ws.cell(row, 2).number_format = "[$-409]#,##0"
        ws.cell(row, 3).value = "Y5-Y8"
        ws.cell(row, 4).value = "EU + Americas"
        ws.cell(row, 5).value = "EU + Americas"
        style_data_row(ws, row, 1, 5, is_total=True)
        row += 1

        # Phase 3 Total SAM (Full Platform) - FORMULA
        phase3_total_row = row
        ws.cell(row, 1).value = "  Phase 3 Total SAM (Full Platform)"
        ws.cell(row, 1).font = Font(bold=True, size=10)
        ws.cell(row, 2).value = f"=B{phase2_cum_row}+B{phase3_incr_row}"
        ws.cell(row, 2).number_format = "[$-409]#,##0"
        ws.cell(row, 3).value = "Y5-Y8"
        ws.cell(row, 4).value = "Phase 2 Cum + Phase 3 Incr"
        ws.cell(row, 5).value = "Global SAM"
        style_data_row(ws, row, 1, 5, is_total=True)
        self.row_refs["sam_total"] = row
        row += 3

        # ----------------------------------------------------------------
        # SOM - YEAR-WISE PROJECTIONS (KEY FOR DOWNSTREAM CALCULATIONS)
        # ----------------------------------------------------------------
        style_section_header(
            ws.cell(row, 1),
            "SOM - SERVICEABLE OBTAINABLE MARKET (Year-by-Year Projections)",
            bg=Colors.SECTION_A_CAT,
        )
        ws.merge_cells(f"A{row}:J{row}")
        row += 1

        # Year-wise headers
        som_year_headers = ["Metric", "Y1", "Y2", "Y3", "Y4", "Y5", "Y6", "Y7", "Y8"]
        style_column_headers(ws, row, som_year_headers, bg=Colors.MEDIUM_BLUE)
        row += 1

        self.row_refs["som_start"] = row

        # Geography row (text, not numbers)
        ws.cell(row, 1).value = "Geography"
        for col, geo in enumerate(
            [
                "India",
                "India",
                "+SE Asia",
                "+SE Asia",
                "+Global",
                "+Global",
                "+Global",
                "Global",
            ],
            2,
        ):
            ws.cell(row, col).value = geo
        row += 1

        # Active SAM row - store for penetration calculation
        sam_row = row
        ws.cell(row, 1).value = "Active SAM ($M)"
        for col, sam in enumerate([1800, 1800, 2880, 2880, 4320, 4800, 5280, 5760], 2):
            ws.cell(row, col).value = sam
            ws.cell(row, col).number_format = "[$-409]#,##0"
        row += 1

        # Revenue by Stream section header
        style_category_header(ws, row, "REVENUE BY STREAM ($M)")
        row += 1

        # Individual revenue streams
        sw_rev_row = row
        ws.cell(row, 1).value = "  Software Revenue"
        ws.cell(row, 1).font = Font(size=10)
        for col, val in enumerate([0.2, 1.3, 3.9, 9.0, 18.6, 31.0, 47.7, 64.0], 2):
            ws.cell(row, col).value = val
            ws.cell(row, col).number_format = "[$-409]#,##0.0"
        row += 1

        tooling_svc_row = row
        ws.cell(row, 1).value = "  Services - Tooling"
        ws.cell(row, 1).font = Font(size=10)
        for col, val in enumerate([0.2, 0.8, 2.1, 4.0, 7.4, 12.4, 18.7, 24.2], 2):
            ws.cell(row, col).value = val
            ws.cell(row, col).number_format = "[$-409]#,##0.0"
        row += 1

        ems_svc_row = row
        ws.cell(row, 1).value = "  Services - EMS"
        ws.cell(row, 1).font = Font(size=10)
        for col, val in enumerate([0.1, 0.3, 0.7, 1.5, 3.0, 5.0, 6.5, 7.4], 2):
            ws.cell(row, col).value = val
            ws.cell(row, col).number_format = "[$-409]#,##0.0"
        row += 1

        hw_rev_row = row
        ws.cell(row, 1).value = "  Hardware Revenue"
        ws.cell(row, 1).font = Font(size=10)
        for col, val in enumerate([0.0, 0.1, 0.3, 0.5, 1.0, 1.6, 4.1, 7.4], 2):
            ws.cell(row, col).value = val
            ws.cell(row, col).number_format = "[$-409]#,##0.0"
        row += 1

        cons_rev_row = row
        ws.cell(row, 1).value = "  Consumables Revenue"
        ws.cell(row, 1).font = Font(size=10)
        for col, val in enumerate([0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 0.0, 1.0], 2):
            ws.cell(row, col).value = val
            ws.cell(row, col).number_format = "[$-409]#,##0.0"
        row += 1

        # TOTAL REVENUE - FORMULA summing all streams
        total_rev_row = row
        ws.cell(row, 1).value = "TOTAL REVENUE ($M)"
        ws.cell(row, 1).font = Font(bold=True)
        for col in range(2, 10):
            col_letter = get_column_letter(col)
            ws.cell(row, col).value = (
                f"=SUM({col_letter}{sw_rev_row}:{col_letter}{cons_rev_row})"
            )
            ws.cell(row, col).number_format = "[$-409]#,##0.0"
        style_data_row(ws, row, 1, 9, is_total=True)
        row += 2

        # Customers row
        cust_row = row
        ws.cell(row, 1).value = "Customers"
        ws.cell(row, 1).font = Font(bold=True)
        for col, cust in enumerate([8, 30, 85, 220, 550, 1350, 2900, 5500], 2):
            ws.cell(row, col).value = cust
            ws.cell(row, col).number_format = "[$-409]#,##0"
        row += 1

        # Avg ARPU - FORMULA (Revenue / Customers)
        arpu_row = row
        ws.cell(row, 1).value = "Avg ARPU ($)"
        for col in range(2, 10):
            col_letter = get_column_letter(col)
            ws.cell(row, col).value = (
                f"=IF({col_letter}{cust_row}>0,({col_letter}{total_rev_row}*1000000)/{col_letter}{cust_row},0)"
            )
            ws.cell(row, col).number_format = "$#,##0"
        row += 1

        # YoY Growth - FORMULA
        yoy_row = row
        ws.cell(row, 1).value = "YoY Growth (%)"
        ws.cell(row, 2).value = ""  # Y1 has no prior year
        for col in range(3, 10):
            col_letter = get_column_letter(col)
            prev_col = get_column_letter(col - 1)
            ws.cell(row, col).value = (
                f"=IF({prev_col}{total_rev_row}>0,({col_letter}{total_rev_row}-{prev_col}{total_rev_row})/{prev_col}{total_rev_row},0)"
            )
            ws.cell(row, col).number_format = "0%"
        row += 1

        # SAM Penetration - FORMULA (Revenue / SAM)
        pen_row = row
        ws.cell(row, 1).value = "SAM Penetration (%)"
        ws.cell(row, 1).font = Font(bold=True)
        for col in range(2, 10):
            col_letter = get_column_letter(col)
            ws.cell(row, col).value = (
                f"=IF({col_letter}{sam_row}>0,{col_letter}{total_rev_row}/{col_letter}{sam_row},0)"
            )
            ws.cell(row, col).number_format = "0.00%"
        row += 1

        self.row_refs["som_target"] = total_rev_row
        row += 3

        # ----------------------------------------------------------------
        # KEY BENCHMARKS & UNIT ECONOMICS
        # ----------------------------------------------------------------
        style_section_header(
            ws.cell(row, 1), "KEY BENCHMARKS & UNIT ECONOMICS", bg=Colors.SECTION_A_CAT
        )
        ws.merge_cells(f"A{row}:J{row}")
        row += 1

        bench_headers = ["Metric", "Value", "Unit", "Source [Ref#]", "Notes"]
        style_column_headers(ws, row, bench_headers, bg=Colors.MEDIUM_BLUE)
        row += 1

        # Customer Acquisition section
        style_category_header(ws, row, "CUSTOMER ACQUISITION")
        row += 1

        plg_cac_row = row
        ws.cell(row, 1).value = "PLG/Inbound CAC"
        ws.cell(row, 2).value = 2500
        ws.cell(row, 2).number_format = "[$-409]#,##0"
        ws.cell(row, 3).value = "USD"
        ws.cell(row, 4).value = "[22] SaaS Capital"
        ws.cell(row, 5).value = "70% of customers"
        row += 1

        var_cac_row = row
        ws.cell(row, 1).value = "VAR Channel CAC"
        ws.cell(row, 2).value = 12000
        ws.cell(row, 2).number_format = "[$-409]#,##0"
        ws.cell(row, 3).value = "USD"
        ws.cell(row, 4).value = "Industry benchmark"
        ws.cell(row, 5).value = "25% of customers"
        row += 1

        direct_cac_row = row
        ws.cell(row, 1).value = "Direct Sales CAC"
        ws.cell(row, 2).value = 45000
        ws.cell(row, 2).number_format = "[$-409]#,##0"
        ws.cell(row, 3).value = "USD"
        ws.cell(row, 4).value = "Industry benchmark"
        ws.cell(row, 5).value = "5% of customers"
        row += 1

        # Blended CAC - FORMULA (weighted average)
        ws.cell(row, 1).value = "Blended CAC"
        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 2).value = (
            f"=(B{plg_cac_row}*0.70)+(B{var_cac_row}*0.25)+(B{direct_cac_row}*0.05)"
        )
        ws.cell(row, 2).number_format = "[$-409]#,##0"
        ws.cell(row, 3).value = "USD"
        ws.cell(row, 4).value = "Channel-weighted avg"
        ws.cell(row, 5).value = "70% PLG + 25% VAR + 5% Direct"
        style_data_row(ws, row, 1, 5, is_total=True)
        blended_cac_row = row
        row += 2

        # Lifetime Value section
        style_category_header(ws, row, "LIFETIME VALUE (LTV)")
        row += 1

        # Churn rates for LTV calculation
        smb_churn_row = row
        ws.cell(row, 1).value = "SMB Churn Rate"
        ws.cell(row, 2).value = 0.17
        ws.cell(row, 2).number_format = "0%"
        ws.cell(row, 3).value = "%"
        ws.cell(row, 4).value = "[23] ChartMogul"
        ws.cell(row, 5).value = "B2B SaaS benchmark"
        row += 1

        mm_churn_row = row
        ws.cell(row, 1).value = "Mid-Market Churn Rate"
        ws.cell(row, 2).value = 0.11
        ws.cell(row, 2).number_format = "0%"
        ws.cell(row, 3).value = "%"
        ws.cell(row, 4).value = "[23] ChartMogul"
        ws.cell(row, 5).value = "B2B SaaS benchmark"
        row += 1

        ent_churn_row = row
        ws.cell(row, 1).value = "Enterprise Churn Rate"
        ws.cell(row, 2).value = 0.06
        ws.cell(row, 2).number_format = "0%"
        ws.cell(row, 3).value = "%"
        ws.cell(row, 4).value = "[23] ChartMogul"
        ws.cell(row, 5).value = "B2B SaaS benchmark"
        row += 1

        # ARPU by segment
        smb_arpu_row = row
        ws.cell(row, 1).value = "SMB ARPU"
        ws.cell(row, 2).value = 6500
        ws.cell(row, 2).number_format = "[$-409]#,##0"
        ws.cell(row, 3).value = "USD/yr"
        ws.cell(row, 4).value = "Pricing model"
        row += 1

        mm_arpu_row = row
        ws.cell(row, 1).value = "Mid-Market ARPU"
        ws.cell(row, 2).value = 24000
        ws.cell(row, 2).number_format = "[$-409]#,##0"
        ws.cell(row, 3).value = "USD/yr"
        ws.cell(row, 4).value = "Pricing model"
        row += 1

        ent_arpu_row = row
        ws.cell(row, 1).value = "Enterprise ARPU"
        ws.cell(row, 2).value = 37000
        ws.cell(row, 2).number_format = "[$-409]#,##0"
        ws.cell(row, 3).value = "USD/yr"
        ws.cell(row, 4).value = "Pricing model"
        row += 1

        # LTV calculations - FORMULA (ARPU / Churn = Lifetime * ARPU)
        smb_ltv_row = row
        ws.cell(row, 1).value = "SMB LTV"
        ws.cell(row, 2).value = f"=B{smb_arpu_row}/B{smb_churn_row}"
        ws.cell(row, 2).number_format = "[$-409]#,##0"
        ws.cell(row, 3).value = "USD"
        ws.cell(row, 4).value = "[23] ChartMogul"
        ws.cell(row, 5).value = "5.9 yr lifetime"
        row += 1

        mm_ltv_row = row
        ws.cell(row, 1).value = "Mid-Market LTV"
        ws.cell(row, 2).value = f"=B{mm_arpu_row}/B{mm_churn_row}"
        ws.cell(row, 2).number_format = "[$-409]#,##0"
        ws.cell(row, 3).value = "USD"
        ws.cell(row, 4).value = "[23] ChartMogul"
        ws.cell(row, 5).value = "9.1 yr lifetime"
        row += 1

        ent_ltv_row = row
        ws.cell(row, 1).value = "Enterprise LTV"
        ws.cell(row, 2).value = f"=B{ent_arpu_row}/B{ent_churn_row}"
        ws.cell(row, 2).number_format = "[$-409]#,##0"
        ws.cell(row, 3).value = "USD"
        ws.cell(row, 4).value = "[23] ChartMogul"
        ws.cell(row, 5).value = "16.7 yr lifetime"
        row += 1

        # Blended LTV - FORMULA (weighted by segment mix: 60% SMB, 30% MM, 10% Ent)
        ws.cell(row, 1).value = "Blended LTV"
        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 2).value = (
            f"=(B{smb_ltv_row}*0.60)+(B{mm_ltv_row}*0.30)+(B{ent_ltv_row}*0.10)"
        )
        ws.cell(row, 2).number_format = "[$-409]#,##0"
        ws.cell(row, 3).value = "USD"
        ws.cell(row, 4).value = "Segment-weighted"
        ws.cell(row, 5).value = "60% SMB + 30% MM + 10% Ent"
        style_data_row(ws, row, 1, 5, is_total=True)
        blended_ltv_row = row
        row += 2

        # LTV:CAC Ratios section
        style_category_header(ws, row, "LTV:CAC RATIOS")
        row += 1

        # LTV:CAC - FORMULAS
        ws.cell(row, 1).value = "SMB LTV:CAC"
        ws.cell(row, 2).value = f"=B{smb_ltv_row}/B{plg_cac_row}"
        ws.cell(row, 2).number_format = "0.0"
        ws.cell(row, 3).value = "x"
        ws.cell(row, 4).value = "LTV / PLG CAC"
        ws.cell(row, 5).value = "Excellent (>3x target)"
        row += 1

        ws.cell(row, 1).value = "Mid-Market LTV:CAC"
        ws.cell(row, 2).value = f"=B{mm_ltv_row}/B{var_cac_row}"
        ws.cell(row, 2).number_format = "0.0"
        ws.cell(row, 3).value = "x"
        ws.cell(row, 4).value = "LTV / VAR CAC"
        ws.cell(row, 5).value = "Excellent"
        row += 1

        ws.cell(row, 1).value = "Enterprise LTV:CAC"
        ws.cell(row, 2).value = f"=B{ent_ltv_row}/B{direct_cac_row}"
        ws.cell(row, 2).number_format = "0.0"
        ws.cell(row, 3).value = "x"
        ws.cell(row, 4).value = "LTV / Direct CAC"
        ws.cell(row, 5).value = "Strong"
        row += 1

        ws.cell(row, 1).value = "Blended LTV:CAC"
        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 2).value = f"=B{blended_ltv_row}/B{blended_cac_row}"
        ws.cell(row, 2).number_format = "0.0"
        ws.cell(row, 3).value = "x"
        ws.cell(row, 4).value = "Blended LTV / Blended CAC"
        ws.cell(row, 5).value = "Very healthy (target >3x)"
        style_data_row(ws, row, 1, 5, is_total=True)
        row += 2

        # Gross Margins section
        style_category_header(ws, row, "GROSS MARGINS")
        row += 1

        sw_gm_row = row
        ws.cell(row, 1).value = "Software Gross Margin"
        ws.cell(row, 2).value = 0.80
        ws.cell(row, 2).number_format = "0%"
        ws.cell(row, 3).value = "%"
        ws.cell(row, 4).value = "[22] SaaS Capital"
        ws.cell(row, 5).value = "SaaS industry standard"
        row += 1

        svc_gm_row = row
        ws.cell(row, 1).value = "Services Gross Margin"
        ws.cell(row, 2).value = 0.45
        ws.cell(row, 2).number_format = "0%"
        ws.cell(row, 3).value = "%"
        ws.cell(row, 4).value = "Industry benchmark"
        ws.cell(row, 5).value = "Professional services"
        row += 1

        hw_gm_row = row
        ws.cell(row, 1).value = "Hardware Gross Margin"
        ws.cell(row, 2).value = 0.35
        ws.cell(row, 2).number_format = "0%"
        ws.cell(row, 3).value = "%"
        ws.cell(row, 4).value = "Industry benchmark"
        ws.cell(row, 5).value = "Equipment resale"
        row += 1

        cons_gm_row = row
        ws.cell(row, 1).value = "Consumables Gross Margin"
        ws.cell(row, 2).value = 0.50
        ws.cell(row, 2).number_format = "0%"
        ws.cell(row, 3).value = "%"
        ws.cell(row, 4).value = "Industry benchmark"
        ws.cell(row, 5).value = "Materials markup"
        row += 1

        # Blended Gross Margin - FORMULA (revenue-weighted)
        ws.cell(row, 1).value = "Blended Gross Margin (Y8)"
        ws.cell(row, 1).font = Font(bold=True)
        # Y8 mix: Software 62%, Services 30%, Hardware 7%, Consumables 1%
        ws.cell(row, 2).value = (
            f"=(B{sw_gm_row}*0.62)+(B{svc_gm_row}*0.30)+(B{hw_gm_row}*0.07)+(B{cons_gm_row}*0.01)"
        )
        ws.cell(row, 2).number_format = "0%"
        ws.cell(row, 3).value = "%"
        ws.cell(row, 4).value = "Revenue-weighted"
        ws.cell(row, 5).value = "Y8 mix: 62% SW, 30% Svc, 7% HW, 1% Cons"
        style_data_row(ws, row, 1, 5, is_total=True)
        row += 3

        # ----------------------------------------------------------------
        # INDUSTRY GROWTH RATES (CAGR)
        # ----------------------------------------------------------------
        style_section_header(
            ws.cell(row, 1),
            "INDUSTRY GROWTH RATES (CAGR 2024-2030)",
            bg=Colors.SECTION_A_CAT,
        )
        ws.merge_cells(f"A{row}:J{row}")
        row += 1

        cagr_headers = ["Market Segment", "CAGR", "Period", "Source [Ref#]", "Notes"]
        style_column_headers(ws, row, cagr_headers, bg=Colors.MEDIUM_BLUE)
        row += 1

        cagr_data = [
            (
                "CAD Software Market",
                0.062,
                "2024-2030",
                "[1] Future Market Insights",
                "Fastest growing segment",
            ),
            (
                "CAM Software Market",
                0.058,
                "2024-2030",
                "[2] Mordor Intelligence",
                "Manufacturing digitalization",
            ),
            (
                "Industrial 3D Printing",
                0.106,
                "2024-2030",
                "[6] GM Insights",
                "Rapid tooling adoption",
            ),
            (
                "3D Printing Materials",
                0.135,
                "2024-2030",
                "[8] Grand View Research",
                "Polymer + metal materials",
            ),
            (
                "Engineering Services Outsourcing",
                0.089,
                "2024-2030",
                "[32] Mordor Intelligence",
                "Global talent arbitrage",
            ),
            (
                "Industrial Robotics",
                0.12,
                "2023-2030",
                "[21] IFR World Robotics",
                "Automation demand driver",
            ),
        ]

        for name, cagr, period, source, notes in cagr_data:
            ws.cell(row, 1).value = name
            ws.cell(row, 1).font = Font(size=10)
            ws.cell(row, 2).value = cagr
            ws.cell(row, 2).number_format = "0.0%"
            ws.cell(row, 3).value = period
            ws.cell(row, 4).value = source
            ws.cell(row, 5).value = notes
            row += 1

        row += 2

        # ----------------------------------------------------------------
        # GOVERNMENT POLICY INVESTMENT (Tailwinds)
        # ----------------------------------------------------------------
        style_section_header(
            ws.cell(row, 1),
            "GOVERNMENT POLICY INVESTMENT (Key Tailwinds)",
            bg=Colors.SECTION_A_CAT,
        )
        ws.merge_cells(f"A{row}:J{row}")
        row += 1

        policy_headers = [
            "Policy/Initiative",
            "Investment ($B)",
            "Region",
            "Sector",
            "Source [Ref#]",
        ]
        style_column_headers(ws, row, policy_headers, bg=Colors.MEDIUM_BLUE)
        row += 1

        policy_data = [
            ("U.S. CHIPS Act", 52.7, "USA", "Semiconductor Manufacturing", "[40]"),
            ("U.S. Inflation Reduction Act", 369, "USA", "Clean Energy/EV", "[41]"),
            ("India PLI Scheme", 26, "India", "14 Manufacturing Sectors", "[42]"),
            ("India Semiconductor Mission", 10, "India", "Fab + ATMP", "[43]"),
            ("EU Chips Act", 43, "EU", "Semiconductor", "[44]"),
            ("EU Green Deal", 1000, "EU", "Clean Energy Transition", "[45]"),
            ("Germany Industry 4.0", 20, "Germany", "Smart Manufacturing", "[46]"),
        ]

        for name, investment, region, sector, source in policy_data:
            ws.cell(row, 1).value = name
            ws.cell(row, 1).font = Font(size=10)
            ws.cell(row, 2).value = investment
            ws.cell(row, 2).number_format = "[$-409]#,##0.0"
            ws.cell(row, 3).value = region
            ws.cell(row, 4).value = sector
            ws.cell(row, 5).value = source
            row += 1

        # Total Policy Investment - FORMULA
        ws.cell(row, 1).value = "TOTAL POLICY INVESTMENT"
        ws.cell(row, 1).font = Font(bold=True)
        policy_start = row - len(policy_data)
        ws.cell(row, 2).value = f"=SUM(B{policy_start}:B{row-1})"
        ws.cell(row, 2).number_format = "[$-409]#,##0.0"
        ws.cell(row, 3).value = "Global"
        ws.cell(row, 4).value = "Manufacturing Tailwinds"
        style_data_row(ws, row, 1, 5, is_total=True)
        row += 3

        # ----------------------------------------------------------------
        # COMPETITOR BENCHMARKS
        # ----------------------------------------------------------------
        style_section_header(
            ws.cell(row, 1), "COMPETITOR BENCHMARKS", bg=Colors.SECTION_A_CAT
        )
        ws.merge_cells(f"A{row}:J{row}")
        row += 1

        comp_headers = ["Company", "Revenue ($M)", "Segment", "Key Metric", "Source"]
        style_column_headers(ws, row, comp_headers, bg=Colors.MEDIUM_BLUE)
        row += 1

        comp_data = [
            (
                "Trinckle (Paramate)",
                15,
                "Generative Design",
                "BMW partnership",
                "Crunchbase",
            ),
            ("Zoo Corp (KittyCAD)", 8, "API-first CAD", "$25M Series A", "Crunchbase"),
            ("Autodesk", 5500, "Incumbent CAD", "80% gross margin", "10-K Filing"),
            (
                "Dassault Systèmes",
                6200,
                "Incumbent PLM",
                "85% gross margin",
                "Annual Report",
            ),
            ("PTC", 2100, "Incumbent CAD/PLM", "SaaS transition", "10-K Filing"),
            ("Stratasys", 630, "Industrial 3DP", "45% gross margin", "10-K Filing"),
            ("3D Systems", 490, "Industrial 3DP", "40% gross margin", "10-K Filing"),
        ]

        for name, revenue, segment, metric, source in comp_data:
            ws.cell(row, 1).value = name
            ws.cell(row, 1).font = Font(size=10)
            ws.cell(row, 2).value = revenue
            ws.cell(row, 2).number_format = "[$-409]#,##0"
            ws.cell(row, 3).value = segment
            ws.cell(row, 4).value = metric
            ws.cell(row, 5).value = source
            row += 1

        row += 2

        # ================================================================
        # SECTION B: FULL SOURCE DOCUMENTATION WITH URLS
        # ================================================================
        style_section_header(ws.cell(row, 1), "SECTION B: FULL SOURCE DOCUMENTATION")
        ws.merge_cells(f"A{row}:J{row}")
        row += 2

        source_headers = ["Ref#", "Source Name", "Data Point", "Value", "URL"]
        style_column_headers(ws, row, source_headers, bg=Colors.MEDIUM_BLUE)
        row += 1

        # Full source citations with URLs
        sources = [
            ("", "MARKET RESEARCH - TAM SOURCES", "", "", ""),
            (
                "[1]",
                "Future Market Insights",
                "Global CAD Software Market",
                "$12.2B (2025)",
                "https://www.futuremarketinsights.com/reports/computer-aided-design-cad-market",
            ),
            (
                "[2]",
                "Mordor Intelligence",
                "Global CAM Software Market",
                "$3.45B (2025)",
                "https://www.mordorintelligence.com/industry-reports/computer-aided-manufacturing-market",
            ),
            (
                "[3]",
                "Mordor Intelligence",
                "India CAD Market",
                "$620M (2025)",
                "https://www.mordorintelligence.com/industry-reports/india-cad-market",
            ),
            (
                "[6]",
                "GM Insights",
                "Industrial 3D Printing Market",
                "$18.3B (2025)",
                "https://www.gminsights.com/industry-analysis/industrial-3d-printing-market",
            ),
            (
                "[7]",
                "GM Insights",
                "India 3D Printing Market",
                "$1.24B (2025)",
                "https://www.gminsights.com/industry-analysis/india-3d-printing-market",
            ),
            (
                "[8]",
                "Grand View Research",
                "3D Printing Materials Market",
                "$3.88B (2025)",
                "https://www.grandviewresearch.com/industry-analysis/3d-printing-materials-market",
            ),
            (
                "[9]",
                "MarketsandMarkets",
                "CAM Software Market",
                "$3.40B (2024)",
                "https://www.marketsandmarkets.com/Market-Reports/computer-aided-manufacturing-market-251259446.html",
            ),
            ("", "", "", "", ""),
            ("", "INDUSTRY GROWTH RATES (CAGR)", "", "", ""),
            (
                "[1]",
                "CAD Software Market CAGR",
                "2024-2030",
                "6.2%",
                "Future Market Insights",
            ),
            (
                "[2]",
                "CAM Software Market CAGR",
                "2024-2030",
                "7.1%",
                "Mordor Intelligence",
            ),
            ("[6]", "Industrial 3D Printing CAGR", "2024-2030", "18.5%", "GM Insights"),
            (
                "[8]",
                "3D Printing Materials CAGR",
                "2024-2030",
                "21.2%",
                "Grand View Research",
            ),
            (
                "[32]",
                "Engineering Services CAGR",
                "2024-2030",
                "12.8%",
                "Mordor Intelligence",
            ),
            (
                "[21]",
                "Robot Installations CAGR",
                "2023-2030",
                "28.5%",
                "IFR World Robotics",
            ),
            ("", "", "", "", ""),
            ("", "GOVERNMENT POLICY INVESTMENT (WHY NOW)", "", "", ""),
            (
                "[40]",
                "U.S. CHIPS Act",
                "Semiconductor Manufacturing",
                "$52B (2022-2027)",
                "https://www.commerce.gov/chips",
            ),
            (
                "[41]",
                "U.S. Inflation Reduction Act",
                "Clean Energy Manufacturing",
                "$369B (2022-2032)",
                "https://www.whitehouse.gov/cleanenergy/inflation-reduction-act-guidebook/",
            ),
            (
                "[42]",
                "India PLI Scheme",
                "14 Sectors Incentives",
                "₹1.97L Cr / $24B",
                "https://www.investindia.gov.in/production-linked-incentives-schemes",
            ),
            (
                "[43]",
                "India Semiconductor Mission",
                "Fab + ATMP Investment",
                "₹76K Cr / $10B",
                "https://www.meity.gov.in/esdm/semiconductors",
            ),
            (
                "[44]",
                "EU Chips Act",
                "European Semiconductor",
                "€43B (2023-2030)",
                "https://digital-strategy.ec.europa.eu/en/policies/european-chips-act",
            ),
            (
                "[45]",
                "EU Green Deal",
                "Clean Energy Transition",
                "€1T (2020-2050)",
                "https://commission.europa.eu/strategy-and-policy/priorities-2019-2024/european-green-deal_en",
            ),
            (
                "[46]",
                "Germany Industry 4.0",
                "Automation Investment",
                "€40B (2020-2030)",
                "https://www.plattform-i40.de/",
            ),
            ("", "", "", "", ""),
            ("", "ROBOTICS DEMAND MULTIPLIER", "", "", ""),
            (
                "[21]",
                "IFR World Robotics",
                "Global Robot Stock 2023",
                "3.9M units",
                "https://ifr.org/worldrobotics/",
            ),
            (
                "[21]",
                "IFR World Robotics",
                "Annual Installations 2023",
                "550K units/year",
                "https://ifr.org/worldrobotics/",
            ),
            (
                "[21]",
                "IFR World Robotics",
                "Projected Stock 2030",
                "9.5M units",
                "https://ifr.org/worldrobotics/",
            ),
            (
                "[21]",
                "IFR World Robotics",
                "Fixtures per Robot",
                "5-15 fixtures/robot",
                "https://ifr.org/worldrobotics/",
            ),
            (
                "[21]",
                "IFR World Robotics",
                "Fixture Demand Growth",
                "540% by 2030",
                "https://ifr.org/worldrobotics/",
            ),
            ("", "", "", "", ""),
            ("", "INDUSTRY DATA SOURCES", "", "", ""),
            (
                "[10]",
                "Stratasys Annual Report",
                "Tooling Applications %",
                "20-25% of market",
                "https://investors.stratasys.com/",
            ),
            (
                "[11]",
                "3D Systems 10-K Filing",
                "Jigs/Fixtures Segment",
                "Fastest growing",
                "https://www.sec.gov/cgi-bin/browse-edgar?action=getcompany&CIK=0000910638",
            ),
            (
                "[13]",
                "IBEF",
                "India Manufacturing Output",
                "$490B (2025)",
                "https://www.ibef.org/industry/manufacturing-sector-india",
            ),
            (
                "[14]",
                "ASEAN Statistics",
                "SE Asia Manufacturing",
                "$100B+ (2025)",
                "https://www.aseanstats.org/",
            ),
            ("", "", "", "", ""),
            ("", "SERVICES TAM VALIDATION (Triple-Validated)", "", "", ""),
            (
                "[32]",
                "Mordor Intelligence",
                "Engineering Services Outsourcing",
                "$315.61B (2025)",
                "https://www.mordorintelligence.com/industry-reports/global-engineering-services-outsourcing-market",
            ),
            (
                "[33]",
                "Market Report Analytics",
                "Mechanical Eng Outsourcing",
                "$50B (2025)",
                "https://www.marketreportanalytics.com/reports/mechanical-engineering-outsourcing-75833",
            ),
            (
                "[34]",
                "SNS Insider",
                "Contract Mfg & Design Services",
                "$621.36B (2024)",
                "https://www.snsinsider.com/reports/contract-manufacturing-and-design-services-market-4119",
            ),
            (
                "[35]",
                "Verified Market Research",
                "Engineering Design Services",
                "$148.62B (2032)",
                "https://www.verifiedmarketresearch.com/product/engineering-design-service-market/",
            ),
            ("", "", "", "", ""),
            ("", "COMPETITOR BENCHMARKS", "", "", ""),
            (
                "[50]",
                "Trinckle (FixtureMate)",
                "Direct competitor - 12yr head start",
                "Est. $5-10M ARR",
                "https://trinckle.com/",
            ),
            (
                "[51]",
                "Zoo/KittyCAD",
                "AI CAD - $20M Series A",
                "$99/mo Pro tier",
                "https://zoo.dev/",
            ),
            (
                "[52]",
                "Autodesk",
                "Incumbent CAD - NASDAQ:ADSK",
                "$4.5B mfg revenue",
                "https://investors.autodesk.com/",
            ),
            (
                "[53]",
                "Dassault Systèmes",
                "SOLIDWORKS owner",
                "$6.0B revenue",
                "https://www.3ds.com/investors/",
            ),
            (
                "[54]",
                "PTC",
                "Creo + IoT - NASDAQ:PTC",
                "$2.1B revenue",
                "https://investor.ptc.com/",
            ),
            ("", "", "", "", ""),
            ("", "FUNDING COMPARABLES", "", "", ""),
            (
                "[60]",
                "CB Insights Q3 2024",
                "Median Seed Deal Size",
                "$3.0M",
                "https://www.cbinsights.com/research/report/venture-trends-q3-2024/",
            ),
            (
                "[61]",
                "CB Insights Q3 2024",
                "Median Seed Valuation",
                "$13.5M",
                "https://www.cbinsights.com/research/report/venture-trends-q3-2024/",
            ),
            (
                "[62]",
                "Bessemer State of Cloud",
                "B2B SaaS Series A Median",
                "$10-12M",
                "https://www.bvp.com/atlas/state-of-the-cloud",
            ),
            (
                "[63]",
                "Bessemer State of Cloud",
                "Series A Valuation Multiple",
                "5-7x ARR",
                "https://www.bvp.com/atlas/state-of-the-cloud",
            ),
            (
                "[64]",
                "Carta Data",
                "Founder Ownership at Exit",
                "15-25% (US)",
                "https://carta.com/blog/",
            ),
            (
                "[65]",
                "Y Combinator",
                "Recommended Seed Runway",
                "18-24 months",
                "https://www.ycombinator.com/library/",
            ),
            ("", "", "", "", ""),
            ("", "EXIT MULTIPLES & M&A COMPARABLES", "", "", ""),
            (
                "[70]",
                "Software Equity Group",
                "B2B SaaS EV/Revenue",
                "6-10x revenue",
                "https://softwareequity.com/",
            ),
            (
                "[71]",
                "Autodesk Acquisitions",
                "Avg. premium paid",
                "4-8x revenue",
                "Investor relations",
            ),
            (
                "[72]",
                "PTC/Arena Solutions",
                "Manufacturing cloud M&A",
                "$715M / 8x rev",
                "https://www.ptc.com/en/news/",
            ),
            (
                "[73]",
                "Autodesk/Upchain",
                "PLM SaaS acquisition",
                "Undisclosed/5-7x",
                "https://investors.autodesk.com/",
            ),
            (
                "[74]",
                "Hexagon/Vero Software",
                "CAD/CAM acquisition",
                "$250M / 4x rev",
                "https://hexagon.com/",
            ),
            ("", "", "", "", ""),
            ("", "BENCHMARK & VALIDATION SOURCES", "", "", ""),
            (
                "[20]",
                "Deloitte Manufacturing Institute",
                "Skills Gap Study 2024",
                "2.1M unfilled jobs",
                "https://www2.deloitte.com/us/en/pages/manufacturing/articles/future-of-manufacturing-skills-gap-study.html",
            ),
            (
                "[21]",
                "IFR World Robotics",
                "Robot Stock & Fixtures",
                "3.9M robots (2023)",
                "https://ifr.org/worldrobotics/",
            ),
            (
                "[22]",
                "SaaS Capital",
                "SaaS Valuation Metrics",
                "LTV:CAC 3.2:1 median",
                "https://www.saas-capital.com/research/saas-company-valuation-metrics/",
            ),
            (
                "[23]",
                "ChartMogul",
                "SaaS Churn Benchmarks",
                "SMB 8-10%, Ent 3-5%",
                "https://chartmogul.com/blog/saas-churn-rate-benchmarks/",
            ),
            (
                "[24]",
                "Pacific Crest SaaS Survey",
                "Revenue Growth Benchmarks",
                "40-60% early stage",
                "https://www.keybanccapitalmarkets.com/insights/saas-survey/",
            ),
            ("", "", "", "", ""),
            ("", "CUSTOMER VALIDATION", "", "", ""),
            (
                "[17]",
                "Honda India (Pilot)",
                "Fixture Design Automation",
                "$200K/year contract",
                "Under NDA - available for investor DD",
            ),
            (
                "[18]",
                "TVS Motor (Pilot)",
                "Assembly Line Tooling",
                "$150K/year contract",
                "Under NDA - available for investor DD",
            ),
            (
                "[19]",
                "Toyota India (Early Engagement)",
                "Tier-2 Supplier Network",
                "250+ suppliers",
                "Early engagement stage",
            ),
        ]

        for ref, source_name, data_point, value, url in sources:
            ws.cell(row, 1).value = ref
            if source_name and source_name.isupper():
                ws.cell(row, 2).value = source_name
                ws.cell(row, 2).font = Font(bold=True)
            else:
                ws.cell(row, 2).value = source_name
            ws.cell(row, 3).value = data_point
            ws.cell(row, 4).value = value
            ws.cell(row, 5).value = url
            if url and url.startswith("http"):
                ws.cell(row, 5).font = Font(color="1A4CB3", underline="single")
            row += 1

        row += 2

        # Data confidence summary
        ws.cell(row, 1).value = "DATA CONFIDENCE ASSESSMENT"
        ws.cell(row, 1).font = Font(bold=True)
        row += 1

        confidence = [
            (
                "HIGH Confidence",
                "TAM streams ($42B): All validated by tier-1 research firms",
            ),
            (
                "HIGH Confidence",
                "SAM regional filters: Verified against IBEF, ASEAN, GM Insights",
            ),
            (
                "MEDIUM-HIGH",
                "Services TAM: Triple-validated ($20B converges from 3 methodologies)",
            ),
            (
                "MEDIUM-HIGH",
                "SOM projections: Customer-validated (Honda/TVS pilots converting)",
            ),
            ("MEDIUM", "Year-wise growth rates: Based on Bessemer SaaS benchmarks"),
        ]

        for level, description in confidence:
            ws.cell(row, 1).value = level
            ws.cell(row, 1).font = Font(bold=True)
            ws.cell(row, 2).value = description
            ws.merge_cells(f"B{row}:E{row}")
            row += 1

    def _build_assumptions(self):
        """Sheet 2: Assumptions - All input parameters."""
        print("  âš™ï¸  Building Assumptions...")
        ws = self.wb.create_sheet("Assumptions")

        # Column widths
        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 10
        for i in range(4, 4 + self.num_years):
            ws.column_dimensions[get_column_letter(i)].width = 12

        # Title
        style_title(ws["A1"], "FINANCIAL MODEL ASSUMPTIONS")
        ws.merge_cells(f"A1:{get_column_letter(3 + self.num_years)}1")

        row = 3

        # General Parameters
        style_section_header(ws.cell(row, 1), "GENERAL PARAMETERS")
        ws.merge_cells(f"A{row}:C{row}")
        row += 1

        general = self.config.get("general", {})
        params = [
            ("Tax Rate", general.get("tax_rate", 0.25), "%"),
            ("CapEx Year 0", general.get("capex_y0", 150000), "USD"),
            ("CapEx Annual", general.get("capex_annual", 50000), "USD"),
            ("Depreciation Period", general.get("depreciation_years", 5), "Years"),
            ("Debtor Days", general.get("debtor_days", 45), "Days"),
            ("Creditor Days", general.get("creditor_days", 30), "Days"),
            ("Interest Rate", general.get("interest_rate", 0.10), "%"),
            ("Cost Inflation", general.get("cost_inflation", 0.05), "%"),
        ]

        self.row_refs["assumptions_start"] = row
        for name, value, unit in params:
            ws.cell(row, 1).value = name
            ws.cell(row, 2).value = value
            ws.cell(row, 3).value = unit
            if unit == "%":
                ws.cell(row, 2).number_format = "0.00%"
            elif unit == "USD":
                ws.cell(row, 2).number_format = "[$-409]#,##0"

            # Store row references for key params
            if name == "Tax Rate":
                self.row_refs["tax_rate"] = row
            elif name == "Depreciation Period":
                self.row_refs["depreciation_years"] = row
            elif name == "Cost Inflation":
                self.row_refs["cost_inflation"] = row
            row += 1

        row += 1

        # Revenue Streams
        style_section_header(ws.cell(row, 1), "REVENUE STREAMS")
        ws.merge_cells(f"A{row}:{get_column_letter(3 + self.num_years)}{row}")
        row += 1

        # Revenue headers
        headers = ["Stream", "Price", "Unit"] + get_year_headers(self.num_years)
        style_column_headers(
            ws, row, headers, bg=Colors.MEDIUM_BLUE, end_col=3 + self.num_years
        )
        row += 1

        # Revenue stream data
        self.row_refs["revenue_streams_start"] = row
        streams = self.config.get("revenue_streams", [])

        for stream in streams:
            name = stream.get("name", "Product")
            price = stream.get("price", 1000)
            volume_y0 = stream.get("volume", 10)
            growth = stream.get("growth", 0.25)

            # Price row
            ws.cell(row, 1).value = f"{name}: Price"
            ws.cell(row, 2).value = price
            ws.cell(row, 2).number_format = "[$-409]#,##0"
            ws.cell(row, 3).value = "USD"
            row += 1

            # Volume row with growth formula
            ws.cell(row, 1).value = f"{name}: Volume"
            ws.cell(row, 3).value = "Units"
            for yr in range(self.num_years):
                col = 4 + yr
                if yr == 0:
                    ws.cell(row, col).value = volume_y0
                else:
                    prev_col = get_column_letter(col - 1)
                    ws.cell(row, col).value = f"=ROUND({prev_col}{row}*(1+{growth}),0)"
                ws.cell(row, col).number_format = "[$-409]#,##0"
            row += 1

            # Growth rate row
            ws.cell(row, 1).value = f"{name}: Growth"
            ws.cell(row, 2).value = growth
            ws.cell(row, 2).number_format = "0.0%"
            ws.cell(row, 3).value = "%"
            row += 1

            # COGS % row
            ws.cell(row, 1).value = f"{name}: COGS %"
            ws.cell(row, 2).value = stream.get("cogs_pct", 0.30)
            ws.cell(row, 2).number_format = "0.0%"
            ws.cell(row, 3).value = "%"
            row += 1

        self.row_refs["revenue_streams_end"] = row - 1
        row += 1

        # Fixed Costs
        style_section_header(ws.cell(row, 1), "FIXED COSTS (Annual)")
        ws.merge_cells(f"A{row}:C{row}")
        row += 1

        self.row_refs["fixed_costs_start"] = row
        fixed_costs = self.config.get("fixed_costs", {})
        if isinstance(fixed_costs, list):
            fixed_costs = {item["name"]: item["value"] for item in fixed_costs}

        for name, value in fixed_costs.items():
            ws.cell(row, 1).value = name
            ws.cell(row, 2).value = value
            ws.cell(row, 2).number_format = "[$-409]#,##0"
            ws.cell(row, 3).value = "USD"
            row += 1

        self.row_refs["fixed_costs_end"] = row - 1
        row += 1

        # Funding Parameters
        style_section_header(ws.cell(row, 1), "FUNDING PARAMETERS")
        ws.merge_cells(f"A{row}:C{row}")
        row += 1

        funding = self.config.get("funding", {})
        funding_params = [
            ("Seed Round", funding.get("seed", 3000000), "USD"),
            ("Series A", funding.get("series_a", 10000000), "USD"),
            ("Series B", funding.get("series_b", 25000000), "USD"),
            ("Seed Timing (Year)", funding.get("seed_year", 0), "Year"),
            ("Series A Timing (Year)", funding.get("series_a_year", 2), "Year"),
            ("Series B Timing (Year)", funding.get("series_b_year", 4), "Year"),
        ]

        self.row_refs["funding_start"] = row
        for name, value, unit in funding_params:
            ws.cell(row, 1).value = name
            ws.cell(row, 2).value = value
            if unit == "USD":
                ws.cell(row, 2).number_format = "[$-409]#,##0"
            ws.cell(row, 3).value = unit
            row += 1
        self.row_refs["funding_end"] = row - 1
        row += 1

        # Customer Segmentation
        style_section_header(ws.cell(row, 1), "CUSTOMER SEGMENTATION (Year 8)")
        ws.merge_cells(f"A{row}:F{row}")
        row += 1

        customers = self.config.get("customers", {})
        segments = customers.get("segments", {})

        # Headers for customer segments
        segment_headers = ["Segment", "Count", "% Mix", "ARPU", "Churn", "GM %"]
        for col, header in enumerate(segment_headers, 1):
            ws.cell(row, col).value = header
            style_header(ws.cell(row, col), bg=Colors.MEDIUM_BLUE, size=10)
        row += 1

        self.row_refs["customers_start"] = row
        for seg_name, seg_data in [
            ("SMB", segments.get("smb", {"count": 3850, "pct": 0.70, "arpu": 9300, "churn": 0.17})),
            ("Mid-Market", segments.get("midmarket", {"count": 1375, "pct": 0.25, "arpu": 37200, "churn": 0.11})),
            ("Enterprise", segments.get("enterprise", {"count": 275, "pct": 0.05, "arpu": 62000, "churn": 0.06})),
        ]:
            ws.cell(row, 1).value = seg_name
            ws.cell(row, 2).value = seg_data.get("count", 0)
            ws.cell(row, 2).number_format = "[$-409]#,##0"
            ws.cell(row, 3).value = seg_data.get("pct", 0)
            ws.cell(row, 3).number_format = "0.0%"
            ws.cell(row, 4).value = seg_data.get("arpu", 0)
            ws.cell(row, 4).number_format = "[$-409]#,##0"
            ws.cell(row, 5).value = seg_data.get("churn", 0)
            ws.cell(row, 5).number_format = "0.0%"
            ws.cell(row, 6).value = seg_data.get("gross_margin", 0.65)
            ws.cell(row, 6).number_format = "0.0%"
            row += 1

        # Total customers row
        ws.cell(row, 1).value = "TOTAL"
        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 2).value = customers.get("total_y8", 5500)
        ws.cell(row, 2).number_format = "[$-409]#,##0"
        ws.cell(row, 2).font = Font(bold=True)
        row += 2

        # Unit Economics
        style_section_header(ws.cell(row, 1), "UNIT ECONOMICS (Blended)")
        ws.merge_cells(f"A{row}:C{row}")
        row += 1

        unit_econ = self.config.get("unit_economics", {})
        unit_params = [
            ("Blended ARPU", unit_econ.get("blended_arpu", 18909), "USD"),
            ("Blended CAC", unit_econ.get("blended_cac", 7000), "USD"),
            ("Blended Churn Rate", unit_econ.get("blended_churn", 0.15), "%"),
            ("Blended Lifetime (years)", unit_econ.get("blended_lifetime", 6.7), "Years"),
            ("LTV", unit_econ.get("ltv", 82000), "USD"),
            ("LTV:CAC Ratio", unit_econ.get("ltv_cac_ratio", 11.7), "x"),
            ("Payback Period (months)", unit_econ.get("payback_months", 6), "Months"),
            ("Net Revenue Retention", unit_econ.get("nrr_target", 1.11), "%"),
        ]

        self.row_refs["unit_econ_start"] = row
        for name, value, unit in unit_params:
            ws.cell(row, 1).value = name
            ws.cell(row, 2).value = value
            if unit == "USD":
                ws.cell(row, 2).number_format = "[$-409]#,##0"
            elif unit == "%":
                ws.cell(row, 2).number_format = "0.0%" if value < 2 else "0%"
            elif unit == "x":
                ws.cell(row, 2).number_format = "0.0"
            ws.cell(row, 3).value = unit
            row += 1
        row += 1

        # Pricing Architecture
        style_section_header(ws.cell(row, 1), "SOFTWARE PRICING ARCHITECTURE")
        ws.merge_cells(f"A{row}:C{row}")
        row += 1

        pricing = self.config.get("pricing", {})
        pricing_params = [
            ("Base Price/Seat/Year", pricing.get("software_per_seat", 2400), "USD"),
            ("SMB Avg Seats", pricing.get("smb_seats", 2.0), "Seats"),
            ("Mid-Market Avg Seats", pricing.get("midmarket_seats", 8.0), "Seats"),
            ("Enterprise Avg Seats", pricing.get("enterprise_seats", 25.0), "Seats"),
            ("Mid-Market Discount", pricing.get("midmarket_discount", 0.10), "%"),
            ("Enterprise Discount", pricing.get("enterprise_discount", 0.20), "%"),
        ]

        for name, value, unit in pricing_params:
            ws.cell(row, 1).value = name
            ws.cell(row, 2).value = value
            if unit == "USD":
                ws.cell(row, 2).number_format = "[$-409]#,##0"
            elif unit == "%":
                ws.cell(row, 2).number_format = "0.0%"
            elif unit == "Seats":
                ws.cell(row, 2).number_format = "0.0"
            ws.cell(row, 3).value = unit
            row += 1
        row += 1

        # Geographic Expansion
        style_section_header(ws.cell(row, 1), "GEOGRAPHIC EXPANSION PHASES")
        ws.merge_cells(f"A{row}:D{row}")
        row += 1

        geo_headers = ["Phase", "Regions", "Years", "SAM"]
        for col, header in enumerate(geo_headers, 1):
            ws.cell(row, col).value = header
            style_header(ws.cell(row, col), bg=Colors.MEDIUM_BLUE, size=10)
        row += 1

        geographic = self.config.get("geographic", {})
        geo_phases = [
            ("Phase 1", geographic.get("phase1", {}).get("regions", ["India"]),
             geographic.get("phase1", {}).get("years", "Y1-Y2"), 1800000000),
            ("Phase 2", geographic.get("phase2", {}).get("regions", ["India", "SE Asia"]),
             geographic.get("phase2", {}).get("years", "Y3-Y4"), 2880000000),
            ("Phase 3", geographic.get("phase3", {}).get("regions", ["India", "SE Asia", "US/EU"]),
             geographic.get("phase3", {}).get("years", "Y5-Y8"), 5760000000),
        ]

        for phase, regions, years, sam in geo_phases:
            ws.cell(row, 1).value = phase
            ws.cell(row, 2).value = ", ".join(regions) if isinstance(regions, list) else str(regions)
            ws.cell(row, 3).value = years
            ws.cell(row, 4).value = sam
            ws.cell(row, 4).number_format = "[$-409]#,##0"
            row += 1
        row += 1

        # Efficiency Metrics
        style_section_header(ws.cell(row, 1), "EFFICIENCY METRICS (Targets)")
        ws.merge_cells(f"A{row}:C{row}")
        row += 1

        efficiency = self.config.get("efficiency_metrics", {})
        eff_params = [
            ("Revenue/Employee (Y8)", efficiency.get("revenue_per_employee_y8", 274000), "USD"),
            ("Burn Multiple (Y1-Y4)", efficiency.get("burn_multiple_y1_y4", 0.8), "x"),
            ("ARR per $ Raised", efficiency.get("arr_per_dollar_raised", 2.74), "x"),
            ("Time to Profitability", efficiency.get("time_to_profitability", "Y4"), "Year"),
            ("S&M as % Revenue (Y8)", efficiency.get("sm_as_pct_revenue_y8", 0.24), "%"),
            ("R&D as % Revenue (Y8)", efficiency.get("rd_as_pct_revenue_y8", 0.07), "%"),
            ("G&A as % Revenue (Y8)", efficiency.get("ga_as_pct_revenue_y8", 0.07), "%"),
            ("Rule of 40 (Y8)", efficiency.get("rule_of_40_y8", 59), "Score"),
        ]

        for name, value, unit in eff_params:
            ws.cell(row, 1).value = name
            ws.cell(row, 2).value = value
            if unit == "USD":
                ws.cell(row, 2).number_format = "[$-409]#,##0"
            elif unit == "%":
                ws.cell(row, 2).number_format = "0.0%"
            elif unit == "x":
                ws.cell(row, 2).number_format = "0.00"
            ws.cell(row, 3).value = unit
            row += 1

    def _build_headcount(self):
        """Sheet 3: Headcount Plan."""
        print("  ðŸ‘¥ Building Headcount Plan...")
        ws = self.wb.create_sheet("Headcount Plan")

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 15
        for i in range(3, 3 + self.num_years):
            ws.column_dimensions[get_column_letter(i)].width = 12

        # Title
        style_title(ws["A1"], "HEADCOUNT PLAN")
        ws.merge_cells(f"A1:{get_column_letter(2 + self.num_years)}1")

        row = 3

        # Headers
        headers = ["Department", "Avg Salary"] + get_year_headers(self.num_years)
        style_column_headers(ws, row, headers, end_col=2 + self.num_years)
        row += 1

        # Headcount data
        headcount = self.config.get("headcount", {})
        departments = [
            (
                "Engineering",
                headcount.get("engineering_salary", 80000),
                headcount.get("engineering_y0", 5),
                0.40,
            ),
            (
                "Sales & Marketing",
                headcount.get("sales_salary", 60000),
                headcount.get("sales_y0", 3),
                0.50,
            ),
            (
                "Operations",
                headcount.get("ops_salary", 50000),
                headcount.get("ops_y0", 2),
                0.35,
            ),
            ("G&A", headcount.get("ga_salary", 70000), headcount.get("ga_y0", 2), 0.25),
        ]

        self.row_refs["headcount_start"] = row
        for dept, salary, y0_count, growth in departments:
            ws.cell(row, 1).value = dept
            ws.cell(row, 2).value = salary
            ws.cell(row, 2).number_format = "[$-409]#,##0"

            for yr in range(self.num_years):
                col = 3 + yr
                if yr == 0:
                    ws.cell(row, col).value = y0_count
                else:
                    prev_col = get_column_letter(col - 1)
                    ws.cell(row, col).value = f"=ROUND({prev_col}{row}*(1+{growth}),0)"
                ws.cell(row, col).number_format = "[$-409]#,##0"
            row += 1

        self.row_refs["headcount_end"] = row - 1

        # Total headcount
        ws.cell(row, 1).value = "TOTAL HEADCOUNT"
        ws.cell(row, 1).font = Font(bold=True)
        for yr in range(self.num_years):
            col = 3 + yr
            col_letter = get_column_letter(col)
            ws.cell(row, col).value = (
                f'=SUM({col_letter}{self.row_refs["headcount_start"]}:{col_letter}{self.row_refs["headcount_end"]})'
            )
            ws.cell(row, col).number_format = "[$-409]#,##0"
            ws.cell(row, col).font = Font(bold=True)
        style_data_row(ws, row, 1, 2 + self.num_years, is_total=True)
        self.row_refs["headcount_total"] = row
        row += 2

        # Total salary cost
        ws.cell(row, 1).value = "TOTAL SALARY COST"
        ws.cell(row, 1).font = Font(bold=True)
        for yr in range(self.num_years):
            col = 3 + yr
            col_letter = get_column_letter(col)
            # Sum of (headcount Ã— salary) for each department
            formula_parts = []
            for dept_row in range(
                self.row_refs["headcount_start"], self.row_refs["headcount_end"] + 1
            ):
                formula_parts.append(f"({col_letter}{dept_row}*B{dept_row})")
            ws.cell(row, col).value = f'={"+".join(formula_parts)}'
            ws.cell(row, col).number_format = "[$-409]#,##0"
            ws.cell(row, col).font = Font(bold=True)
        style_data_row(ws, row, 1, 2 + self.num_years, is_total=True)
        self.row_refs["salary_cost_total"] = row

    def _build_revenue(self):
        """Sheet 4: Revenue calculations."""
        print("  ðŸ“ˆ Building Revenue...")
        ws = self.wb.create_sheet("Revenue")

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 10
        for i in range(3, 3 + self.num_years):
            ws.column_dimensions[get_column_letter(i)].width = 14

        # Title
        style_title(ws["A1"], "REVENUE PROJECTIONS")
        ws.merge_cells(f"A1:{get_column_letter(2 + self.num_years)}1")

        row = 3

        # Headers
        headers = ["Revenue Stream", "Unit"] + get_year_headers(self.num_years)
        style_column_headers(ws, row, headers, end_col=2 + self.num_years)
        row += 1

        # Revenue calculations per stream
        streams = self.config.get("revenue_streams", [])
        self.row_refs["revenue_start"] = row

        # Calculate which rows in Assumptions correspond to each stream
        # Each stream takes 4 rows (Price, Volume, Growth, COGS%)
        assumptions_stream_start = self.row_refs.get("revenue_streams_start", 17)

        for idx, stream in enumerate(streams):
            name = stream.get("name", f"Stream {idx+1}")
            ws.cell(row, 1).value = name
            ws.cell(row, 2).value = "USD"

            # Price is at assumptions_stream_start + (idx * 4)
            # Volume is at assumptions_stream_start + (idx * 4) + 1
            price_row = assumptions_stream_start + (idx * 4)
            volume_row = price_row + 1

            for yr in range(self.num_years):
                col = 3 + yr
                yr_col = get_column_letter(
                    4 + yr
                )  # Assumptions years start at column D
                # Revenue = Price Ã— Volume
                ws.cell(row, col).value = (
                    f"=Assumptions!$B${price_row}*Assumptions!{yr_col}${volume_row}"
                )
                ws.cell(row, col).number_format = "[$-409]#,##0"
            row += 1

        self.row_refs["revenue_end"] = row - 1
        row += 1

        # Total Revenue
        ws.cell(row, 1).value = "TOTAL REVENUE"
        ws.cell(row, 1).font = Font(bold=True, size=11)
        ws.cell(row, 2).value = "USD"
        for yr in range(self.num_years):
            col = 3 + yr
            col_letter = get_column_letter(col)
            ws.cell(row, col).value = (
                f'=SUM({col_letter}{self.row_refs["revenue_start"]}:{col_letter}{self.row_refs["revenue_end"]})'
            )
            ws.cell(row, col).number_format = "[$-409]#,##0"
            ws.cell(row, col).font = Font(bold=True)
        style_data_row(ws, row, 1, 2 + self.num_years, is_total=True)
        self.row_refs["revenue_total"] = row
        row += 2

        # Revenue Mix %
        style_section_header(ws.cell(row, 1), "REVENUE MIX %", bg=Colors.MEDIUM_BLUE)
        ws.merge_cells(f"A{row}:{get_column_letter(2 + self.num_years)}{row}")
        row += 1

        for idx, stream in enumerate(streams):
            name = stream.get("name", f"Stream {idx+1}")
            ws.cell(row, 1).value = f"{name} %"
            ws.cell(row, 2).value = "%"

            stream_row = self.row_refs["revenue_start"] + idx
            for yr in range(self.num_years):
                col = 3 + yr
                col_letter = get_column_letter(col)
                ws.cell(row, col).value = (
                    f'=IF({col_letter}{self.row_refs["revenue_total"]}=0,0,{col_letter}{stream_row}/{col_letter}{self.row_refs["revenue_total"]})'
                )
                ws.cell(row, col).number_format = "0.0%"
            row += 1

    def _build_operating_costs(self):
        """Sheet 5: Operating Costs - COGS, Fixed, S&M."""
        print("  ðŸ’° Building Operating Costs...")
        ws = self.wb.create_sheet("Operating Costs")

        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 10
        for i in range(3, 3 + self.num_years):
            ws.column_dimensions[get_column_letter(i)].width = 14

        # Title
        style_title(ws["A1"], "OPERATING COSTS")
        ws.merge_cells(f"A1:{get_column_letter(2 + self.num_years)}1")

        row = 3

        # COGS Section
        style_section_header(ws.cell(row, 1), "COST OF GOODS SOLD (COGS)")
        ws.merge_cells(f"A{row}:{get_column_letter(2 + self.num_years)}{row}")
        row += 1

        # COGS per stream
        streams = self.config.get("revenue_streams", [])
        self.row_refs["cogs_start"] = row
        assumptions_stream_start = self.row_refs.get("revenue_streams_start", 17)

        for idx, stream in enumerate(streams):
            name = stream.get("name", f"Stream {idx+1}")
            ws.cell(row, 1).value = f"COGS: {name}"
            ws.cell(row, 2).value = "USD"

            # COGS % is at assumptions_stream_start + (idx * 4) + 3
            cogs_pct_row = assumptions_stream_start + (idx * 4) + 3
            revenue_row = self.row_refs["revenue_start"] + idx

            for yr in range(self.num_years):
                col = 3 + yr
                col_letter = get_column_letter(col)
                # COGS = Revenue Ã— COGS%
                ws.cell(row, col).value = (
                    f"=Revenue!{col_letter}{revenue_row}*Assumptions!$B${cogs_pct_row}"
                )
                ws.cell(row, col).number_format = "[$-409]#,##0"
            row += 1

        self.row_refs["cogs_end"] = row - 1

        # Total COGS
        ws.cell(row, 1).value = "TOTAL COGS"
        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 2).value = "USD"
        for yr in range(self.num_years):
            col = 3 + yr
            col_letter = get_column_letter(col)
            ws.cell(row, col).value = (
                f'=SUM({col_letter}{self.row_refs["cogs_start"]}:{col_letter}{self.row_refs["cogs_end"]})'
            )
            ws.cell(row, col).number_format = "[$-409]#,##0"
            ws.cell(row, col).font = Font(bold=True)
        style_data_row(ws, row, 1, 2 + self.num_years, is_total=True)
        self.row_refs["cogs_total"] = row
        row += 2

        # Fixed Costs Section
        style_section_header(ws.cell(row, 1), "FIXED COSTS")
        ws.merge_cells(f"A{row}:{get_column_letter(2 + self.num_years)}{row}")
        row += 1

        # Salary costs from Headcount sheet
        ws.cell(row, 1).value = "Salaries & Benefits"
        ws.cell(row, 2).value = "USD"
        for yr in range(self.num_years):
            col = 3 + yr
            col_letter = get_column_letter(col)
            ws.cell(row, col).value = (
                f"='Headcount Plan'!{col_letter}{self.row_refs.get('salary_cost_total', 10)}"
            )
            ws.cell(row, col).number_format = "[$-409]#,##0"
        self.row_refs["fixed_salaries"] = row
        row += 1

        # Other fixed costs from Assumptions
        self.row_refs["other_fixed_start"] = row
        fixed_costs = self.config.get("fixed_costs", {})
        if isinstance(fixed_costs, list):
            fixed_costs = {item["name"]: item["value"] for item in fixed_costs}

        # Skip salaries if in fixed_costs (already from Headcount)
        cost_inflation = self.config.get("general", {}).get("cost_inflation", 0.05)

        for name, value in fixed_costs.items():
            if "salary" not in name.lower() and "salaries" not in name.lower():
                ws.cell(row, 1).value = name
                ws.cell(row, 2).value = "USD"
                for yr in range(self.num_years):
                    col = 3 + yr
                    if yr == 0:
                        ws.cell(row, col).value = value
                    else:
                        prev_col = get_column_letter(col - 1)
                        ws.cell(row, col).value = (
                            f"={prev_col}{row}*(1+{cost_inflation})"
                        )
                    ws.cell(row, col).number_format = "[$-409]#,##0"
                row += 1

        self.row_refs["other_fixed_end"] = row - 1

        # Total Fixed Costs
        ws.cell(row, 1).value = "TOTAL FIXED COSTS"
        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 2).value = "USD"
        for yr in range(self.num_years):
            col = 3 + yr
            col_letter = get_column_letter(col)
            ws.cell(row, col).value = (
                f'={col_letter}{self.row_refs["fixed_salaries"]}+SUM({col_letter}{self.row_refs["other_fixed_start"]}:{col_letter}{self.row_refs["other_fixed_end"]})'
            )
            ws.cell(row, col).number_format = "[$-409]#,##0"
            ws.cell(row, col).font = Font(bold=True)
        style_data_row(ws, row, 1, 2 + self.num_years, is_total=True)
        self.row_refs["fixed_total"] = row
        row += 2

        # Total Operating Costs
        ws.cell(row, 1).value = "TOTAL OPERATING COSTS"
        ws.cell(row, 1).font = Font(bold=True, size=11)
        ws.cell(row, 2).value = "USD"
        for yr in range(self.num_years):
            col = 3 + yr
            col_letter = get_column_letter(col)
            ws.cell(row, col).value = (
                f'={col_letter}{self.row_refs["cogs_total"]}+{col_letter}{self.row_refs["fixed_total"]}'
            )
            ws.cell(row, col).number_format = "[$-409]#,##0"
            ws.cell(row, col).font = Font(bold=True)
        style_data_row(ws, row, 1, 2 + self.num_years, is_total=True)
        self.row_refs["opex_total"] = row

    def _build_pnl(self):
        """Sheet 6: Profit & Loss Statement."""
        print("  ðŸ“Š Building P&L...")
        ws = self.wb.create_sheet("P&L")

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 10
        for i in range(3, 3 + self.num_years):
            ws.column_dimensions[get_column_letter(i)].width = 14

        # Title
        style_title(ws["A1"], "PROFIT & LOSS STATEMENT")
        ws.merge_cells(f"A1:{get_column_letter(2 + self.num_years)}1")

        row = 3

        # Headers
        headers = ["Line Item", "Unit"] + get_year_headers(self.num_years)
        style_column_headers(ws, row, headers, end_col=2 + self.num_years)
        row += 1

        # Revenue
        ws.cell(row, 1).value = "Revenue"
        ws.cell(row, 2).value = "USD"
        for yr in range(self.num_years):
            col = 3 + yr
            col_letter = get_column_letter(col)
            ws.cell(row, col).value = (
                f'=Revenue!{col_letter}{self.row_refs["revenue_total"]}'
            )
            ws.cell(row, col).number_format = "[$-409]#,##0"
        self.row_refs["pnl_revenue"] = row
        row += 1

        # COGS
        ws.cell(row, 1).value = "Cost of Goods Sold"
        ws.cell(row, 2).value = "USD"
        for yr in range(self.num_years):
            col = 3 + yr
            col_letter = get_column_letter(col)
            ws.cell(row, col).value = (
                f"='Operating Costs'!{col_letter}{self.row_refs['cogs_total']}"
            )
            ws.cell(row, col).number_format = "[$-409]#,##0"
        self.row_refs["pnl_cogs"] = row
        row += 1

        # Gross Profit
        ws.cell(row, 1).value = "GROSS PROFIT"
        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 2).value = "USD"
        for yr in range(self.num_years):
            col = 3 + yr
            col_letter = get_column_letter(col)
            ws.cell(row, col).value = (
                f'={col_letter}{self.row_refs["pnl_revenue"]}-{col_letter}{self.row_refs["pnl_cogs"]}'
            )
            ws.cell(row, col).number_format = "[$-409]#,##0"
            ws.cell(row, col).font = Font(bold=True)
        self.row_refs["pnl_gross_profit"] = row
        row += 1

        # Gross Margin %
        ws.cell(row, 1).value = "Gross Margin %"
        ws.cell(row, 2).value = "%"
        for yr in range(self.num_years):
            col = 3 + yr
            col_letter = get_column_letter(col)
            ws.cell(row, col).value = (
                f'=IF({col_letter}{self.row_refs["pnl_revenue"]}=0,0,{col_letter}{self.row_refs["pnl_gross_profit"]}/{col_letter}{self.row_refs["pnl_revenue"]})'
            )
            ws.cell(row, col).number_format = "0.0%"
        row += 2

        # Operating Expenses
        ws.cell(row, 1).value = "Operating Expenses"
        ws.cell(row, 2).value = "USD"
        for yr in range(self.num_years):
            col = 3 + yr
            col_letter = get_column_letter(col)
            ws.cell(row, col).value = (
                f"='Operating Costs'!{col_letter}{self.row_refs['fixed_total']}"
            )
            ws.cell(row, col).number_format = "[$-409]#,##0"
        self.row_refs["pnl_opex"] = row
        row += 1

        # EBITDA
        ws.cell(row, 1).value = "EBITDA"
        ws.cell(row, 1).font = Font(bold=True, size=11)
        ws.cell(row, 2).value = "USD"
        for yr in range(self.num_years):
            col = 3 + yr
            col_letter = get_column_letter(col)
            ws.cell(row, col).value = (
                f'={col_letter}{self.row_refs["pnl_gross_profit"]}-{col_letter}{self.row_refs["pnl_opex"]}'
            )
            ws.cell(row, col).number_format = "[$-409]#,##0"
            ws.cell(row, col).font = Font(bold=True)
        style_data_row(ws, row, 1, 2 + self.num_years, is_total=True)
        self.row_refs["pnl_ebitda"] = row
        row += 1

        # EBITDA Margin %
        ws.cell(row, 1).value = "EBITDA Margin %"
        ws.cell(row, 2).value = "%"
        for yr in range(self.num_years):
            col = 3 + yr
            col_letter = get_column_letter(col)
            ws.cell(row, col).value = (
                f'=IF({col_letter}{self.row_refs["pnl_revenue"]}=0,0,{col_letter}{self.row_refs["pnl_ebitda"]}/{col_letter}{self.row_refs["pnl_revenue"]})'
            )
            ws.cell(row, col).number_format = "0.0%"
        row += 2

        # Depreciation
        general = self.config.get("general", {})
        capex_y0 = general.get("capex_y0", 150000)
        capex_annual = general.get("capex_annual", 50000)
        dep_years = general.get("depreciation_years", 5)

        ws.cell(row, 1).value = "Depreciation"
        ws.cell(row, 2).value = "USD"
        for yr in range(self.num_years):
            col = 3 + yr
            # Simplified depreciation: Y0 capex / dep_years + annual capex / dep_years
            dep_value = capex_y0 / dep_years
            if yr > 0:
                dep_value += (capex_annual * min(yr, dep_years)) / dep_years
            ws.cell(row, col).value = round(dep_value)
            ws.cell(row, col).number_format = "[$-409]#,##0"
        self.row_refs["pnl_depreciation"] = row
        row += 1

        # EBIT
        ws.cell(row, 1).value = "EBIT (Operating Profit)"
        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 2).value = "USD"
        for yr in range(self.num_years):
            col = 3 + yr
            col_letter = get_column_letter(col)
            ws.cell(row, col).value = (
                f'={col_letter}{self.row_refs["pnl_ebitda"]}-{col_letter}{self.row_refs["pnl_depreciation"]}'
            )
            ws.cell(row, col).number_format = "[$-409]#,##0"
            ws.cell(row, col).font = Font(bold=True)
        self.row_refs["pnl_ebit"] = row
        row += 2

        # Interest (placeholder - 0 if no debt)
        ws.cell(row, 1).value = "Interest Expense"
        ws.cell(row, 2).value = "USD"
        for yr in range(self.num_years):
            col = 3 + yr
            ws.cell(row, col).value = 0
            ws.cell(row, col).number_format = "[$-409]#,##0"
        self.row_refs["pnl_interest"] = row
        row += 1

        # PBT
        ws.cell(row, 1).value = "PBT (Profit Before Tax)"
        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 2).value = "USD"
        for yr in range(self.num_years):
            col = 3 + yr
            col_letter = get_column_letter(col)
            ws.cell(row, col).value = (
                f'={col_letter}{self.row_refs["pnl_ebit"]}-{col_letter}{self.row_refs["pnl_interest"]}'
            )
            ws.cell(row, col).number_format = "[$-409]#,##0"
            ws.cell(row, col).font = Font(bold=True)
        self.row_refs["pnl_pbt"] = row
        row += 1

        # Tax
        tax_rate = general.get("tax_rate", 0.25)
        ws.cell(row, 1).value = f"Tax ({int(tax_rate*100)}%)"
        ws.cell(row, 2).value = "USD"
        for yr in range(self.num_years):
            col = 3 + yr
            col_letter = get_column_letter(col)
            ws.cell(row, col).value = (
                f'=MAX(0,{col_letter}{self.row_refs["pnl_pbt"]}*{tax_rate})'
            )
            ws.cell(row, col).number_format = "[$-409]#,##0"
        self.row_refs["pnl_tax"] = row
        row += 1

        # PAT (Net Income)
        ws.cell(row, 1).value = "NET INCOME (PAT)"
        ws.cell(row, 1).font = Font(bold=True, size=11)
        ws.cell(row, 2).value = "USD"
        for yr in range(self.num_years):
            col = 3 + yr
            col_letter = get_column_letter(col)
            ws.cell(row, col).value = (
                f'={col_letter}{self.row_refs["pnl_pbt"]}-{col_letter}{self.row_refs["pnl_tax"]}'
            )
            ws.cell(row, col).number_format = "[$-409]#,##0"
            ws.cell(row, col).font = Font(bold=True)
        style_data_row(ws, row, 1, 2 + self.num_years, is_total=True)
        self.row_refs["pnl_net_income"] = row
        row += 1

        # Net Margin %
        ws.cell(row, 1).value = "Net Margin %"
        ws.cell(row, 2).value = "%"
        for yr in range(self.num_years):
            col = 3 + yr
            col_letter = get_column_letter(col)
            ws.cell(row, col).value = (
                f'=IF({col_letter}{self.row_refs["pnl_revenue"]}=0,0,{col_letter}{self.row_refs["pnl_net_income"]}/{col_letter}{self.row_refs["pnl_revenue"]})'
            )
            ws.cell(row, col).number_format = "0.0%"

    def _build_cash_flow(self):
        """Sheet 7: Cash Flow Statement."""
        print("  ðŸ’µ Building Cash Flow...")
        ws = self.wb.create_sheet("Cash Flow")

        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 10
        for i in range(3, 3 + self.num_years):
            ws.column_dimensions[get_column_letter(i)].width = 14

        # Title
        style_title(ws["A1"], "CASH FLOW STATEMENT")
        ws.merge_cells(f"A1:{get_column_letter(2 + self.num_years)}1")

        row = 3

        # Headers
        headers = ["Line Item", "Unit"] + get_year_headers(self.num_years)
        style_column_headers(ws, row, headers, end_col=2 + self.num_years)
        row += 1

        # Operating Activities
        style_section_header(
            ws.cell(row, 1), "OPERATING ACTIVITIES", bg=Colors.MEDIUM_BLUE
        )
        ws.merge_cells(f"A{row}:{get_column_letter(2 + self.num_years)}{row}")
        row += 1

        # Net Income
        ws.cell(row, 1).value = "Net Income"
        ws.cell(row, 2).value = "USD"
        for yr in range(self.num_years):
            col = 3 + yr
            col_letter = get_column_letter(col)
            ws.cell(row, col).value = (
                f"='P&L'!{col_letter}{self.row_refs['pnl_net_income']}"
            )
            ws.cell(row, col).number_format = "[$-409]#,##0"
        self.row_refs["cf_net_income"] = row
        row += 1

        # Add back Depreciation
        ws.cell(row, 1).value = "+ Depreciation"
        ws.cell(row, 2).value = "USD"
        for yr in range(self.num_years):
            col = 3 + yr
            col_letter = get_column_letter(col)
            ws.cell(row, col).value = (
                f"='P&L'!{col_letter}{self.row_refs['pnl_depreciation']}"
            )
            ws.cell(row, col).number_format = "[$-409]#,##0"
        self.row_refs["cf_depreciation"] = row
        row += 1

        # Working Capital Change (simplified)
        ws.cell(row, 1).value = "Working Capital Change"
        ws.cell(row, 2).value = "USD"
        for yr in range(self.num_years):
            col = 3 + yr
            ws.cell(row, col).value = 0  # Simplified - would need AR/AP calculation
            ws.cell(row, col).number_format = "[$-409]#,##0"
        self.row_refs["cf_wc_change"] = row
        row += 1

        # Operating Cash Flow
        ws.cell(row, 1).value = "Operating Cash Flow"
        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 2).value = "USD"
        for yr in range(self.num_years):
            col = 3 + yr
            col_letter = get_column_letter(col)
            ws.cell(row, col).value = (
                f'={col_letter}{self.row_refs["cf_net_income"]}+{col_letter}{self.row_refs["cf_depreciation"]}-{col_letter}{self.row_refs["cf_wc_change"]}'
            )
            ws.cell(row, col).number_format = "[$-409]#,##0"
            ws.cell(row, col).font = Font(bold=True)
        style_data_row(ws, row, 1, 2 + self.num_years, is_total=True)
        self.row_refs["cf_operating"] = row
        row += 2

        # Investing Activities
        style_section_header(
            ws.cell(row, 1), "INVESTING ACTIVITIES", bg=Colors.MEDIUM_BLUE
        )
        ws.merge_cells(f"A{row}:{get_column_letter(2 + self.num_years)}{row}")
        row += 1

        general = self.config.get("general", {})
        capex_y0 = general.get("capex_y0", 150000)
        capex_annual = general.get("capex_annual", 50000)

        ws.cell(row, 1).value = "Capital Expenditure"
        ws.cell(row, 2).value = "USD"
        for yr in range(self.num_years):
            col = 3 + yr
            capex = capex_y0 if yr == 0 else capex_annual
            ws.cell(row, col).value = -capex
            ws.cell(row, col).number_format = "[$-409]#,##0"
        self.row_refs["cf_capex"] = row
        row += 1

        # Investing Cash Flow
        ws.cell(row, 1).value = "Investing Cash Flow"
        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 2).value = "USD"
        for yr in range(self.num_years):
            col = 3 + yr
            col_letter = get_column_letter(col)
            ws.cell(row, col).value = f'={col_letter}{self.row_refs["cf_capex"]}'
            ws.cell(row, col).number_format = "[$-409]#,##0"
            ws.cell(row, col).font = Font(bold=True)
        style_data_row(ws, row, 1, 2 + self.num_years, is_total=True)
        self.row_refs["cf_investing"] = row
        row += 2

        # Financing Activities
        style_section_header(
            ws.cell(row, 1), "FINANCING ACTIVITIES", bg=Colors.MEDIUM_BLUE
        )
        ws.merge_cells(f"A{row}:{get_column_letter(2 + self.num_years)}{row}")
        row += 1

        funding = self.config.get("funding", {})
        seed = funding.get("seed", 3000000)
        seed_year = funding.get("seed_year", 0)
        series_a = funding.get("series_a", 10000000)
        series_a_year = funding.get("series_a_year", 2)
        series_b = funding.get("series_b", 25000000)
        series_b_year = funding.get("series_b_year", 4)

        ws.cell(row, 1).value = "Equity Raised"
        ws.cell(row, 2).value = "USD"
        for yr in range(self.num_years):
            col = 3 + yr
            equity = 0
            if yr == seed_year:
                equity += seed
            if yr == series_a_year:
                equity += series_a
            if yr == series_b_year:
                equity += series_b
            ws.cell(row, col).value = equity
            ws.cell(row, col).number_format = "[$-409]#,##0"
        self.row_refs["cf_equity"] = row
        row += 1

        # Financing Cash Flow
        ws.cell(row, 1).value = "Financing Cash Flow"
        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 2).value = "USD"
        for yr in range(self.num_years):
            col = 3 + yr
            col_letter = get_column_letter(col)
            ws.cell(row, col).value = f'={col_letter}{self.row_refs["cf_equity"]}'
            ws.cell(row, col).number_format = "[$-409]#,##0"
            ws.cell(row, col).font = Font(bold=True)
        style_data_row(ws, row, 1, 2 + self.num_years, is_total=True)
        self.row_refs["cf_financing"] = row
        row += 2

        # Net Cash Flow
        ws.cell(row, 1).value = "NET CASH FLOW"
        ws.cell(row, 1).font = Font(bold=True, size=11)
        ws.cell(row, 2).value = "USD"
        for yr in range(self.num_years):
            col = 3 + yr
            col_letter = get_column_letter(col)
            ws.cell(row, col).value = (
                f'={col_letter}{self.row_refs["cf_operating"]}+{col_letter}{self.row_refs["cf_investing"]}+{col_letter}{self.row_refs["cf_financing"]}'
            )
            ws.cell(row, col).number_format = "[$-409]#,##0"
            ws.cell(row, col).font = Font(bold=True)
        style_data_row(ws, row, 1, 2 + self.num_years, is_total=True)
        self.row_refs["cf_net"] = row
        row += 2

        # Cumulative Cash
        ws.cell(row, 1).value = "CUMULATIVE CASH"
        ws.cell(row, 1).font = Font(bold=True, size=11)
        ws.cell(row, 2).value = "USD"
        for yr in range(self.num_years):
            col = 3 + yr
            col_letter = get_column_letter(col)
            if yr == 0:
                ws.cell(row, col).value = f'={col_letter}{self.row_refs["cf_net"]}'
            else:
                prev_col = get_column_letter(col - 1)
                ws.cell(row, col).value = (
                    f'={prev_col}{row}+{col_letter}{self.row_refs["cf_net"]}'
                )
            ws.cell(row, col).number_format = "[$-409]#,##0"
            ws.cell(row, col).font = Font(bold=True)
        style_data_row(ws, row, 1, 2 + self.num_years, is_total=True)
        self.row_refs["cf_cumulative"] = row

    def _build_balance_sheet(self):
        """Sheet 8: Balance Sheet."""
        print("  ðŸ“‹ Building Balance Sheet...")
        ws = self.wb.create_sheet("Balance Sheet")

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 10
        for i in range(3, 3 + self.num_years):
            ws.column_dimensions[get_column_letter(i)].width = 14

        # Title
        style_title(ws["A1"], "BALANCE SHEET")
        ws.merge_cells(f"A1:{get_column_letter(2 + self.num_years)}1")

        row = 3

        # Headers
        headers = ["Line Item", "Unit"] + get_year_headers(self.num_years)
        style_column_headers(ws, row, headers, end_col=2 + self.num_years)
        row += 1

        # ASSETS
        style_section_header(ws.cell(row, 1), "ASSETS", bg=Colors.MEDIUM_BLUE)
        ws.merge_cells(f"A{row}:{get_column_letter(2 + self.num_years)}{row}")
        row += 1

        # Cash
        ws.cell(row, 1).value = "Cash & Equivalents"
        ws.cell(row, 2).value = "USD"
        for yr in range(self.num_years):
            col = 3 + yr
            col_letter = get_column_letter(col)
            ws.cell(row, col).value = (
                f"='Cash Flow'!{col_letter}{self.row_refs['cf_cumulative']}"
            )
            ws.cell(row, col).number_format = "[$-409]#,##0"
        self.row_refs["bs_cash"] = row
        row += 1

        # Fixed Assets (simplified - Capex accumulation minus depreciation)
        ws.cell(row, 1).value = "Net Fixed Assets"
        ws.cell(row, 2).value = "USD"
        general = self.config.get("general", {})
        capex_y0 = general.get("capex_y0", 150000)
        capex_annual = general.get("capex_annual", 50000)
        dep_years = general.get("depreciation_years", 5)

        for yr in range(self.num_years):
            col = 3 + yr
            # Cumulative capex - cumulative depreciation
            cum_capex = capex_y0 + (capex_annual * yr)
            cum_dep = (capex_y0 / dep_years) * min(yr + 1, dep_years)
            if yr > 0:
                cum_dep += sum(
                    (capex_annual / dep_years) * min(yr - y, dep_years)
                    for y in range(yr)
                )
            ws.cell(row, col).value = round(max(0, cum_capex - cum_dep))
            ws.cell(row, col).number_format = "[$-409]#,##0"
        self.row_refs["bs_fixed_assets"] = row
        row += 1

        # Total Assets
        ws.cell(row, 1).value = "TOTAL ASSETS"
        ws.cell(row, 1).font = Font(bold=True, size=11)
        ws.cell(row, 2).value = "USD"
        for yr in range(self.num_years):
            col = 3 + yr
            col_letter = get_column_letter(col)
            ws.cell(row, col).value = (
                f'={col_letter}{self.row_refs["bs_cash"]}+{col_letter}{self.row_refs["bs_fixed_assets"]}'
            )
            ws.cell(row, col).number_format = "[$-409]#,##0"
            ws.cell(row, col).font = Font(bold=True)
        style_data_row(ws, row, 1, 2 + self.num_years, is_total=True)
        self.row_refs["bs_total_assets"] = row
        row += 2

        # LIABILITIES & EQUITY
        style_section_header(
            ws.cell(row, 1), "LIABILITIES & EQUITY", bg=Colors.MEDIUM_BLUE
        )
        ws.merge_cells(f"A{row}:{get_column_letter(2 + self.num_years)}{row}")
        row += 1

        # Liabilities (simplified - 0 for now)
        ws.cell(row, 1).value = "Total Liabilities"
        ws.cell(row, 2).value = "USD"
        for yr in range(self.num_years):
            col = 3 + yr
            ws.cell(row, col).value = 0
            ws.cell(row, col).number_format = "[$-409]#,##0"
        self.row_refs["bs_liabilities"] = row
        row += 2

        # Equity
        ws.cell(row, 1).value = "Paid-in Capital"
        ws.cell(row, 2).value = "USD"
        funding = self.config.get("funding", {})
        seed = funding.get("seed", 3000000)
        seed_year = funding.get("seed_year", 0)
        series_a = funding.get("series_a", 10000000)
        series_a_year = funding.get("series_a_year", 2)
        series_b = funding.get("series_b", 25000000)
        series_b_year = funding.get("series_b_year", 4)

        for yr in range(self.num_years):
            col = 3 + yr
            cum_equity = 0
            if yr >= seed_year:
                cum_equity += seed
            if yr >= series_a_year:
                cum_equity += series_a
            if yr >= series_b_year:
                cum_equity += series_b
            ws.cell(row, col).value = cum_equity
            ws.cell(row, col).number_format = "[$-409]#,##0"
        self.row_refs["bs_paid_capital"] = row
        row += 1

        # Retained Earnings
        ws.cell(row, 1).value = "Retained Earnings"
        ws.cell(row, 2).value = "USD"
        for yr in range(self.num_years):
            col = 3 + yr
            col_letter = get_column_letter(col)
            if yr == 0:
                ws.cell(row, col).value = (
                    f"='P&L'!{col_letter}{self.row_refs['pnl_net_income']}"
                )
            else:
                prev_col = get_column_letter(col - 1)
                ws.cell(row, col).value = (
                    f"={prev_col}{row}+'P&L'!{col_letter}{self.row_refs['pnl_net_income']}"
                )
            ws.cell(row, col).number_format = "[$-409]#,##0"
        self.row_refs["bs_retained_earnings"] = row
        row += 1

        # Total Equity
        ws.cell(row, 1).value = "Total Equity"
        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 2).value = "USD"
        for yr in range(self.num_years):
            col = 3 + yr
            col_letter = get_column_letter(col)
            ws.cell(row, col).value = (
                f'={col_letter}{self.row_refs["bs_paid_capital"]}+{col_letter}{self.row_refs["bs_retained_earnings"]}'
            )
            ws.cell(row, col).number_format = "[$-409]#,##0"
            ws.cell(row, col).font = Font(bold=True)
        self.row_refs["bs_total_equity"] = row
        row += 1

        # Total Liabilities & Equity
        ws.cell(row, 1).value = "TOTAL LIABILITIES & EQUITY"
        ws.cell(row, 1).font = Font(bold=True, size=11)
        ws.cell(row, 2).value = "USD"
        for yr in range(self.num_years):
            col = 3 + yr
            col_letter = get_column_letter(col)
            ws.cell(row, col).value = (
                f'={col_letter}{self.row_refs["bs_liabilities"]}+{col_letter}{self.row_refs["bs_total_equity"]}'
            )
            ws.cell(row, col).number_format = "[$-409]#,##0"
            ws.cell(row, col).font = Font(bold=True)
        style_data_row(ws, row, 1, 2 + self.num_years, is_total=True)
        self.row_refs["bs_total_le"] = row
        row += 2

        # Balance Check
        ws.cell(row, 1).value = "BALANCE CHECK (Assets = L+E)"
        ws.cell(row, 1).font = Font(bold=True, color=Colors.GRAY)
        for yr in range(self.num_years):
            col = 3 + yr
            col_letter = get_column_letter(col)
            ws.cell(row, col).value = (
                f'={col_letter}{self.row_refs["bs_total_assets"]}-{col_letter}{self.row_refs["bs_total_le"]}'
            )
            ws.cell(row, col).number_format = "[$-409]#,##0"
        self.row_refs["bs_check"] = row

    def _build_summary(self):
        """Sheet 9: KPI Summary Dashboard."""
        print("  ðŸ“ˆ Building Summary...")
        ws = self.wb.create_sheet("Summary")

        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 10
        for i in range(3, 3 + self.num_years):
            ws.column_dimensions[get_column_letter(i)].width = 14

        # Title
        style_title(ws["A1"], "KEY PERFORMANCE INDICATORS")
        ws.merge_cells(f"A1:{get_column_letter(2 + self.num_years)}1")

        row = 3

        # Headers
        headers = ["Metric", "Unit"] + get_year_headers(self.num_years)
        style_column_headers(ws, row, headers, end_col=2 + self.num_years)
        row += 1

        # Revenue Metrics
        style_section_header(ws.cell(row, 1), "REVENUE METRICS", bg=Colors.MEDIUM_BLUE)
        ws.merge_cells(f"A{row}:{get_column_letter(2 + self.num_years)}{row}")
        row += 1

        kpis = [
            (
                "Total Revenue",
                "USD",
                f"'P&L'!{{col}}{self.row_refs['pnl_revenue']}",
                "[$-409]#,##0",
            ),
            ("Revenue Growth %", "%", None, "0.0%"),  # Calculated below
            (
                "Gross Margin %",
                "%",
                f"'P&L'!{{col}}{self.row_refs['pnl_gross_profit']}/'P&L'!{{col}}{self.row_refs['pnl_revenue']}",
                "0.0%",
            ),
            (
                "EBITDA Margin %",
                "%",
                f"'P&L'!{{col}}{self.row_refs['pnl_ebitda']}/'P&L'!{{col}}{self.row_refs['pnl_revenue']}",
                "0.0%",
            ),
        ]

        for name, unit, formula_template, num_format in kpis:
            ws.cell(row, 1).value = name
            ws.cell(row, 2).value = unit

            for yr in range(self.num_years):
                col = 3 + yr
                col_letter = get_column_letter(col)

                if name == "Revenue Growth %":
                    if yr == 0:
                        ws.cell(row, col).value = 0
                    else:
                        prev_col = get_column_letter(col - 1)
                        rev_row = self.row_refs["pnl_revenue"]
                        ws.cell(row, col).value = (
                            f"=IF('P&L'!{prev_col}{rev_row}=0,0,('P&L'!{col_letter}{rev_row}-'P&L'!{prev_col}{rev_row})/'P&L'!{prev_col}{rev_row})"
                        )
                elif formula_template:
                    formula = formula_template.replace("{col}", col_letter)
                    ws.cell(row, col).value = (
                        f"=IF('P&L'!{col_letter}{self.row_refs['pnl_revenue']}=0,0,{formula})"
                    )

                ws.cell(row, col).number_format = num_format
            row += 1

        row += 1

        # Team Metrics
        style_section_header(ws.cell(row, 1), "TEAM METRICS", bg=Colors.MEDIUM_BLUE)
        ws.merge_cells(f"A{row}:{get_column_letter(2 + self.num_years)}{row}")
        row += 1

        ws.cell(row, 1).value = "Total Headcount"
        ws.cell(row, 2).value = "FTE"
        for yr in range(self.num_years):
            col = 3 + yr
            col_letter = get_column_letter(col)
            ws.cell(row, col).value = (
                f"='Headcount Plan'!{col_letter}{self.row_refs.get('headcount_total', 8)}"
            )
            ws.cell(row, col).number_format = "[$-409]#,##0"
        row += 1

        ws.cell(row, 1).value = "Revenue per Employee"
        ws.cell(row, 2).value = "USD"
        for yr in range(self.num_years):
            col = 3 + yr
            col_letter = get_column_letter(col)
            hc_row = self.row_refs.get("headcount_total", 8)
            rev_row = self.row_refs["pnl_revenue"]
            ws.cell(row, col).value = (
                f"=IF('Headcount Plan'!{col_letter}{hc_row}=0,0,'P&L'!{col_letter}{rev_row}/'Headcount Plan'!{col_letter}{hc_row})"
            )
            ws.cell(row, col).number_format = "[$-409]#,##0"
        row += 2

        # Cash Metrics
        style_section_header(ws.cell(row, 1), "CASH METRICS", bg=Colors.MEDIUM_BLUE)
        ws.merge_cells(f"A{row}:{get_column_letter(2 + self.num_years)}{row}")
        row += 1

        ws.cell(row, 1).value = "Cash Balance"
        ws.cell(row, 2).value = "USD"
        for yr in range(self.num_years):
            col = 3 + yr
            col_letter = get_column_letter(col)
            ws.cell(row, col).value = (
                f"='Cash Flow'!{col_letter}{self.row_refs['cf_cumulative']}"
            )
            ws.cell(row, col).number_format = "[$-409]#,##0"
        row += 1

        ws.cell(row, 1).value = "Net Cash Flow"
        ws.cell(row, 2).value = "USD"
        for yr in range(self.num_years):
            col = 3 + yr
            col_letter = get_column_letter(col)
            ws.cell(row, col).value = (
                f"='Cash Flow'!{col_letter}{self.row_refs['cf_net']}"
            )
            ws.cell(row, col).number_format = "[$-409]#,##0"

    def _build_sensitivity(self):
        """Sheet 10: Sensitivity Analysis."""
        print("  ðŸ”¬ Building Sensitivity Analysis...")
        ws = self.wb.create_sheet("Sensitivity Analysis")

        ws.column_dimensions["A"].width = 30
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 15

        # Title
        style_title(ws["A1"], "SENSITIVITY ANALYSIS")
        ws.merge_cells("A1:D1")

        row = 3

        # Scenario headers
        style_section_header(ws.cell(row, 1), "SCENARIO COMPARISON")
        ws.merge_cells(f"A{row}:D{row}")
        row += 1

        headers = ["Metric", "Downside (-20%)", "Base Case", "Upside (+20%)"]
        style_column_headers(ws, row, headers, bg=Colors.MEDIUM_BLUE, end_col=4)
        row += 1

        ws.cell(row, 1).value = "Year 8 Revenue"
        ws.cell(row, 2).value = "=0.8*C" + str(row)
        ws.cell(row, 3).value = "='P&L'!K4"  # Year 8 revenue (column K = 11 = Year 8)
        ws.cell(row, 4).value = "=1.2*C" + str(row)
        for col in range(2, 5):
            ws.cell(row, col).number_format = "[$-409]#,##0"
        row += 1

        ws.cell(row, 1).value = "Year 8 EBITDA"
        ws.cell(row, 2).value = "=0.7*C" + str(row)  # More sensitive
        ws.cell(row, 3).value = f"='P&L'!K{self.row_refs['pnl_ebitda']}"
        ws.cell(row, 4).value = "=1.3*C" + str(row)
        for col in range(2, 5):
            ws.cell(row, col).number_format = "[$-409]#,##0"
        row += 1

        ws.cell(row, 1).value = "Year 8 Cash"
        ws.cell(row, 2).value = "=0.75*C" + str(row)
        ws.cell(row, 3).value = f"='Cash Flow'!K{self.row_refs['cf_cumulative']}"
        ws.cell(row, 4).value = "=1.25*C" + str(row)
        for col in range(2, 5):
            ws.cell(row, col).number_format = "[$-409]#,##0"

    def _build_valuation(self):
        """Sheet 11: DCF Valuation."""
        print("  ðŸ’Ž Building Valuation...")
        ws = self.wb.create_sheet("Valuation")

        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 15

        # Title
        style_title(ws["A1"], "VALUATION ANALYSIS")
        ws.merge_cells("A1:B1")

        row = 3

        # DCF Inputs
        style_section_header(ws.cell(row, 1), "DCF ASSUMPTIONS")
        ws.merge_cells(f"A{row}:B{row}")
        row += 1

        dcf_inputs = [
            ("Discount Rate (WACC)", 0.15, "0.0%"),
            ("Terminal Growth Rate", 0.03, "0.0%"),
            ("Exit Multiple (EV/Revenue)", 5.0, "0.0x"),
        ]

        for name, value, fmt in dcf_inputs:
            ws.cell(row, 1).value = name
            ws.cell(row, 2).value = value
            ws.cell(row, 2).number_format = fmt
            row += 1

        row += 1

        # Valuation Summary
        style_section_header(ws.cell(row, 1), "VALUATION SUMMARY")
        ws.merge_cells(f"A{row}:B{row}")
        row += 1

        ws.cell(row, 1).value = "Year 8 Revenue"
        ws.cell(row, 2).value = f"='P&L'!K{self.row_refs['pnl_revenue']}"
        ws.cell(row, 2).number_format = "[$-409]#,##0"
        row += 1

        ws.cell(row, 1).value = "Implied Exit Valuation (5x Rev)"
        ws.cell(row, 2).value = f"=B{row-1}*5"
        ws.cell(row, 2).number_format = "[$-409]#,##0"
        ws.cell(row, 2).font = Font(bold=True)

    def _build_breakeven(self):
        """Sheet 12: Break-even Analysis."""
        print("  âš–ï¸  Building Break-even Analysis...")
        ws = self.wb.create_sheet("Break-even Analysis")

        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 15

        # Title
        style_title(ws["A1"], "BREAK-EVEN ANALYSIS")
        ws.merge_cells("A1:B1")

        row = 3

        ws.cell(row, 1).value = "Analysis to be populated based on contribution margin"
        ws.cell(row, 1).font = Font(italic=True, color=Colors.GRAY)

    def _build_cap_table(self):
        """Sheet 13: Funding & Cap Table."""
        print("  ðŸ“Š Building Cap Table...")
        ws = self.wb.create_sheet("Funding Cap Table")

        ws.column_dimensions["A"].width = 25
        ws.column_dimensions["B"].width = 15
        ws.column_dimensions["C"].width = 15
        ws.column_dimensions["D"].width = 15
        ws.column_dimensions["E"].width = 15

        # Title
        style_title(ws["A1"], "FUNDING & CAP TABLE")
        ws.merge_cells("A1:E1")

        row = 3

        # Funding Rounds
        style_section_header(ws.cell(row, 1), "FUNDING ROUNDS")
        ws.merge_cells(f"A{row}:E{row}")
        row += 1

        headers = ["Round", "Amount", "Pre-Money", "Post-Money", "Dilution"]
        style_column_headers(ws, row, headers, bg=Colors.MEDIUM_BLUE, end_col=5)
        row += 1

        funding = self.config.get("funding", {})
        rounds = [
            ("Seed", funding.get("seed", 3000000), funding.get("seed_pre", 10000000)),
            (
                "Series A",
                funding.get("series_a", 10000000),
                funding.get("series_a_pre", 30000000),
            ),
            (
                "Series B",
                funding.get("series_b", 25000000),
                funding.get("series_b_pre", 75000000),
            ),
        ]

        for round_name, amount, pre_money in rounds:
            ws.cell(row, 1).value = round_name
            ws.cell(row, 2).value = amount
            ws.cell(row, 2).number_format = "[$-409]#,##0"
            ws.cell(row, 3).value = pre_money
            ws.cell(row, 3).number_format = "[$-409]#,##0"
            ws.cell(row, 4).value = f"=B{row}+C{row}"
            ws.cell(row, 4).number_format = "[$-409]#,##0"
            ws.cell(row, 5).value = f"=B{row}/D{row}"
            ws.cell(row, 5).number_format = "0.0%"
            row += 1

        # Total Raised
        row += 1
        ws.cell(row, 1).value = "TOTAL RAISED"
        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 2).value = f"=SUM(B{row-4}:B{row-2})"
        ws.cell(row, 2).number_format = "[$-409]#,##0"
        ws.cell(row, 2).font = Font(bold=True)
        style_data_row(ws, row, 1, 5, is_total=True)

    def _build_charts_data(self):
        """Sheet 14: Charts Data."""
        print("  ðŸ“‰ Building Charts Data...")
        ws = self.wb.create_sheet("Charts Data")

        ws.column_dimensions["A"].width = 25
        for i in range(2, 2 + self.num_years):
            ws.column_dimensions[get_column_letter(i)].width = 12

        # Title
        style_title(ws["A1"], "CHARTS DATA")
        ws.merge_cells(f"A1:{get_column_letter(1 + self.num_years)}1")

        row = 3

        # Revenue data for charts
        headers = ["Metric"] + get_year_headers(self.num_years)
        style_column_headers(
            ws, row, headers, bg=Colors.MEDIUM_BLUE, end_col=1 + self.num_years
        )
        row += 1

        ws.cell(row, 1).value = "Revenue"
        for yr in range(self.num_years):
            col = 2 + yr
            col_letter = get_column_letter(3 + yr)  # P&L starts at column C
            ws.cell(row, col).value = (
                f"='P&L'!{col_letter}{self.row_refs['pnl_revenue']}"
            )
            ws.cell(row, col).number_format = "[$-409]#,##0"
        row += 1

        ws.cell(row, 1).value = "EBITDA"
        for yr in range(self.num_years):
            col = 2 + yr
            col_letter = get_column_letter(3 + yr)
            ws.cell(row, col).value = (
                f"='P&L'!{col_letter}{self.row_refs['pnl_ebitda']}"
            )
            ws.cell(row, col).number_format = "[$-409]#,##0"
        row += 1

        ws.cell(row, 1).value = "Net Income"
        for yr in range(self.num_years):
            col = 2 + yr
            col_letter = get_column_letter(3 + yr)
            ws.cell(row, col).value = (
                f"='P&L'!{col_letter}{self.row_refs['pnl_net_income']}"
            )
            ws.cell(row, col).number_format = "[$-409]#,##0"
        row += 1

        ws.cell(row, 1).value = "Cash Balance"
        for yr in range(self.num_years):
            col = 2 + yr
            col_letter = get_column_letter(3 + yr)
            ws.cell(row, col).value = (
                f"='Cash Flow'!{col_letter}{self.row_refs['cf_cumulative']}"
            )
            ws.cell(row, col).number_format = "[$-409]#,##0"

    def save(self, filepath: str):
        """Save the workbook to file."""
        os.makedirs(
            os.path.dirname(filepath) if os.path.dirname(filepath) else ".",
            exist_ok=True,
        )
        self.wb.save(filepath)
        print(f"\nâœ… Saved: {filepath}")
        return filepath


# =============================================================================
# MAIN ENTRY POINT
# =============================================================================


def load_config(config_path: str) -> Dict[str, Any]:
    """Load configuration from JSON file and transform to expected format."""
    if not os.path.exists(config_path):
        print(f"WARNING: Config not found: {config_path}")
        return {}

    with open(config_path, "r", encoding="utf-8") as f:
        raw_config = json.load(f)

    # Transform RapidTools config format to builder format
    config = {}

    # Company info
    if isinstance(raw_config.get("company"), dict):
        config["company"] = raw_config["company"].get("name", "Company")
    else:
        config["company"] = raw_config.get("company", "Company")

    # General parameters
    config["general"] = {
        "tax_rate": 0.25,
        "capex_y0": 150000,
        "capex_annual": 50000,
        "depreciation_years": 5,
        "debtor_days": 45,
        "creditor_days": 30,
        "interest_rate": 0.10,
        "cost_inflation": 0.05,
    }

    # Revenue streams - transform from nested dict to list format
    revenue = raw_config.get("revenue", {})
    streams_data = revenue.get("streams", {})
    margins = raw_config.get("gross_margins", {})

    config["revenue_streams"] = []

    # Map stream names and calculate initial volumes from projections
    stream_mapping = {
        "software": {
            "name": "Software",
            "price": 2400,
            "cogs_pct": 1 - margins.get("software", 0.82),
        },
        "tooling_services": {
            "name": "Tooling Services",
            "price": 5000,
            "cogs_pct": 1 - margins.get("tooling_services", 0.44),
        },
        "ems": {
            "name": "Enterprise Managed Services",
            "price": 20000,
            "cogs_pct": 1 - margins.get("ems", 0.28),
        },
        "hardware": {
            "name": "Hardware Sales",
            "price": 15000,
            "cogs_pct": 1 - margins.get("hardware", 0.26),
        },
        "consumables": {
            "name": "Consumables",
            "price": 2000,
            "cogs_pct": 1 - margins.get("consumables", 0.35),
        },
    }

    for key, mapping in stream_mapping.items():
        stream_revenue = streams_data.get(key, {})
        y1_rev = stream_revenue.get("Y1", 0)
        y8_rev = stream_revenue.get("Y8", 0)

        # Calculate implied volume from revenue / price
        price = mapping["price"]
        volume_y1 = max(1, int(y1_rev / price)) if y1_rev > 0 else 0

        # Calculate implied growth rate
        if y1_rev > 0 and y8_rev > y1_rev:
            growth = (y8_rev / y1_rev) ** (1 / 7) - 1  # 7 years from Y1 to Y8
        else:
            growth = 0.30  # Default 30% growth

        config["revenue_streams"].append(
            {
                "name": mapping["name"],
                "price": price,
                "volume": volume_y1 if volume_y1 > 0 else 10,
                "growth": min(growth, 1.0),  # Cap at 100% growth
                "cogs_pct": mapping["cogs_pct"],
            }
        )

    # Fixed costs
    costs = raw_config.get("costs", {})
    fixed_breakdown = costs.get("fixed_cost_breakdown_y4", {})
    config["fixed_costs"] = {
        "Facilities": fixed_breakdown.get("facilities", 264000),
        "Technology": fixed_breakdown.get("technology", 378000),
        "R&D Infrastructure": fixed_breakdown.get("rd_infra", 168000),
        "Professional Services": fixed_breakdown.get("professional_services", 216000),
        "R&D Partnerships": fixed_breakdown.get("rd_partnerships", 114000),
        "Other G&A": fixed_breakdown.get("other", 610000),
    }

    # Headcount
    headcount = raw_config.get("headcount", {})
    avg_costs = headcount.get("avg_cost_per_fte", {})
    by_year = headcount.get("by_year", {})

    config["headcount"] = {
        "engineering_salary": avg_costs.get("Y1", 50000) * 1.2,  # Engineering premium
        "engineering_y0": int(by_year.get("Y1", 16) * 0.35),
        "sales_salary": avg_costs.get("Y1", 50000),
        "sales_y0": int(by_year.get("Y1", 16) * 0.25),
        "ops_salary": avg_costs.get("Y1", 50000) * 0.8,
        "ops_y0": int(by_year.get("Y1", 16) * 0.25),
        "ga_salary": avg_costs.get("Y1", 50000) * 0.9,
        "ga_y0": int(by_year.get("Y1", 16) * 0.15),
    }

    # Funding
    funding = raw_config.get("funding", {})
    rounds = funding.get("rounds", {})
    seed = rounds.get("seed", {})
    series_a = rounds.get("series_a", {})
    series_b = rounds.get("series_b", {})

    config["funding"] = {
        "seed": seed.get("amount", 3000000),
        "seed_year": 0,
        "seed_pre": seed.get("pre_money", 10000000),
        "series_a": series_a.get("amount", 10000000),
        "series_a_year": 2,
        "series_a_pre": series_a.get("pre_money", 30000000),
        "series_b": series_b.get("amount", 25000000),
        "series_b_year": 4,
        "series_b_pre": series_b.get("pre_money", 75000000),
    }

    # Market sizing
    market = raw_config.get("market", {})
    config["tam"] = {
        "software": 10000,  # $10B
        "hardware": 4000,
        "consumables": 8000,
        "services": 20000,
    }
    config["sam"] = {
        "india": market.get("sam", 1800000000) / 1000000,  # Convert to millions
        "se_asia": market.get("sam", 1800000000) / 1000000 * 0.6,
    }
    config["som"] = {
        "year8_revenue": market.get("som_y8", 104000000)
        / 1000000,  # Convert to millions
    }

    # Customer segmentation
    customers_raw = raw_config.get("customers", {})
    segments_raw = customers_raw.get("segments", {})
    config["customers"] = {
        "total_y8": customers_raw.get("total_y8", 5500),
        "segments": {
            "smb": segments_raw.get("smb", {
                "count": 3850, "pct": 0.70, "arpu": 9300, "churn": 0.17, "gross_margin": 0.70
            }),
            "midmarket": segments_raw.get("midmarket", {
                "count": 1375, "pct": 0.25, "arpu": 37200, "churn": 0.11, "gross_margin": 0.65
            }),
            "enterprise": segments_raw.get("enterprise", {
                "count": 275, "pct": 0.05, "arpu": 62000, "churn": 0.06, "gross_margin": 0.60
            }),
        }
    }

    # Unit economics
    unit_econ_raw = raw_config.get("unit_economics", {})
    config["unit_economics"] = {
        "blended_arpu": unit_econ_raw.get("blended_arpu", 18909),
        "blended_cac": unit_econ_raw.get("blended_cac", 7000),
        "blended_churn": unit_econ_raw.get("blended_churn", 0.15),
        "blended_lifetime": unit_econ_raw.get("blended_lifetime", 6.7),
        "ltv": unit_econ_raw.get("ltv", 82000),
        "ltv_cac_ratio": unit_econ_raw.get("ltv_cac_ratio", 11.7),
        "payback_months": unit_econ_raw.get("payback_months", 6),
        "nrr_target": unit_econ_raw.get("nrr_target", 1.11),
    }

    # Pricing
    pricing_raw = raw_config.get("pricing", {})
    config["pricing"] = {
        "software_per_seat": pricing_raw.get("software_per_seat", 2400),
        "smb_seats": pricing_raw.get("smb_seats", 2.0),
        "midmarket_seats": pricing_raw.get("midmarket_seats", 8.0),
        "enterprise_seats": pricing_raw.get("enterprise_seats", 25.0),
        "midmarket_discount": pricing_raw.get("midmarket_discount", 0.10),
        "enterprise_discount": pricing_raw.get("enterprise_discount", 0.20),
    }

    # Geographic expansion
    geographic_raw = raw_config.get("geographic", {})
    config["geographic"] = {
        "phase1": geographic_raw.get("phase1", {"regions": ["India"], "years": "Y1-Y2"}),
        "phase2": geographic_raw.get("phase2", {"regions": ["India", "SE Asia"], "years": "Y3-Y4"}),
        "phase3": geographic_raw.get("phase3", {"regions": ["India", "SE Asia", "US/EU"], "years": "Y5-Y8"}),
    }

    # Efficiency metrics
    efficiency_raw = raw_config.get("efficiency_metrics", {})
    config["efficiency_metrics"] = {
        "revenue_per_employee_y8": efficiency_raw.get("revenue_per_employee_y8", 274000),
        "burn_multiple_y1_y4": efficiency_raw.get("burn_multiple_y1_y4", 0.8),
        "arr_per_dollar_raised": efficiency_raw.get("arr_per_dollar_raised", 2.74),
        "time_to_profitability": efficiency_raw.get("time_to_profitability", "Y4"),
        "sm_as_pct_revenue_y8": efficiency_raw.get("sm_as_pct_revenue_y8", 0.24),
        "rd_as_pct_revenue_y8": efficiency_raw.get("rd_as_pct_revenue_y8", 0.07),
        "ga_as_pct_revenue_y8": efficiency_raw.get("ga_as_pct_revenue_y8", 0.07),
        "rule_of_40_y8": efficiency_raw.get("rule_of_40_y8", 59),
    }

    return config


def create_default_config() -> Dict[str, Any]:
    """Create default configuration for testing."""
    return {
        "company": "TestCompany",
        "general": {
            "tax_rate": 0.25,
            "capex_y0": 150000,
            "capex_annual": 50000,
            "depreciation_years": 5,
            "debtor_days": 45,
            "creditor_days": 30,
            "interest_rate": 0.10,
            "cost_inflation": 0.05,
        },
        "revenue_streams": [
            {
                "name": "Software",
                "price": 2500,
                "volume": 25,
                "growth": 0.50,
                "cogs_pct": 0.15,
            },
            {
                "name": "Hardware",
                "price": 5000,
                "volume": 15,
                "growth": 0.35,
                "cogs_pct": 0.60,
            },
            {
                "name": "Services",
                "price": 10000,
                "volume": 5,
                "growth": 0.40,
                "cogs_pct": 0.45,
            },
        ],
        "fixed_costs": {
            "Office & Utilities": 36000,
            "Marketing": 60000,
            "R&D": 72000,
            "Legal & Compliance": 24000,
            "Insurance": 12000,
        },
        "headcount": {
            "engineering_salary": 80000,
            "engineering_y0": 5,
            "sales_salary": 60000,
            "sales_y0": 3,
            "ops_salary": 50000,
            "ops_y0": 2,
            "ga_salary": 70000,
            "ga_y0": 2,
        },
        "funding": {
            "seed": 3000000,
            "seed_year": 0,
            "seed_pre": 10000000,
            "series_a": 10000000,
            "series_a_year": 2,
            "series_a_pre": 30000000,
            "series_b": 25000000,
            "series_b_year": 4,
            "series_b_pre": 75000000,
        },
        "tam": {
            "software": 10000,
            "hardware": 4000,
            "consumables": 8000,
            "services": 20000,
        },
        "sam": {"india": 1800, "se_asia": 1080},
        "som": {"year8_revenue": 104},
    }


def main():
    parser = argparse.ArgumentParser(
        description="Build complete 14-sheet financial model",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  python build_complete_financial_model.py --config .tmp/rapidtools/config/rapidtools_config.json
  python build_complete_financial_model.py --company "TestCompany" --years 8
  python build_complete_financial_model.py --company "RapidTools" --output .tmp/rapidtools_model.xlsx

The script creates all 14 sheets with proper cross-sheet formulas:
  1. Sources & References    9. Summary
  2. Assumptions            10. Sensitivity Analysis
  3. Headcount Plan         11. Valuation
  4. Revenue                12. Break-even Analysis
  5. Operating Costs        13. Funding Cap Table
  6. P&L                    14. Charts Data
  7. Cash Flow
  8. Balance Sheet
        """,
    )
    parser.add_argument("--config", "-c", help="Path to JSON config file")
    parser.add_argument("--company", help="Company name (for default config)")
    parser.add_argument(
        "--years",
        type=int,
        default=11,
        help="Number of years (default: 11 = Year 0 to Year 10)",
    )
    parser.add_argument("--output", "-o", help="Output Excel file path")
    parser.add_argument(
        "--validate", "-v", action="store_true", help="Validate after creation"
    )

    args = parser.parse_args()

    # Load or create config
    if args.config:
        config = load_config(args.config)
        # Handle both flat and nested company config
        company_data = config.get("company", {})
        if isinstance(company_data, dict):
            company = company_data.get(
                "name", company_data.get("product", "FinancialModel")
            )
        else:
            company = company_data or "FinancialModel"
    elif args.company:
        config = create_default_config()
        config["company"] = args.company
        company = args.company
    else:
        print("ERROR: Provide --config or --company")
        sys.exit(1)

    # Determine output path
    if args.output:
        output_path = args.output
    else:
        os.makedirs(".tmp", exist_ok=True)
        output_path = f'.tmp/{company.replace(" ", "_")}_complete_model.xlsx'

    # Build the model
    builder = FinancialModelBuilder(config, num_years=args.years)
    builder.build()
    builder.save(output_path)

    # Validate if requested
    if args.validate:
        print("\nValidating formulas...")
        try:
            from validate_excel_model import validate_excel_model

            success, report = validate_excel_model(output_path, verbose=True)
            print(report)
            if not success:
                sys.exit(1)
        except ImportError:
            print("WARNING: validate_excel_model not available")

    print(f"\nNext steps:")
    print(f"  1. Review in Excel: start {output_path}")
    print(f"  2. Validate formulas:")
    print(f"     python execution/validate_excel_model.py --file {output_path}")
    print(f"  3. Upload to Google Sheets:")
    print(f"     python execution/sync_to_cloud.py --file {output_path}")


if __name__ == "__main__":
    main()
