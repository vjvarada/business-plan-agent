#!/usr/bin/env python3
"""
Clean Financial Model Builder (Config-Driven, Zero Hardcoded Data)
==================================================================

Creates a 14-sheet financial model purely from a JSON config file.
No business-specific data is hardcoded in this script.

Reference: directives/SHEET_BUILD_GUIDE.md

Usage:
    python execution/build_financial_model.py --config config.json
    python execution/build_financial_model.py --config config.json --up-to "P&L"
    python execution/build_financial_model.py --config config.json --sheets "Assumptions" "Revenue"
"""

import argparse
import json
import os
import sys
from typing import Any, Dict, List, Optional

try:
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


# =============================================================================
# COLOR PALETTE (matches directives/SHEET_BUILD_GUIDE.md)
# =============================================================================
class Colors:
    TITLE_BLUE = "335080"
    DARK_BLUE = "336699"
    MEDIUM_BLUE = "6699CC"
    SECTION_A_CAT = "4D80B3"
    LIGHT_BLUE = "D8EAF9"
    LIGHT_GRAY = "F2F2F2"
    GREEN = "E5F8E5"
    WHITE = "FFFFFF"
    BLACK = "000000"
    URL_BLUE = "1A4CB3"
    GRAY = "808080"


# =============================================================================
# STYLING HELPERS
# =============================================================================
THIN_BORDER = Border(
    left=Side(style="thin"), right=Side(style="thin"),
    top=Side(style="thin"), bottom=Side(style="thin"),
)


def style_title(cell, text, bg=Colors.TITLE_BLUE):
    cell.value = text
    cell.font = Font(name="Calibri", size=14, bold=True, color=Colors.WHITE)
    cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center")


def style_header(cell, bg=Colors.DARK_BLUE, fg=Colors.WHITE, size=12, bold=True):
    cell.font = Font(name="Calibri", size=size, bold=bold, color=fg)
    cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = THIN_BORDER


def style_section_header(cell, text, bg=Colors.DARK_BLUE):
    cell.value = text
    style_header(cell, bg=bg, size=11)


def style_column_headers(ws, row, headers, bg=Colors.MEDIUM_BLUE, end_col=10):
    for col, header in enumerate(headers, 1):
        c = ws.cell(row, col)
        c.value = header
        style_header(c, bg=bg, size=10)
    for col in range(len(headers) + 1, end_col + 1):
        ws.cell(row, col).fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")


def style_category_header(ws, row, text, end_col=10):
    ws.cell(row, 1).value = text
    ws.cell(row, 1).font = Font(name="Calibri", size=10, bold=True, color=Colors.BLACK)
    for col in range(1, end_col + 1):
        ws.cell(row, col).fill = PatternFill(
            start_color=Colors.LIGHT_BLUE, end_color=Colors.LIGHT_BLUE, fill_type="solid"
        )


def style_total_row(ws, row, start_col, end_col):
    for col in range(start_col, end_col + 1):
        ws.cell(row, col).fill = PatternFill(
            start_color=Colors.GREEN, end_color=Colors.GREEN, fill_type="solid"
        )


def year_headers(n=11, starting_year=None):
    """Generate year column headers. Uses actual years if starting_year is set."""
    if starting_year:
        return [str(starting_year + i) for i in range(n)]
    return [f"Year {i}" for i in range(n)]


# =============================================================================
# CONFIG NORMALIZER
# =============================================================================
def normalize_config(raw: Dict[str, Any]) -> Dict[str, Any]:
    """Normalize various config formats into a canonical structure."""
    cfg = dict(raw)

    # Company name
    if isinstance(cfg.get("company"), dict):
        cfg["company_name"] = cfg["company"].get("name", "Company")
    elif "company_name" not in cfg:
        cfg["company_name"] = cfg.get("company", "Company")

    # General defaults
    defaults = {
        "tax_rate": 0.25, "capex_y0": 150000, "capex_annual": 50000,
        "depreciation_years": 5, "debtor_days": 45, "creditor_days": 30,
        "interest_rate": 0.10, "cost_inflation": 0.05,
    }
    gen = cfg.get("general", {})
    if not isinstance(gen, dict):
        gen = {}
    for k, v in defaults.items():
        gen.setdefault(k, v)
    cfg["general"] = gen

    # Currency (default USD)
    cfg.setdefault("currency", gen.get("currency", "USD"))

    # Starting year (default current year + 1)
    if "starting_year" not in cfg:
        from datetime import datetime
        cfg["starting_year"] = gen.get("starting_year", datetime.now().year + 1)

    # Tax rate at top level
    if "tax_rate" not in cfg:
        cfg["tax_rate"] = gen["tax_rate"]

    # Revenue streams: ensure list of dicts with required keys
    streams = cfg.get("revenue_streams", [])
    for s in streams:
        s.setdefault("name", "Product")
        s.setdefault("price", 1000)
        s.setdefault("volume", 10)
        s.setdefault("growth", 0.25)
        s.setdefault("cogs_pct", 0.30)
    cfg["revenue_streams"] = streams

    # Fixed costs: normalize to list of {name, annual_cost}
    fc = cfg.get("fixed_costs", [])
    if isinstance(fc, dict):
        fc = [{"name": k, "annual_cost": v} for k, v in fc.items()]
    elif isinstance(fc, list):
        normalized = []
        for item in fc:
            if isinstance(item, dict):
                name = item.get("name", item.get("category", "Cost"))
                cost = item.get("annual_cost", item.get("amount", item.get("value", 0)))
                normalized.append({"name": name, "annual_cost": cost})
        fc = normalized
    cfg["fixed_costs"] = fc

    # Headcount: normalize to departments list
    hc = cfg.get("headcount", {})
    if "departments" not in hc:
        # Legacy flat format: detect all <prefix>_salary keys dynamically
        known_prefixes = [
            ("engineering", "Engineering"), ("sales", "Sales & Marketing"),
            ("ops", "Operations"), ("ga", "G&A"), ("product", "Product"),
            ("data", "Data Science"), ("support", "Customer Support"),
            ("marketing", "Marketing"), ("finance", "Finance"),
            ("hr", "Human Resources"), ("design", "Design"),
            ("devops", "DevOps"), ("research", "Research"),
        ]
        # Also detect any unknown prefixes from _salary keys
        detected = set()
        for key in hc:
            if key.endswith("_salary"):
                pfx = key.replace("_salary", "")
                detected.add(pfx)

        depts = []
        for prefix, label in known_prefixes:
            salary = hc.get(f"{prefix}_salary")
            y0 = hc.get(f"{prefix}_y0")
            if salary is not None or y0 is not None:
                growth = hc.get(f"{prefix}_growth", 0.30)
                depts.append({"name": label, "salary": salary or 60000,
                              "y0_count": y0 or 0, "growth": growth})
                detected.discard(prefix)

        # Handle unknown prefixes
        for pfx in sorted(detected):
            salary = hc.get(f"{pfx}_salary", 60000)
            y0 = hc.get(f"{pfx}_y0", 0)
            growth = hc.get(f"{pfx}_growth", 0.30)
            label = pfx.replace("_", " ").title()
            if y0 > 0 or salary > 0:
                depts.append({"name": label, "salary": salary, "y0_count": y0, "growth": growth})

        if not depts:
            depts = [{"name": "Team", "salary": 60000, "y0_count": 5, "growth": 0.30}]
        hc["departments"] = depts
    cfg["headcount"] = hc

    # Funding: normalize to rounds list
    fund = cfg.get("funding", {})
    if "rounds" not in fund:
        rounds = []
        for key, label in [("seed", "Seed"), ("series_a", "Series A"), ("series_b", "Series B")]:
            amt = fund.get(key, 0)
            yr = fund.get(f"{key}_year", 0)
            pre = fund.get(f"{key}_pre", amt * 3)
            if amt > 0:
                rounds.append({"name": label, "amount": amt, "year": yr, "pre_money": pre})
        fund["rounds"] = rounds
    cfg["funding"] = fund

    # TAM: normalize to streams list
    tam = cfg.get("tam", {})
    if "streams" not in tam:
        streams_tam = []
        for k, v in tam.items():
            if isinstance(v, (int, float)):
                streams_tam.append({"name": k.title(), "value_m": v, "source": "", "confidence": "MEDIUM"})
        tam["streams"] = streams_tam
    cfg["tam"] = tam

    # SAM: normalize to regions list
    sam = cfg.get("sam", {})
    if "regions" not in sam:
        regions = []
        for k, v in sam.items():
            if isinstance(v, (int, float)):
                regions.append({"name": k.replace("_", " ").title(), "value_m": v, "years": ""})
        sam["regions"] = regions
    cfg["sam"] = sam

    # SOM defaults (accept terminal_revenue_m or year8_revenue_m)
    som = cfg.get("som", {})
    if "year8_revenue_m" not in som:
        som["year8_revenue_m"] = som.get("terminal_revenue_m", som.get("year8_revenue", 0))
    cfg["som"] = som

    # Valuation defaults
    val = cfg.get("valuation", {})
    val.setdefault("wacc", 0.15)
    val.setdefault("terminal_growth", 0.03)
    val.setdefault("exit_multiple", 5.0)
    val.setdefault("terminal_year", 8)  # 0-indexed year offset for valuation/sensitivity
    cfg["valuation"] = val

    # Customer acquisition defaults
    ca = cfg.get("customer_acquisition", {})
    ca.setdefault("cac", 0)
    ca.setdefault("churn_rate", 0.05)
    ca.setdefault("new_customers_y0", 0)
    cfg["customer_acquisition"] = ca

    return cfg


# =============================================================================
# SHEET SEQUENCE
# =============================================================================
SHEET_SEQUENCE = [
    "Sources & References",
    "Assumptions",
    "Headcount Plan",
    "Revenue",
    "Operating Costs",
    "P&L",
    "Cash Flow",
    "Balance Sheet",
    "Summary",
    "Sensitivity Analysis",
    "Valuation",
    "Break-even Analysis",
    "Funding Cap Table",
    "Charts Data",
]


# =============================================================================
# FINANCIAL MODEL BUILDER
# =============================================================================
class FinancialModelBuilder:
    """Builds a 14-sheet financial model from config. Zero hardcoded data."""

    def __init__(self, config: Dict[str, Any], num_years: int = 11):
        self.config = normalize_config(config)
        self.num_years = num_years
        self.wb = Workbook()
        if "Sheet" in self.wb.sheetnames:
            del self.wb["Sheet"]
        self.row_refs: Dict[str, int] = {}
        self._end_col = 2 + num_years  # last year column index
        self.currency = self.config.get("currency", "USD")
        self.starting_year = self.config.get("starting_year")
        # Terminal year for valuation/sensitivity (0-indexed year offset)
        # Defaults to year 8, but capped at last available year
        val_cfg = self.config.get("valuation", {})
        self._terminal_yr = min(val_cfg.get("terminal_year", 8), num_years - 1)

    def _year_headers(self):
        """Year column headers using starting_year from config."""
        return year_headers(self.num_years, self.starting_year)

    # ---- public API ----

    def build_all(self) -> "Workbook":
        """Build all 14 sheets in dependency order."""
        print(f"Building 14-sheet financial model for: {self.config['company_name']}")
        print("=" * 60)
        for sheet_name in SHEET_SEQUENCE:
            self._build_sheet(sheet_name)
        print("=" * 60)
        print(f"Created {len(self.wb.sheetnames)} sheets")
        return self.wb

    def build_up_to(self, target: str) -> "Workbook":
        """Build sheets up to and including target."""
        if target not in SHEET_SEQUENCE:
            print(f"ERROR: Unknown sheet '{target}'. Valid: {SHEET_SEQUENCE}")
            sys.exit(1)
        idx = SHEET_SEQUENCE.index(target)
        sheets_to_build = SHEET_SEQUENCE[: idx + 1]
        print(f"Building sheets 1-{idx+1} (up to '{target}')...")
        print("=" * 60)
        for name in sheets_to_build:
            self._build_sheet(name)
        print("=" * 60)
        print(f"Created {len(self.wb.sheetnames)} sheets")
        return self.wb

    def build_sheets(self, names: List[str]) -> "Workbook":
        """Build specific sheets (must respect dependency order)."""
        for name in names:
            if name not in SHEET_SEQUENCE:
                print(f"ERROR: Unknown sheet '{name}'.")
                sys.exit(1)
        # Always build in order
        ordered = [s for s in SHEET_SEQUENCE if s in names]
        print(f"Building {len(ordered)} sheets...")
        for name in ordered:
            self._build_sheet(name)
        print(f"Created {len(self.wb.sheetnames)} sheets")
        return self.wb

    def save(self, filepath: str) -> str:
        os.makedirs(os.path.dirname(filepath) or ".", exist_ok=True)
        self.wb.save(filepath)
        print(f"\nSaved: {filepath}")
        return filepath

    # ---- internal dispatch ----

    def _build_sheet(self, name: str):
        dispatch = {
            "Sources & References": self._build_sources,
            "Assumptions": self._build_assumptions,
            "Headcount Plan": self._build_headcount,
            "Revenue": self._build_revenue,
            "Operating Costs": self._build_operating_costs,
            "P&L": self._build_pnl,
            "Cash Flow": self._build_cash_flow,
            "Balance Sheet": self._build_balance_sheet,
            "Summary": self._build_summary,
            "Sensitivity Analysis": self._build_sensitivity,
            "Valuation": self._build_valuation,
            "Break-even Analysis": self._build_breakeven,
            "Funding Cap Table": self._build_cap_table,
            "Charts Data": self._build_charts_data,
        }
        fn = dispatch.get(name)
        if fn:
            fn()
        else:
            print(f"  WARNING: No builder for '{name}'")

    # ---- helpers ----

    def _set_col_widths(self, ws, label_width=30, val_width=15, year_width=14):
        ws.column_dimensions["A"].width = label_width
        ws.column_dimensions["B"].width = val_width
        for i in range(3, 3 + self.num_years):
            ws.column_dimensions[get_column_letter(i)].width = year_width

    # ================================================================
    # SHEET 1: Sources & References
    # ================================================================
    def _build_sources(self):
        print("  [1/14] Sources & References...")
        ws = self.wb.create_sheet("Sources & References")
        ws.column_dimensions["A"].width = 48
        for c in "BCDE":
            ws.column_dimensions[c].width = 15

        style_title(ws["A1"], "SOURCES & REFERENCES")
        ws.merge_cells("A1:E1")
        row = 3

        # --- SECTION A: KEY METRICS ---
        style_section_header(ws.cell(row, 1), "SECTION A: KEY METRICS")
        ws.merge_cells(f"A{row}:E{row}")
        row += 2

        # TAM
        style_section_header(ws.cell(row, 1), "TAM \u2014 TOTAL ADDRESSABLE MARKET", bg=Colors.SECTION_A_CAT)
        ws.merge_cells(f"A{row}:E{row}")
        row += 1
        style_column_headers(ws, row, ["Revenue Stream", "Value ($M)", "Source", "Confidence", "Notes"], end_col=5)
        row += 1

        tam_streams = self.config.get("tam", {}).get("streams", [])
        tam_start = row
        for ts in tam_streams:
            ws.cell(row, 1).value = ts.get("name", "")
            ws.cell(row, 1).font = Font(size=10)
            ws.cell(row, 2).value = ts.get("value_m", 0)
            ws.cell(row, 2).number_format = "#,##0"
            ws.cell(row, 3).value = ts.get("source", "")
            ws.cell(row, 4).value = ts.get("confidence", "MEDIUM")
            ws.cell(row, 5).value = ts.get("notes", "")
            row += 1
        tam_end = row - 1

        ws.cell(row, 1).value = "TOTAL TAM"
        ws.cell(row, 1).font = Font(bold=True, size=10)
        if tam_start <= tam_end:
            ws.cell(row, 2).value = f"=SUM(B{tam_start}:B{tam_end})"
        else:
            ws.cell(row, 2).value = 0
        ws.cell(row, 2).number_format = "#,##0"
        ws.cell(row, 2).font = Font(bold=True)
        style_total_row(ws, row, 1, 5)
        self.row_refs["tam_total"] = row
        row += 2

        # SAM
        style_section_header(ws.cell(row, 1), "SAM \u2014 SERVICEABLE ADDRESSABLE MARKET", bg=Colors.SECTION_A_CAT)
        ws.merge_cells(f"A{row}:E{row}")
        row += 1
        style_column_headers(ws, row, ["Region", "SAM ($M)", "Years", "Notes", "Source"], end_col=5)
        row += 1

        sam_regions = self.config.get("sam", {}).get("regions", [])
        sam_start = row
        for sr in sam_regions:
            ws.cell(row, 1).value = sr.get("name", "")
            ws.cell(row, 1).font = Font(size=10)
            ws.cell(row, 2).value = sr.get("value_m", 0)
            ws.cell(row, 2).number_format = "#,##0"
            ws.cell(row, 3).value = sr.get("years", "")
            row += 1
        sam_end = row - 1

        ws.cell(row, 1).value = "TOTAL SAM"
        ws.cell(row, 1).font = Font(bold=True, size=10)
        if sam_start <= sam_end:
            ws.cell(row, 2).value = f"=SUM(B{sam_start}:B{sam_end})"
        else:
            ws.cell(row, 2).value = 0
        ws.cell(row, 2).number_format = "#,##0"
        ws.cell(row, 2).font = Font(bold=True)
        style_total_row(ws, row, 1, 5)
        self.row_refs["sam_total"] = row
        row += 2

        # SOM
        style_section_header(ws.cell(row, 1), "SOM \u2014 SERVICEABLE OBTAINABLE MARKET", bg=Colors.SECTION_A_CAT)
        ws.merge_cells(f"A{row}:E{row}")
        row += 1
        som = self.config.get("som", {})
        terminal_label = f"Year {self._terminal_yr}"
        ws.cell(row, 1).value = f"{terminal_label} Revenue Target"
        ws.cell(row, 1).font = Font(size=10)
        ws.cell(row, 2).value = som.get("year8_revenue_m", som.get("terminal_revenue_m", 0))
        ws.cell(row, 2).number_format = "#,##0"
        ws.cell(row, 3).value = f"{self.currency} M"
        self.row_refs["som_revenue"] = row
        row += 1

        ws.cell(row, 1).value = "SAM Penetration"
        ws.cell(row, 1).font = Font(size=10, bold=True)
        sam_total_row = self.row_refs.get("sam_total", row - 3)
        ws.cell(row, 2).value = f"=IF(B{sam_total_row}=0,0,B{row-1}/B{sam_total_row})"
        ws.cell(row, 2).number_format = "0.00%"
        row += 2

        # --- SECTION B: SOURCE DOCUMENTATION ---
        style_section_header(ws.cell(row, 1), "SECTION B: SOURCE DOCUMENTATION")
        ws.merge_cells(f"A{row}:E{row}")
        row += 2
        style_column_headers(ws, row, ["Ref#", "Source", "Data Point", "Value", "URL"], end_col=5)
        row += 1

        # Placeholder rows for research citations
        for i in range(1, 6):
            ws.cell(row, 1).value = f"[{i}]"
            ws.cell(row, 1).font = Font(size=10, color=Colors.GRAY)
            ws.cell(row, 2).value = "(to be populated during research)"
            ws.cell(row, 2).font = Font(size=10, italic=True, color=Colors.GRAY)
            row += 1

    # ================================================================
    # SHEET 2: Assumptions
    # ================================================================
    def _build_assumptions(self):
        print("  [2/14] Assumptions...")
        ws = self.wb.create_sheet("Assumptions")
        self._set_col_widths(ws, label_width=35)

        style_title(ws["A1"], "FINANCIAL MODEL ASSUMPTIONS")
        ws.merge_cells(f"A1:{get_column_letter(self._end_col)}1")
        row = 3

        # --- General Parameters ---
        style_section_header(ws.cell(row, 1), "GENERAL PARAMETERS")
        ws.merge_cells(f"A{row}:C{row}")
        row += 1

        gen = self.config["general"]
        cur = self.currency
        params = [
            ("Tax Rate", gen["tax_rate"], "%"),
            ("CapEx Year 0", gen["capex_y0"], cur),
            ("CapEx Annual", gen["capex_annual"], cur),
            ("Depreciation Period", gen["depreciation_years"], "Years"),
            ("Debtor Days", gen["debtor_days"], "Days"),
            ("Creditor Days", gen["creditor_days"], "Days"),
            ("Interest Rate", gen["interest_rate"], "%"),
            ("Cost Inflation", gen["cost_inflation"], "%"),
        ]

        self.row_refs["assumptions_start"] = row
        for name, value, unit in params:
            ws.cell(row, 1).value = name
            ws.cell(row, 2).value = value
            ws.cell(row, 3).value = unit
            if unit == "%":
                ws.cell(row, 2).number_format = "0.00%"
            elif unit == cur:
                ws.cell(row, 2).number_format = "#,##0"
            # Track key rows
            key_map = {"Tax Rate": "tax_rate", "Depreciation Period": "depreciation_years",
                       "Cost Inflation": "cost_inflation"}
            if name in key_map:
                self.row_refs[key_map[name]] = row
            row += 1
        row += 1

        # --- Revenue Streams ---
        style_section_header(ws.cell(row, 1), "REVENUE STREAMS")
        ws.merge_cells(f"A{row}:{get_column_letter(self._end_col)}{row}")
        row += 1

        headers = ["Stream", "Price", "Unit"] + self._year_headers()
        style_column_headers(ws, row, headers, end_col=self._end_col)
        row += 1

        self.row_refs["revenue_streams_start"] = row
        for stream in self.config["revenue_streams"]:
            name = stream["name"]
            price = stream["price"]
            volume = stream["volume"]
            growth = stream["growth"]
            cogs = stream["cogs_pct"]

            # Price row
            ws.cell(row, 1).value = f"{name}: Price"
            ws.cell(row, 2).value = price
            ws.cell(row, 2).number_format = "#,##0"
            ws.cell(row, 3).value = self.currency
            row += 1

            # Volume row (with growth formulas)
            ws.cell(row, 1).value = f"{name}: Volume"
            ws.cell(row, 3).value = "Units"
            for yr in range(self.num_years):
                col = 4 + yr
                if yr == 0:
                    ws.cell(row, col).value = volume
                else:
                    prev = get_column_letter(col - 1)
                    ws.cell(row, col).value = f"=ROUND({prev}{row}*(1+{growth}),0)"
                ws.cell(row, col).number_format = "#,##0"
            row += 1

            # Growth row
            ws.cell(row, 1).value = f"{name}: Growth"
            ws.cell(row, 2).value = growth
            ws.cell(row, 2).number_format = "0.0%"
            ws.cell(row, 3).value = "%"
            row += 1

            # COGS % row
            ws.cell(row, 1).value = f"{name}: COGS %"
            ws.cell(row, 2).value = cogs
            ws.cell(row, 2).number_format = "0.0%"
            ws.cell(row, 3).value = "%"
            row += 1

        self.row_refs["revenue_streams_end"] = row - 1
        row += 1

        # --- Fixed Costs ---
        style_section_header(ws.cell(row, 1), "FIXED COSTS (Annual)")
        ws.merge_cells(f"A{row}:C{row}")
        row += 1

        self.row_refs["fixed_costs_start"] = row
        for cost in self.config["fixed_costs"]:
            ws.cell(row, 1).value = cost["name"]
            ws.cell(row, 2).value = cost["annual_cost"]
            ws.cell(row, 2).number_format = "#,##0"
            ws.cell(row, 3).value = self.currency
            row += 1
        self.row_refs["fixed_costs_end"] = row - 1
        row += 1

        # --- Funding Parameters ---
        style_section_header(ws.cell(row, 1), "FUNDING PARAMETERS")
        ws.merge_cells(f"A{row}:C{row}")
        row += 1

        self.row_refs["funding_start"] = row
        for rnd in self.config["funding"].get("rounds", []):
            ws.cell(row, 1).value = f"{rnd['name']} Amount"
            ws.cell(row, 2).value = rnd["amount"]
            ws.cell(row, 2).number_format = "#,##0"
            ws.cell(row, 3).value = self.currency
            row += 1
            ws.cell(row, 1).value = f"{rnd['name']} Timing"
            ws.cell(row, 2).value = rnd.get("year", 0)
            ws.cell(row, 3).value = "Year"
            row += 1
            ws.cell(row, 1).value = f"{rnd['name']} Pre-Money"
            ws.cell(row, 2).value = rnd.get("pre_money", 0)
            ws.cell(row, 2).number_format = "#,##0"
            ws.cell(row, 3).value = self.currency
            row += 1
        self.row_refs["funding_end"] = row - 1
        row += 1

        # --- Customer Acquisition ---
        ca = self.config.get("customer_acquisition", {})
        if ca.get("cac") or ca.get("churn_rate") or ca.get("new_customers_y0"):
            style_section_header(ws.cell(row, 1), "CUSTOMER ACQUISITION")
            ws.merge_cells(f"A{row}:C{row}")
            row += 1

            self.row_refs["ca_start"] = row
            ca_params = [
                ("CAC (Cost per Customer)", ca.get("cac", 0), self.currency),
                ("Churn Rate (Annual)", ca.get("churn_rate", 0.05), "%"),
                ("New Customers Year 0", ca.get("new_customers_y0", 0), "Count"),
                ("Customer Growth Rate", ca.get("customer_growth", 0.25), "%"),
            ]
            for name, value, unit in ca_params:
                ws.cell(row, 1).value = name
                ws.cell(row, 2).value = value
                if unit == "%":
                    ws.cell(row, 2).number_format = "0.00%"
                elif unit == self.currency:
                    ws.cell(row, 2).number_format = "#,##0"
                ws.cell(row, 3).value = unit
                row += 1
            self.row_refs["ca_end"] = row - 1

    # ================================================================
    # SHEET 3: Headcount Plan
    # ================================================================
    def _build_headcount(self):
        print("  [3/14] Headcount Plan...")
        ws = self.wb.create_sheet("Headcount Plan")
        self._set_col_widths(ws)

        style_title(ws["A1"], "HEADCOUNT PLAN")
        ws.merge_cells(f"A1:{get_column_letter(self._end_col)}1")
        row = 3

        headers = ["Department", "Avg Salary"] + self._year_headers()
        style_column_headers(ws, row, headers, end_col=self._end_col)
        row += 1

        depts = self.config["headcount"]["departments"]
        self.row_refs["headcount_start"] = row

        for dept in depts:
            name = dept["name"]
            salary = dept["salary"]
            y0 = dept["y0_count"]
            growth = dept.get("growth", 0.30)

            ws.cell(row, 1).value = name
            ws.cell(row, 2).value = salary
            ws.cell(row, 2).number_format = "#,##0"
            for yr in range(self.num_years):
                col = 3 + yr
                if yr == 0:
                    ws.cell(row, col).value = y0
                else:
                    prev = get_column_letter(col - 1)
                    ws.cell(row, col).value = f"=ROUND({prev}{row}*(1+{growth}),0)"
                ws.cell(row, col).number_format = "#,##0"
            row += 1

        self.row_refs["headcount_end"] = row - 1

        # Total headcount
        ws.cell(row, 1).value = "TOTAL HEADCOUNT"
        ws.cell(row, 1).font = Font(bold=True)
        for yr in range(self.num_years):
            col = 3 + yr
            cl = get_column_letter(col)
            ws.cell(row, col).value = f'=SUM({cl}{self.row_refs["headcount_start"]}:{cl}{self.row_refs["headcount_end"]})'
            ws.cell(row, col).number_format = "#,##0"
            ws.cell(row, col).font = Font(bold=True)
        style_total_row(ws, row, 1, self._end_col)
        self.row_refs["headcount_total"] = row
        row += 2

        # Total salary cost
        ws.cell(row, 1).value = "TOTAL SALARY COST"
        ws.cell(row, 1).font = Font(bold=True)
        for yr in range(self.num_years):
            col = 3 + yr
            cl = get_column_letter(col)
            parts = []
            for dr in range(self.row_refs["headcount_start"], self.row_refs["headcount_end"] + 1):
                parts.append(f"({cl}{dr}*B{dr})")
            ws.cell(row, col).value = f'={"+".join(parts)}'
            ws.cell(row, col).number_format = "#,##0"
            ws.cell(row, col).font = Font(bold=True)
        style_total_row(ws, row, 1, self._end_col)
        self.row_refs["salary_cost_total"] = row

    # ================================================================
    # SHEET 4: Revenue
    # ================================================================
    def _build_revenue(self):
        print("  [4/14] Revenue...")
        ws = self.wb.create_sheet("Revenue")
        self._set_col_widths(ws)

        style_title(ws["A1"], "REVENUE PROJECTIONS")
        ws.merge_cells(f"A1:{get_column_letter(self._end_col)}1")
        row = 3

        headers = ["Revenue Stream", "Unit"] + self._year_headers()
        style_column_headers(ws, row, headers, end_col=self._end_col)
        row += 1

        streams = self.config["revenue_streams"]
        self.row_refs["revenue_start"] = row
        ass_start = self.row_refs.get("revenue_streams_start", 17)

        for idx, stream in enumerate(streams):
            ws.cell(row, 1).value = stream["name"]
            ws.cell(row, 2).value = self.currency
            price_row = ass_start + (idx * 4)
            volume_row = price_row + 1
            for yr in range(self.num_years):
                col = 3 + yr
                yr_col = get_column_letter(4 + yr)
                ws.cell(row, col).value = f"=Assumptions!$B${price_row}*Assumptions!{yr_col}${volume_row}"
                ws.cell(row, col).number_format = "#,##0"
            row += 1

        self.row_refs["revenue_end"] = row - 1
        row += 1

        # Total Revenue
        ws.cell(row, 1).value = "TOTAL REVENUE"
        ws.cell(row, 1).font = Font(bold=True, size=11)
        ws.cell(row, 2).value = self.currency
        for yr in range(self.num_years):
            col = 3 + yr
            cl = get_column_letter(col)
            ws.cell(row, col).value = f'=SUM({cl}{self.row_refs["revenue_start"]}:{cl}{self.row_refs["revenue_end"]})'
            ws.cell(row, col).number_format = "#,##0"
            ws.cell(row, col).font = Font(bold=True)
        style_total_row(ws, row, 1, self._end_col)
        self.row_refs["revenue_total"] = row
        row += 2

        # Revenue Mix %
        style_section_header(ws.cell(row, 1), "REVENUE MIX %", bg=Colors.MEDIUM_BLUE)
        ws.merge_cells(f"A{row}:{get_column_letter(self._end_col)}{row}")
        row += 1

        for idx, stream in enumerate(streams):
            ws.cell(row, 1).value = f"{stream['name']} %"
            ws.cell(row, 2).value = "%"
            sr = self.row_refs["revenue_start"] + idx
            for yr in range(self.num_years):
                col = 3 + yr
                cl = get_column_letter(col)
                ws.cell(row, col).value = f'=IF({cl}{self.row_refs["revenue_total"]}=0,0,{cl}{sr}/{cl}{self.row_refs["revenue_total"]})'
                ws.cell(row, col).number_format = "0.0%"
            row += 1

    # ================================================================
    # SHEET 5: Operating Costs
    # ================================================================
    def _build_operating_costs(self):
        print("  [5/14] Operating Costs...")
        ws = self.wb.create_sheet("Operating Costs")
        self._set_col_widths(ws, label_width=35)

        style_title(ws["A1"], "OPERATING COSTS")
        ws.merge_cells(f"A1:{get_column_letter(self._end_col)}1")
        row = 3

        # COGS
        style_section_header(ws.cell(row, 1), "COST OF GOODS SOLD (COGS)")
        ws.merge_cells(f"A{row}:{get_column_letter(self._end_col)}{row}")
        row += 1

        streams = self.config["revenue_streams"]
        self.row_refs["cogs_start"] = row
        ass_start = self.row_refs.get("revenue_streams_start", 17)

        for idx, stream in enumerate(streams):
            ws.cell(row, 1).value = f"COGS: {stream['name']}"
            ws.cell(row, 2).value = self.currency
            cogs_pct_row = ass_start + (idx * 4) + 3
            rev_row = self.row_refs["revenue_start"] + idx
            for yr in range(self.num_years):
                col = 3 + yr
                cl = get_column_letter(col)
                ws.cell(row, col).value = f"=Revenue!{cl}{rev_row}*Assumptions!$B${cogs_pct_row}"
                ws.cell(row, col).number_format = "#,##0"
            row += 1

        self.row_refs["cogs_end"] = row - 1

        # Total COGS
        ws.cell(row, 1).value = "TOTAL COGS"
        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 2).value = self.currency
        for yr in range(self.num_years):
            col = 3 + yr
            cl = get_column_letter(col)
            ws.cell(row, col).value = f'=SUM({cl}{self.row_refs["cogs_start"]}:{cl}{self.row_refs["cogs_end"]})'
            ws.cell(row, col).number_format = "#,##0"
            ws.cell(row, col).font = Font(bold=True)
        style_total_row(ws, row, 1, self._end_col)
        self.row_refs["cogs_total"] = row
        row += 2

        # Fixed Costs
        style_section_header(ws.cell(row, 1), "FIXED COSTS")
        ws.merge_cells(f"A{row}:{get_column_letter(self._end_col)}{row}")
        row += 1

        # Salaries from Headcount
        ws.cell(row, 1).value = "Salaries & Benefits"
        ws.cell(row, 2).value = self.currency
        sal_row = self.row_refs.get("salary_cost_total", 10)
        for yr in range(self.num_years):
            col = 3 + yr
            cl = get_column_letter(col)
            ws.cell(row, col).value = f"='Headcount Plan'!{cl}{sal_row}"
            ws.cell(row, col).number_format = "#,##0"
        self.row_refs["fixed_salaries"] = row
        row += 1

        # Other fixed costs
        self.row_refs["other_fixed_start"] = row
        inflation = self.config["general"]["cost_inflation"]
        for cost in self.config["fixed_costs"]:
            cname = cost["name"]
            if "salary" in cname.lower() or "salaries" in cname.lower():
                continue
            ws.cell(row, 1).value = cname
            ws.cell(row, 2).value = self.currency
            for yr in range(self.num_years):
                col = 3 + yr
                if yr == 0:
                    ws.cell(row, col).value = cost["annual_cost"]
                else:
                    prev = get_column_letter(col - 1)
                    ws.cell(row, col).value = f"={prev}{row}*(1+{inflation})"
                ws.cell(row, col).number_format = "#,##0"
            row += 1
        self.row_refs["other_fixed_end"] = row - 1

        # Total Fixed
        ws.cell(row, 1).value = "TOTAL FIXED COSTS"
        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 2).value = self.currency
        for yr in range(self.num_years):
            col = 3 + yr
            cl = get_column_letter(col)
            ws.cell(row, col).value = f'={cl}{self.row_refs["fixed_salaries"]}+SUM({cl}{self.row_refs["other_fixed_start"]}:{cl}{self.row_refs["other_fixed_end"]})'
            ws.cell(row, col).number_format = "#,##0"
            ws.cell(row, col).font = Font(bold=True)
        style_total_row(ws, row, 1, self._end_col)
        self.row_refs["fixed_total"] = row
        row += 2

        # Total Operating Costs
        ws.cell(row, 1).value = "TOTAL OPERATING COSTS"
        ws.cell(row, 1).font = Font(bold=True, size=11)
        ws.cell(row, 2).value = self.currency
        for yr in range(self.num_years):
            col = 3 + yr
            cl = get_column_letter(col)
            ws.cell(row, col).value = f'={cl}{self.row_refs["cogs_total"]}+{cl}{self.row_refs["fixed_total"]}'
            ws.cell(row, col).number_format = "#,##0"
            ws.cell(row, col).font = Font(bold=True)
        style_total_row(ws, row, 1, self._end_col)
        self.row_refs["opex_total"] = row

    # ================================================================
    # SHEET 6: P&L
    # ================================================================
    def _build_pnl(self):
        print("  [6/14] P&L...")
        ws = self.wb.create_sheet("P&L")
        self._set_col_widths(ws)

        style_title(ws["A1"], "PROFIT & LOSS STATEMENT")
        ws.merge_cells(f"A1:{get_column_letter(self._end_col)}1")
        row = 3

        headers = ["Line Item", "Unit"] + self._year_headers()
        style_column_headers(ws, row, headers, end_col=self._end_col)
        row += 1

        def _ref(sheet, ref_key):
            return self.row_refs.get(ref_key, 4)

        # Revenue
        ws.cell(row, 1).value = "Revenue"
        ws.cell(row, 2).value = self.currency
        for yr in range(self.num_years):
            cl = get_column_letter(3 + yr)
            ws.cell(row, 3 + yr).value = f"=Revenue!{cl}{_ref('Revenue', 'revenue_total')}"
            ws.cell(row, 3 + yr).number_format = "#,##0"
        self.row_refs["pnl_revenue"] = row
        row += 1

        # COGS
        ws.cell(row, 1).value = "Cost of Goods Sold"
        ws.cell(row, 2).value = self.currency
        for yr in range(self.num_years):
            cl = get_column_letter(3 + yr)
            ws.cell(row, 3 + yr).value = f"='Operating Costs'!{cl}{_ref('OpCosts', 'cogs_total')}"
            ws.cell(row, 3 + yr).number_format = "#,##0"
        self.row_refs["pnl_cogs"] = row
        row += 1

        # Gross Profit
        ws.cell(row, 1).value = "GROSS PROFIT"
        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 2).value = self.currency
        for yr in range(self.num_years):
            cl = get_column_letter(3 + yr)
            ws.cell(row, 3 + yr).value = f"={cl}{self.row_refs['pnl_revenue']}-{cl}{self.row_refs['pnl_cogs']}"
            ws.cell(row, 3 + yr).number_format = "#,##0"
            ws.cell(row, 3 + yr).font = Font(bold=True)
        self.row_refs["pnl_gross_profit"] = row
        row += 1

        # Gross Margin %
        ws.cell(row, 1).value = "Gross Margin %"
        ws.cell(row, 2).value = "%"
        for yr in range(self.num_years):
            cl = get_column_letter(3 + yr)
            ws.cell(row, 3 + yr).value = f"=IF({cl}{self.row_refs['pnl_revenue']}=0,0,{cl}{self.row_refs['pnl_gross_profit']}/{cl}{self.row_refs['pnl_revenue']})"
            ws.cell(row, 3 + yr).number_format = "0.0%"
        row += 2

        # OpEx
        ws.cell(row, 1).value = "Operating Expenses"
        ws.cell(row, 2).value = self.currency
        for yr in range(self.num_years):
            cl = get_column_letter(3 + yr)
            ws.cell(row, 3 + yr).value = f"='Operating Costs'!{cl}{_ref('OpCosts', 'fixed_total')}"
            ws.cell(row, 3 + yr).number_format = "#,##0"
        self.row_refs["pnl_opex"] = row
        row += 1

        # EBITDA
        ws.cell(row, 1).value = "EBITDA"
        ws.cell(row, 1).font = Font(bold=True, size=11)
        ws.cell(row, 2).value = self.currency
        for yr in range(self.num_years):
            cl = get_column_letter(3 + yr)
            ws.cell(row, 3 + yr).value = f"={cl}{self.row_refs['pnl_gross_profit']}-{cl}{self.row_refs['pnl_opex']}"
            ws.cell(row, 3 + yr).number_format = "#,##0"
            ws.cell(row, 3 + yr).font = Font(bold=True)
        style_total_row(ws, row, 1, self._end_col)
        self.row_refs["pnl_ebitda"] = row
        row += 1

        # EBITDA Margin
        ws.cell(row, 1).value = "EBITDA Margin %"
        ws.cell(row, 2).value = "%"
        for yr in range(self.num_years):
            cl = get_column_letter(3 + yr)
            ws.cell(row, 3 + yr).value = f"=IF({cl}{self.row_refs['pnl_revenue']}=0,0,{cl}{self.row_refs['pnl_ebitda']}/{cl}{self.row_refs['pnl_revenue']})"
            ws.cell(row, 3 + yr).number_format = "0.0%"
        row += 2

        # Depreciation
        gen = self.config["general"]
        capex_y0 = gen["capex_y0"]
        capex_ann = gen["capex_annual"]
        dep_yrs = gen["depreciation_years"]

        ws.cell(row, 1).value = "Depreciation"
        ws.cell(row, 2).value = self.currency
        for yr in range(self.num_years):
            dep = capex_y0 / dep_yrs
            if yr > 0:
                dep += (capex_ann * min(yr, dep_yrs)) / dep_yrs
            ws.cell(row, 3 + yr).value = round(dep)
            ws.cell(row, 3 + yr).number_format = "#,##0"
        self.row_refs["pnl_depreciation"] = row
        row += 1

        # EBIT
        ws.cell(row, 1).value = "EBIT (Operating Profit)"
        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 2).value = self.currency
        for yr in range(self.num_years):
            cl = get_column_letter(3 + yr)
            ws.cell(row, 3 + yr).value = f"={cl}{self.row_refs['pnl_ebitda']}-{cl}{self.row_refs['pnl_depreciation']}"
            ws.cell(row, 3 + yr).number_format = "#,##0"
            ws.cell(row, 3 + yr).font = Font(bold=True)
        self.row_refs["pnl_ebit"] = row
        row += 2

        # Interest
        ws.cell(row, 1).value = "Interest Expense"
        ws.cell(row, 2).value = self.currency
        for yr in range(self.num_years):
            ws.cell(row, 3 + yr).value = 0
            ws.cell(row, 3 + yr).number_format = "#,##0"
        self.row_refs["pnl_interest"] = row
        row += 1

        # PBT
        ws.cell(row, 1).value = "PBT (Profit Before Tax)"
        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 2).value = self.currency
        for yr in range(self.num_years):
            cl = get_column_letter(3 + yr)
            ws.cell(row, 3 + yr).value = f"={cl}{self.row_refs['pnl_ebit']}-{cl}{self.row_refs['pnl_interest']}"
            ws.cell(row, 3 + yr).number_format = "#,##0"
            ws.cell(row, 3 + yr).font = Font(bold=True)
        self.row_refs["pnl_pbt"] = row
        row += 1

        # Tax
        tax_rate = gen["tax_rate"]
        ws.cell(row, 1).value = f"Tax ({int(tax_rate*100)}%)"
        ws.cell(row, 2).value = self.currency
        for yr in range(self.num_years):
            cl = get_column_letter(3 + yr)
            ws.cell(row, 3 + yr).value = f"=MAX(0,{cl}{self.row_refs['pnl_pbt']}*{tax_rate})"
            ws.cell(row, 3 + yr).number_format = "#,##0"
        self.row_refs["pnl_tax"] = row
        row += 1

        # Net Income
        ws.cell(row, 1).value = "NET INCOME (PAT)"
        ws.cell(row, 1).font = Font(bold=True, size=11)
        ws.cell(row, 2).value = self.currency
        for yr in range(self.num_years):
            cl = get_column_letter(3 + yr)
            ws.cell(row, 3 + yr).value = f"={cl}{self.row_refs['pnl_pbt']}-{cl}{self.row_refs['pnl_tax']}"
            ws.cell(row, 3 + yr).number_format = "#,##0"
            ws.cell(row, 3 + yr).font = Font(bold=True)
        style_total_row(ws, row, 1, self._end_col)
        self.row_refs["pnl_net_income"] = row
        row += 1

        # Net Margin
        ws.cell(row, 1).value = "Net Margin %"
        ws.cell(row, 2).value = "%"
        for yr in range(self.num_years):
            cl = get_column_letter(3 + yr)
            ws.cell(row, 3 + yr).value = f"=IF({cl}{self.row_refs['pnl_revenue']}=0,0,{cl}{self.row_refs['pnl_net_income']}/{cl}{self.row_refs['pnl_revenue']})"
            ws.cell(row, 3 + yr).number_format = "0.0%"

    # ================================================================
    # SHEET 7: Cash Flow
    # ================================================================
    def _build_cash_flow(self):
        print("  [7/14] Cash Flow...")
        ws = self.wb.create_sheet("Cash Flow")
        self._set_col_widths(ws, label_width=35)

        style_title(ws["A1"], "CASH FLOW STATEMENT")
        ws.merge_cells(f"A1:{get_column_letter(self._end_col)}1")
        row = 3

        headers = ["Line Item", "Unit"] + self._year_headers()
        style_column_headers(ws, row, headers, end_col=self._end_col)
        row += 1

        # Operating
        style_section_header(ws.cell(row, 1), "OPERATING ACTIVITIES", bg=Colors.MEDIUM_BLUE)
        ws.merge_cells(f"A{row}:{get_column_letter(self._end_col)}{row}")
        row += 1

        ws.cell(row, 1).value = "Net Income"
        ws.cell(row, 2).value = self.currency
        for yr in range(self.num_years):
            cl = get_column_letter(3 + yr)
            ws.cell(row, 3 + yr).value = f"='P&L'!{cl}{self.row_refs['pnl_net_income']}"
            ws.cell(row, 3 + yr).number_format = "#,##0"
        self.row_refs["cf_net_income"] = row
        row += 1

        ws.cell(row, 1).value = "+ Depreciation"
        ws.cell(row, 2).value = self.currency
        for yr in range(self.num_years):
            cl = get_column_letter(3 + yr)
            ws.cell(row, 3 + yr).value = f"='P&L'!{cl}{self.row_refs['pnl_depreciation']}"
            ws.cell(row, 3 + yr).number_format = "#,##0"
        self.row_refs["cf_depreciation"] = row
        row += 1

        ws.cell(row, 1).value = "Working Capital Change"
        ws.cell(row, 2).value = self.currency
        for yr in range(self.num_years):
            ws.cell(row, 3 + yr).value = 0
            ws.cell(row, 3 + yr).number_format = "#,##0"
        self.row_refs["cf_wc_change"] = row
        row += 1

        ws.cell(row, 1).value = "Operating Cash Flow"
        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 2).value = self.currency
        for yr in range(self.num_years):
            cl = get_column_letter(3 + yr)
            ws.cell(row, 3 + yr).value = f"={cl}{self.row_refs['cf_net_income']}+{cl}{self.row_refs['cf_depreciation']}-{cl}{self.row_refs['cf_wc_change']}"
            ws.cell(row, 3 + yr).number_format = "#,##0"
            ws.cell(row, 3 + yr).font = Font(bold=True)
        style_total_row(ws, row, 1, self._end_col)
        self.row_refs["cf_operating"] = row
        row += 2

        # Investing
        style_section_header(ws.cell(row, 1), "INVESTING ACTIVITIES", bg=Colors.MEDIUM_BLUE)
        ws.merge_cells(f"A{row}:{get_column_letter(self._end_col)}{row}")
        row += 1

        gen = self.config["general"]
        ws.cell(row, 1).value = "Capital Expenditure"
        ws.cell(row, 2).value = self.currency
        for yr in range(self.num_years):
            capex = gen["capex_y0"] if yr == 0 else gen["capex_annual"]
            ws.cell(row, 3 + yr).value = -capex
            ws.cell(row, 3 + yr).number_format = "#,##0"
        self.row_refs["cf_capex"] = row
        row += 1

        ws.cell(row, 1).value = "Investing Cash Flow"
        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 2).value = self.currency
        for yr in range(self.num_years):
            cl = get_column_letter(3 + yr)
            ws.cell(row, 3 + yr).value = f"={cl}{self.row_refs['cf_capex']}"
            ws.cell(row, 3 + yr).number_format = "#,##0"
            ws.cell(row, 3 + yr).font = Font(bold=True)
        style_total_row(ws, row, 1, self._end_col)
        self.row_refs["cf_investing"] = row
        row += 2

        # Financing
        style_section_header(ws.cell(row, 1), "FINANCING ACTIVITIES", bg=Colors.MEDIUM_BLUE)
        ws.merge_cells(f"A{row}:{get_column_letter(self._end_col)}{row}")
        row += 1

        rounds = self.config["funding"].get("rounds", [])
        ws.cell(row, 1).value = "Equity Raised"
        ws.cell(row, 2).value = self.currency
        for yr in range(self.num_years):
            equity = sum(r["amount"] for r in rounds if r.get("year", -1) == yr)
            ws.cell(row, 3 + yr).value = equity
            ws.cell(row, 3 + yr).number_format = "#,##0"
        self.row_refs["cf_equity"] = row
        row += 1

        ws.cell(row, 1).value = "Financing Cash Flow"
        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 2).value = self.currency
        for yr in range(self.num_years):
            cl = get_column_letter(3 + yr)
            ws.cell(row, 3 + yr).value = f"={cl}{self.row_refs['cf_equity']}"
            ws.cell(row, 3 + yr).number_format = "#,##0"
            ws.cell(row, 3 + yr).font = Font(bold=True)
        style_total_row(ws, row, 1, self._end_col)
        self.row_refs["cf_financing"] = row
        row += 2

        # Net Cash Flow
        ws.cell(row, 1).value = "NET CASH FLOW"
        ws.cell(row, 1).font = Font(bold=True, size=11)
        ws.cell(row, 2).value = self.currency
        for yr in range(self.num_years):
            cl = get_column_letter(3 + yr)
            ws.cell(row, 3 + yr).value = f"={cl}{self.row_refs['cf_operating']}+{cl}{self.row_refs['cf_investing']}+{cl}{self.row_refs['cf_financing']}"
            ws.cell(row, 3 + yr).number_format = "#,##0"
            ws.cell(row, 3 + yr).font = Font(bold=True)
        style_total_row(ws, row, 1, self._end_col)
        self.row_refs["cf_net"] = row
        row += 2

        # Cumulative
        ws.cell(row, 1).value = "CUMULATIVE CASH"
        ws.cell(row, 1).font = Font(bold=True, size=11)
        ws.cell(row, 2).value = self.currency
        for yr in range(self.num_years):
            cl = get_column_letter(3 + yr)
            if yr == 0:
                ws.cell(row, 3 + yr).value = f"={cl}{self.row_refs['cf_net']}"
            else:
                prev = get_column_letter(2 + yr)
                ws.cell(row, 3 + yr).value = f"={prev}{row}+{cl}{self.row_refs['cf_net']}"
            ws.cell(row, 3 + yr).number_format = "#,##0"
            ws.cell(row, 3 + yr).font = Font(bold=True)
        style_total_row(ws, row, 1, self._end_col)
        self.row_refs["cf_cumulative"] = row

    # ================================================================
    # SHEET 8: Balance Sheet
    # ================================================================
    def _build_balance_sheet(self):
        print("  [8/14] Balance Sheet...")
        ws = self.wb.create_sheet("Balance Sheet")
        self._set_col_widths(ws)

        style_title(ws["A1"], "BALANCE SHEET")
        ws.merge_cells(f"A1:{get_column_letter(self._end_col)}1")
        row = 3

        headers = ["Line Item", "Unit"] + self._year_headers()
        style_column_headers(ws, row, headers, end_col=self._end_col)
        row += 1

        # Assets
        style_section_header(ws.cell(row, 1), "ASSETS", bg=Colors.MEDIUM_BLUE)
        ws.merge_cells(f"A{row}:{get_column_letter(self._end_col)}{row}")
        row += 1

        ws.cell(row, 1).value = "Cash & Equivalents"
        ws.cell(row, 2).value = self.currency
        for yr in range(self.num_years):
            cl = get_column_letter(3 + yr)
            ws.cell(row, 3 + yr).value = f"='Cash Flow'!{cl}{self.row_refs['cf_cumulative']}"
            ws.cell(row, 3 + yr).number_format = "#,##0"
        self.row_refs["bs_cash"] = row
        row += 1

        gen = self.config["general"]
        capex_y0 = gen["capex_y0"]
        capex_ann = gen["capex_annual"]
        dep_yrs = gen["depreciation_years"]

        ws.cell(row, 1).value = "Net Fixed Assets"
        ws.cell(row, 2).value = self.currency
        for yr in range(self.num_years):
            cum_capex = capex_y0 + (capex_ann * yr)
            cum_dep = (capex_y0 / dep_yrs) * min(yr + 1, dep_yrs)
            if yr > 0:
                cum_dep += sum((capex_ann / dep_yrs) * min(yr - y, dep_yrs) for y in range(yr))
            ws.cell(row, 3 + yr).value = round(max(0, cum_capex - cum_dep))
            ws.cell(row, 3 + yr).number_format = "#,##0"
        self.row_refs["bs_fixed_assets"] = row
        row += 1

        ws.cell(row, 1).value = "TOTAL ASSETS"
        ws.cell(row, 1).font = Font(bold=True, size=11)
        ws.cell(row, 2).value = self.currency
        for yr in range(self.num_years):
            cl = get_column_letter(3 + yr)
            ws.cell(row, 3 + yr).value = f"={cl}{self.row_refs['bs_cash']}+{cl}{self.row_refs['bs_fixed_assets']}"
            ws.cell(row, 3 + yr).number_format = "#,##0"
            ws.cell(row, 3 + yr).font = Font(bold=True)
        style_total_row(ws, row, 1, self._end_col)
        self.row_refs["bs_total_assets"] = row
        row += 2

        # Liabilities & Equity
        style_section_header(ws.cell(row, 1), "LIABILITIES & EQUITY", bg=Colors.MEDIUM_BLUE)
        ws.merge_cells(f"A{row}:{get_column_letter(self._end_col)}{row}")
        row += 1

        ws.cell(row, 1).value = "Total Liabilities"
        ws.cell(row, 2).value = self.currency
        for yr in range(self.num_years):
            ws.cell(row, 3 + yr).value = 0
            ws.cell(row, 3 + yr).number_format = "#,##0"
        self.row_refs["bs_liabilities"] = row
        row += 2

        rounds = self.config["funding"].get("rounds", [])
        ws.cell(row, 1).value = "Paid-in Capital"
        ws.cell(row, 2).value = self.currency
        for yr in range(self.num_years):
            cum = sum(r["amount"] for r in rounds if r.get("year", 0) <= yr)
            ws.cell(row, 3 + yr).value = cum
            ws.cell(row, 3 + yr).number_format = "#,##0"
        self.row_refs["bs_paid_capital"] = row
        row += 1

        ws.cell(row, 1).value = "Retained Earnings"
        ws.cell(row, 2).value = self.currency
        for yr in range(self.num_years):
            cl = get_column_letter(3 + yr)
            if yr == 0:
                ws.cell(row, 3 + yr).value = f"='P&L'!{cl}{self.row_refs['pnl_net_income']}"
            else:
                prev = get_column_letter(2 + yr)
                ws.cell(row, 3 + yr).value = f"={prev}{row}+'P&L'!{cl}{self.row_refs['pnl_net_income']}"
            ws.cell(row, 3 + yr).number_format = "#,##0"
        self.row_refs["bs_retained_earnings"] = row
        row += 1

        ws.cell(row, 1).value = "Total Equity"
        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 2).value = self.currency
        for yr in range(self.num_years):
            cl = get_column_letter(3 + yr)
            ws.cell(row, 3 + yr).value = f"={cl}{self.row_refs['bs_paid_capital']}+{cl}{self.row_refs['bs_retained_earnings']}"
            ws.cell(row, 3 + yr).number_format = "#,##0"
            ws.cell(row, 3 + yr).font = Font(bold=True)
        self.row_refs["bs_total_equity"] = row
        row += 1

        ws.cell(row, 1).value = "TOTAL LIABILITIES & EQUITY"
        ws.cell(row, 1).font = Font(bold=True, size=11)
        ws.cell(row, 2).value = self.currency
        for yr in range(self.num_years):
            cl = get_column_letter(3 + yr)
            ws.cell(row, 3 + yr).value = f"={cl}{self.row_refs['bs_liabilities']}+{cl}{self.row_refs['bs_total_equity']}"
            ws.cell(row, 3 + yr).number_format = "#,##0"
            ws.cell(row, 3 + yr).font = Font(bold=True)
        style_total_row(ws, row, 1, self._end_col)
        self.row_refs["bs_total_le"] = row
        row += 2

        ws.cell(row, 1).value = "BALANCE CHECK (Assets = L+E)"
        ws.cell(row, 1).font = Font(bold=True, color=Colors.GRAY)
        for yr in range(self.num_years):
            cl = get_column_letter(3 + yr)
            ws.cell(row, 3 + yr).value = f"={cl}{self.row_refs['bs_total_assets']}-{cl}{self.row_refs['bs_total_le']}"
            ws.cell(row, 3 + yr).number_format = "#,##0"
        self.row_refs["bs_check"] = row

    # ================================================================
    # SHEET 9: Summary
    # ================================================================
    def _build_summary(self):
        print("  [9/14] Summary...")
        ws = self.wb.create_sheet("Summary")
        self._set_col_widths(ws, label_width=35)

        style_title(ws["A1"], "KEY PERFORMANCE INDICATORS")
        ws.merge_cells(f"A1:{get_column_letter(self._end_col)}1")
        row = 3

        headers = ["Metric", "Unit"] + self._year_headers()
        style_column_headers(ws, row, headers, end_col=self._end_col)
        row += 1

        # Revenue Metrics
        style_section_header(ws.cell(row, 1), "REVENUE METRICS", bg=Colors.MEDIUM_BLUE)
        ws.merge_cells(f"A{row}:{get_column_letter(self._end_col)}{row}")
        row += 1

        rev_r = self.row_refs["pnl_revenue"]
        gp_r = self.row_refs["pnl_gross_profit"]
        ebitda_r = self.row_refs["pnl_ebitda"]

        ws.cell(row, 1).value = "Total Revenue"
        ws.cell(row, 2).value = self.currency
        for yr in range(self.num_years):
            cl = get_column_letter(3 + yr)
            ws.cell(row, 3 + yr).value = f"='P&L'!{cl}{rev_r}"
            ws.cell(row, 3 + yr).number_format = "#,##0"
        row += 1

        ws.cell(row, 1).value = "Revenue Growth %"
        ws.cell(row, 2).value = "%"
        for yr in range(self.num_years):
            cl = get_column_letter(3 + yr)
            if yr == 0:
                ws.cell(row, 3 + yr).value = 0
            else:
                prev = get_column_letter(2 + yr)
                ws.cell(row, 3 + yr).value = f"=IF('P&L'!{prev}{rev_r}=0,0,('P&L'!{cl}{rev_r}-'P&L'!{prev}{rev_r})/'P&L'!{prev}{rev_r})"
            ws.cell(row, 3 + yr).number_format = "0.0%"
        row += 1

        ws.cell(row, 1).value = "Gross Margin %"
        ws.cell(row, 2).value = "%"
        for yr in range(self.num_years):
            cl = get_column_letter(3 + yr)
            ws.cell(row, 3 + yr).value = f"=IF('P&L'!{cl}{rev_r}=0,0,'P&L'!{cl}{gp_r}/'P&L'!{cl}{rev_r})"
            ws.cell(row, 3 + yr).number_format = "0.0%"
        row += 1

        ws.cell(row, 1).value = "EBITDA Margin %"
        ws.cell(row, 2).value = "%"
        for yr in range(self.num_years):
            cl = get_column_letter(3 + yr)
            ws.cell(row, 3 + yr).value = f"=IF('P&L'!{cl}{rev_r}=0,0,'P&L'!{cl}{ebitda_r}/'P&L'!{cl}{rev_r})"
            ws.cell(row, 3 + yr).number_format = "0.0%"
        row += 2

        # Team Metrics
        style_section_header(ws.cell(row, 1), "TEAM METRICS", bg=Colors.MEDIUM_BLUE)
        ws.merge_cells(f"A{row}:{get_column_letter(self._end_col)}{row}")
        row += 1

        hc_r = self.row_refs.get("headcount_total", 8)
        ws.cell(row, 1).value = "Total Headcount"
        ws.cell(row, 2).value = "FTE"
        for yr in range(self.num_years):
            cl = get_column_letter(3 + yr)
            ws.cell(row, 3 + yr).value = f"='Headcount Plan'!{cl}{hc_r}"
            ws.cell(row, 3 + yr).number_format = "#,##0"
        row += 1

        ws.cell(row, 1).value = "Revenue per Employee"
        ws.cell(row, 2).value = self.currency
        for yr in range(self.num_years):
            cl = get_column_letter(3 + yr)
            ws.cell(row, 3 + yr).value = f"=IF('Headcount Plan'!{cl}{hc_r}=0,0,'P&L'!{cl}{rev_r}/'Headcount Plan'!{cl}{hc_r})"
            ws.cell(row, 3 + yr).number_format = "#,##0"
        row += 2

        # Cash Metrics
        style_section_header(ws.cell(row, 1), "CASH METRICS", bg=Colors.MEDIUM_BLUE)
        ws.merge_cells(f"A{row}:{get_column_letter(self._end_col)}{row}")
        row += 1

        ws.cell(row, 1).value = "Cash Balance"
        ws.cell(row, 2).value = self.currency
        for yr in range(self.num_years):
            cl = get_column_letter(3 + yr)
            ws.cell(row, 3 + yr).value = f"='Cash Flow'!{cl}{self.row_refs['cf_cumulative']}"
            ws.cell(row, 3 + yr).number_format = "#,##0"
        row += 1

        ws.cell(row, 1).value = "Net Cash Flow"
        ws.cell(row, 2).value = self.currency
        for yr in range(self.num_years):
            cl = get_column_letter(3 + yr)
            ws.cell(row, 3 + yr).value = f"='Cash Flow'!{cl}{self.row_refs['cf_net']}"
            ws.cell(row, 3 + yr).number_format = "#,##0"

    # ================================================================
    # SHEET 10: Sensitivity Analysis
    # ================================================================
    def _build_sensitivity(self):
        print("  [10/14] Sensitivity Analysis...")
        ws = self.wb.create_sheet("Sensitivity Analysis")
        for c, w in [("A", 30), ("B", 15), ("C", 15), ("D", 15)]:
            ws.column_dimensions[c].width = w

        style_title(ws["A1"], "SENSITIVITY ANALYSIS")
        ws.merge_cells("A1:D1")
        row = 3

        style_section_header(ws.cell(row, 1), "SCENARIO COMPARISON")
        ws.merge_cells(f"A{row}:D{row}")
        row += 1

        style_column_headers(ws, row, ["Metric", "Downside (-20%)", "Base Case", "Upside (+20%)"], end_col=4)
        row += 1

        # Terminal year column (dynamic, not hardcoded to Year 8)
        term_col = get_column_letter(3 + self._terminal_yr)
        term_label = f"Year {self._terminal_yr}"

        ws.cell(row, 1).value = f"{term_label} Revenue"
        ws.cell(row, 2).value = f"=0.8*C{row}"
        ws.cell(row, 3).value = f"='P&L'!{term_col}{self.row_refs['pnl_revenue']}"
        ws.cell(row, 4).value = f"=1.2*C{row}"
        for c in range(2, 5):
            ws.cell(row, c).number_format = "#,##0"
        row += 1

        ws.cell(row, 1).value = f"{term_label} EBITDA"
        ws.cell(row, 2).value = f"=0.7*C{row}"
        ws.cell(row, 3).value = f"='P&L'!{term_col}{self.row_refs['pnl_ebitda']}"
        ws.cell(row, 4).value = f"=1.3*C{row}"
        for c in range(2, 5):
            ws.cell(row, c).number_format = "#,##0"
        row += 1

        ws.cell(row, 1).value = f"{term_label} Cash"
        ws.cell(row, 2).value = f"=0.75*C{row}"
        ws.cell(row, 3).value = f"='Cash Flow'!{term_col}{self.row_refs['cf_cumulative']}"
        ws.cell(row, 4).value = f"=1.25*C{row}"
        for c in range(2, 5):
            ws.cell(row, c).number_format = "#,##0"

    # ================================================================
    # SHEET 11: Valuation
    # ================================================================
    def _build_valuation(self):
        print("  [11/14] Valuation...")
        ws = self.wb.create_sheet("Valuation")
        ws.column_dimensions["A"].width = 35
        ws.column_dimensions["B"].width = 15

        val_cfg = self.config.get("valuation", {})
        wacc = val_cfg.get("wacc", 0.15)
        terminal_growth = val_cfg.get("terminal_growth", 0.03)
        exit_multiple = val_cfg.get("exit_multiple", 5.0)

        style_title(ws["A1"], "VALUATION ANALYSIS")
        ws.merge_cells("A1:B1")
        row = 3

        style_section_header(ws.cell(row, 1), "DCF ASSUMPTIONS")
        ws.merge_cells(f"A{row}:B{row}")
        row += 1

        wacc_row = row
        for name, val, fmt in [("Discount Rate (WACC)", wacc, "0.0%"),
                                ("Terminal Growth Rate", terminal_growth, "0.0%"),
                                ("Exit Multiple (EV/Revenue)", exit_multiple, "0.0")]:
            ws.cell(row, 1).value = name
            ws.cell(row, 2).value = val
            ws.cell(row, 2).number_format = fmt
            row += 1
        exit_mult_row = row - 1
        row += 1

        style_section_header(ws.cell(row, 1), "VALUATION SUMMARY")
        ws.merge_cells(f"A{row}:B{row}")
        row += 1

        term_col = get_column_letter(3 + self._terminal_yr)
        term_label = f"Year {self._terminal_yr}"
        ws.cell(row, 1).value = f"{term_label} Revenue"
        ws.cell(row, 2).value = f"='P&L'!{term_col}{self.row_refs['pnl_revenue']}"
        ws.cell(row, 2).number_format = "#,##0"
        rev_row = row
        row += 1

        ws.cell(row, 1).value = f"{term_label} EBITDA"
        ws.cell(row, 2).value = f"='P&L'!{term_col}{self.row_refs['pnl_ebitda']}"
        ws.cell(row, 2).number_format = "#,##0"
        row += 1

        ws.cell(row, 1).value = f"Implied Exit Valuation ({exit_multiple}x Rev)"
        ws.cell(row, 2).value = f"=B{rev_row}*B{exit_mult_row}"
        ws.cell(row, 2).number_format = "#,##0"
        ws.cell(row, 2).font = Font(bold=True)
        exit_val_row = row
        row += 2

        # Total funding raised
        rounds = self.config["funding"].get("rounds", [])
        total_raised = sum(r.get("amount", 0) for r in rounds)
        ws.cell(row, 1).value = "Total Capital Raised"
        ws.cell(row, 2).value = total_raised
        ws.cell(row, 2).number_format = "#,##0"
        raised_row = row
        row += 1

        ws.cell(row, 1).value = "Return Multiple (Exit / Raised)"
        ws.cell(row, 2).value = f"=IF(B{raised_row}=0,0,B{exit_val_row}/B{raised_row})"
        ws.cell(row, 2).number_format = "0.0x"
        ws.cell(row, 2).font = Font(bold=True)

    # ================================================================
    # SHEET 12: Break-even Analysis
    # ================================================================
    def _build_breakeven(self):
        print("  [12/14] Break-even Analysis...")
        ws = self.wb.create_sheet("Break-even Analysis")
        self._set_col_widths(ws)

        style_title(ws["A1"], "BREAK-EVEN ANALYSIS")
        ws.merge_cells(f"A1:{get_column_letter(self._end_col)}1")
        row = 3

        # Per-stream contribution margin
        style_section_header(ws.cell(row, 1), "CONTRIBUTION MARGIN BY STREAM")
        ws.merge_cells(f"A{row}:{get_column_letter(self._end_col)}{row}")
        row += 1

        headers = ["Revenue Stream", "Unit"] + self._year_headers()
        style_column_headers(ws, row, headers, end_col=self._end_col)
        row += 1

        streams = self.config["revenue_streams"]
        rev_start = self.row_refs.get("revenue_start", 4)
        cogs_start = self.row_refs.get("cogs_start", 4)
        stream_rev_rows = []
        stream_cogs_rows = []
        stream_cm_rows = []

        for idx, s in enumerate(streams):
            name = s["name"]
            rev_sheet_row = rev_start + idx
            cogs_sheet_row = cogs_start + idx

            # Revenue per stream (reference Revenue sheet)
            ws.cell(row, 1).value = f"{name}: Revenue"
            ws.cell(row, 2).value = self.currency
            for yr in range(self.num_years):
                cl = get_column_letter(3 + yr)
                ws.cell(row, 3 + yr).value = f"=Revenue!{cl}{rev_sheet_row}"
                ws.cell(row, 3 + yr).number_format = "#,##0"
            stream_rev_rows.append(row)
            row += 1

            # COGS per stream (reference Operating Costs sheet)
            ws.cell(row, 1).value = f"{name}: COGS"
            ws.cell(row, 2).value = self.currency
            for yr in range(self.num_years):
                cl = get_column_letter(3 + yr)
                ws.cell(row, 3 + yr).value = f"='Operating Costs'!{cl}{cogs_sheet_row}"
                ws.cell(row, 3 + yr).number_format = "#,##0"
            stream_cogs_rows.append(row)
            row += 1

            # Contribution margin per stream
            ws.cell(row, 1).value = f"{name}: Contribution Margin"
            ws.cell(row, 2).value = self.currency
            ws.cell(row, 1).font = Font(bold=True)
            for yr in range(self.num_years):
                cl = get_column_letter(3 + yr)
                ws.cell(row, 3 + yr).value = f"={cl}{row-2}-{cl}{row-1}"
                ws.cell(row, 3 + yr).number_format = "#,##0"
            stream_cm_rows.append(row)
            row += 1

        row += 1

        # Total contribution margin
        style_section_header(ws.cell(row, 1), "BREAK-EVEN CALCULATION")
        ws.merge_cells(f"A{row}:{get_column_letter(self._end_col)}{row}")
        row += 1

        ws.cell(row, 1).value = "Total Revenue"
        ws.cell(row, 2).value = self.currency
        for yr in range(self.num_years):
            cl = get_column_letter(3 + yr)
            refs = "+".join(f"{cl}{r}" for r in stream_rev_rows)
            ws.cell(row, 3 + yr).value = f"={refs}"
            ws.cell(row, 3 + yr).number_format = "#,##0"
        total_rev_row = row
        row += 1

        ws.cell(row, 1).value = "Total Variable Costs (COGS)"
        ws.cell(row, 2).value = self.currency
        for yr in range(self.num_years):
            cl = get_column_letter(3 + yr)
            refs = "+".join(f"{cl}{r}" for r in stream_cogs_rows)
            ws.cell(row, 3 + yr).value = f"={refs}"
            ws.cell(row, 3 + yr).number_format = "#,##0"
        total_cogs_row = row
        row += 1

        ws.cell(row, 1).value = "Total Contribution Margin"
        ws.cell(row, 2).value = self.currency
        ws.cell(row, 1).font = Font(bold=True)
        for yr in range(self.num_years):
            cl = get_column_letter(3 + yr)
            ws.cell(row, 3 + yr).value = f"={cl}{total_rev_row}-{cl}{total_cogs_row}"
            ws.cell(row, 3 + yr).number_format = "#,##0"
        total_cm_row = row
        row += 1

        ws.cell(row, 1).value = "Contribution Margin %"
        ws.cell(row, 2).value = "%"
        for yr in range(self.num_years):
            cl = get_column_letter(3 + yr)
            ws.cell(row, 3 + yr).value = f"=IF({cl}{total_rev_row}=0,0,{cl}{total_cm_row}/{cl}{total_rev_row})"
            ws.cell(row, 3 + yr).number_format = "0.0%"
        row += 1

        # Fixed costs (from P&L references)
        ws.cell(row, 1).value = "Total Fixed Costs (OpEx + Headcount)"
        ws.cell(row, 2).value = self.currency
        for yr in range(self.num_years):
            cl = get_column_letter(3 + yr)
            ws.cell(row, 3 + yr).value = f"='P&L'!{cl}{self.row_refs['pnl_opex']}"
            ws.cell(row, 3 + yr).number_format = "#,##0"
        fixed_row = row
        row += 1

        # Break-even revenue
        ws.cell(row, 1).value = "Break-Even Revenue"
        ws.cell(row, 2).value = self.currency
        ws.cell(row, 1).font = Font(bold=True, color="CC0000")
        for yr in range(self.num_years):
            cl = get_column_letter(3 + yr)
            cm_pct_row = total_cm_row + 1
            ws.cell(row, 3 + yr).value = f"=IF({cl}{cm_pct_row}=0,0,{cl}{fixed_row}/{cl}{cm_pct_row})"
            ws.cell(row, 3 + yr).number_format = "#,##0"
        be_rev_row = row
        row += 1

        # Margin of safety
        ws.cell(row, 1).value = "Margin of Safety"
        ws.cell(row, 2).value = "%"
        ws.cell(row, 1).font = Font(bold=True)
        for yr in range(self.num_years):
            cl = get_column_letter(3 + yr)
            ws.cell(row, 3 + yr).value = f"=IF({cl}{total_rev_row}=0,0,({cl}{total_rev_row}-{cl}{be_rev_row})/{cl}{total_rev_row})"
            ws.cell(row, 3 + yr).number_format = "0.0%"
        style_total_row(ws, row, 1, self._end_col)

    # ================================================================
    # SHEET 13: Funding Cap Table
    # ================================================================
    def _build_cap_table(self):
        print("  [13/14] Funding Cap Table...")
        ws = self.wb.create_sheet("Funding Cap Table")
        for c, w in [("A", 25), ("B", 15), ("C", 15), ("D", 15), ("E", 15)]:
            ws.column_dimensions[c].width = w

        style_title(ws["A1"], "FUNDING & CAP TABLE")
        ws.merge_cells("A1:E1")
        row = 3

        style_section_header(ws.cell(row, 1), "FUNDING ROUNDS")
        ws.merge_cells(f"A{row}:E{row}")
        row += 1

        style_column_headers(ws, row, ["Round", "Amount", "Pre-Money", "Post-Money", "Dilution"], end_col=5)
        row += 1

        rounds = self.config["funding"].get("rounds", [])
        first_row = row
        for rnd in rounds:
            ws.cell(row, 1).value = rnd["name"]
            ws.cell(row, 2).value = rnd["amount"]
            ws.cell(row, 2).number_format = "#,##0"
            ws.cell(row, 3).value = rnd.get("pre_money", 0)
            ws.cell(row, 3).number_format = "#,##0"
            ws.cell(row, 4).value = f"=B{row}+C{row}"
            ws.cell(row, 4).number_format = "#,##0"
            ws.cell(row, 5).value = f"=IF(D{row}=0,0,B{row}/D{row})"
            ws.cell(row, 5).number_format = "0.0%"
            row += 1
        last_row = row - 1

        row += 1
        ws.cell(row, 1).value = "TOTAL RAISED"
        ws.cell(row, 1).font = Font(bold=True)
        ws.cell(row, 2).value = f"=SUM(B{first_row}:B{last_row})"
        ws.cell(row, 2).number_format = "#,##0"
        ws.cell(row, 2).font = Font(bold=True)
        style_total_row(ws, row, 1, 5)

    # ================================================================
    # SHEET 14: Charts Data
    # ================================================================
    def _build_charts_data(self):
        print("  [14/14] Charts Data...")
        ws = self.wb.create_sheet("Charts Data")
        ws.column_dimensions["A"].width = 25
        for i in range(2, 2 + self.num_years):
            ws.column_dimensions[get_column_letter(i)].width = 12

        end = 1 + self.num_years
        style_title(ws["A1"], "CHARTS DATA")
        ws.merge_cells(f"A1:{get_column_letter(end)}1")
        row = 3

        headers = ["Metric"] + self._year_headers()
        style_column_headers(ws, row, headers, end_col=end)
        row += 1

        metrics = [
            ("Revenue", "pnl_revenue", "P&L"),
            ("EBITDA", "pnl_ebitda", "P&L"),
            ("Net Income", "pnl_net_income", "P&L"),
            ("Cash Balance", "cf_cumulative", "Cash Flow"),
        ]

        for label, ref_key, sheet in metrics:
            ws.cell(row, 1).value = label
            ref_row = self.row_refs.get(ref_key, 4)
            for yr in range(self.num_years):
                cl = get_column_letter(3 + yr)
                ws.cell(row, 2 + yr).value = f"='{sheet}'!{cl}{ref_row}"
                ws.cell(row, 2 + yr).number_format = "#,##0"
            row += 1


# =============================================================================
# MAIN
# =============================================================================
def main():
    parser = argparse.ArgumentParser(
        description="Build a clean 14-sheet financial model from config (zero hardcoded data)",
        formatter_class=argparse.RawDescriptionHelpFormatter,
    )
    parser.add_argument("--config", "-c", help="Path to JSON config file")
    parser.add_argument("--company", help="Company name (uses minimal defaults)")
    parser.add_argument("--years", type=int, default=11, help="Years (default: 11 = Y0-Y10)")
    parser.add_argument("--output", "-o", help="Output .xlsx path")
    parser.add_argument("--up-to", dest="up_to", help="Build sheets up to this name")
    parser.add_argument("--sheets", nargs="+", help="Build only these specific sheets")
    parser.add_argument("--validate", "-v", action="store_true", help="Validate after build")
    args = parser.parse_args()

    # Load config
    if args.config:
        if not os.path.exists(args.config):
            print(f"ERROR: Config not found: {args.config}")
            sys.exit(1)
        with open(args.config, "r", encoding="utf-8") as f:
            config = json.load(f)
    elif args.company:
        config = {
            "company": args.company,
            "company_name": args.company,
            "starting_year": 2026,
            "tax_rate": 0.25,
            "general": {},
            "revenue_streams": [
                {"name": "Product", "price": 1000, "volume": 10, "growth": 0.30, "cogs_pct": 0.30}
            ],
            "fixed_costs": [{"name": "General & Admin", "annual_cost": 50000}],
            "headcount": {},
            "funding": {"rounds": [{"name": "Seed", "amount": 1000000, "year": 0, "pre_money": 5000000}]},
            "tam": {"streams": []},
            "sam": {"regions": []},
            "som": {"year8_revenue_m": 0},
        }
    else:
        print("ERROR: Provide --config or --company")
        sys.exit(1)

    # Output path
    company_name = config.get("company_name", config.get("company", "Model"))
    if args.output:
        output = args.output
    else:
        safe = company_name.replace(" ", "_")
        os.makedirs(".tmp", exist_ok=True)
        output = f".tmp/{safe}_financial_model.xlsx"

    # Build
    builder = FinancialModelBuilder(config, num_years=args.years)
    if args.up_to:
        builder.build_up_to(args.up_to)
    elif args.sheets:
        builder.build_sheets(args.sheets)
    else:
        builder.build_all()
    builder.save(output)

    # Validate
    if args.validate:
        print("\nValidating formulas...")
        try:
            from validate_excel_model import validate_excel_model
            success, report = validate_excel_model(output, verbose=True)
            print(report)
        except ImportError:
            print("WARNING: validate_excel_model not available")

    print(f"\nNext steps:")
    print(f"  1. Open in Excel: start {output}")
    print(f"  2. Validate: python execution/validate_excel_model.py --file {output}")


if __name__ == "__main__":
    main()
