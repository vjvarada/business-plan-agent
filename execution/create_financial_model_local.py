#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Local-First Financial Model Creator

Creates Excel files (.xlsx) using openpyxl, then validates all formulas using
the 'formulas' library. This workflow allows full formula visibility and
debugging without opening Excel or using Google Sheets API.

Workflow:
    1. CREATE: Build Excel file with openpyxl (formulas, formatting, cross-sheet refs)
    2. VALIDATE: Compute all formulas with 'formulas' library (detect #REF!, #DIV/0!, etc.)
    3. UPLOAD: Push to Google Sheets via sync_to_cloud.py

IMPORTANT:
    This script is for draft/prototype local modeling and formula experiments.
    It does NOT generate the full production 14-sheet template structure.
    For production model creation, use:
        create_financial_model.py --from-template

Usage:
    python create_financial_model_local.py --company "MyCompany" --sample
    python create_financial_model_local.py --company "MyCompany" --config config.json
    python create_financial_model_local.py --company "Test" --validate-only .tmp/test.xlsx

Benefits over Google Sheets API:
    - No rate limits (429 errors)
    - Full formula visibility in code
    - Automatic validation before upload
    - Easy debugging (open in Excel or compute programmatically)
    - Git-trackable changes
"""

import argparse
import json
import os
import sys
from datetime import datetime

try:
    import openpyxl
    from openpyxl import Workbook
    from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
    from openpyxl.utils import get_column_letter
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)


class Colors:
    """Standard color palette for financial models (hex codes for openpyxl)."""

    TITLE_BLUE = "335080"  # Main titles - RGB(0.20, 0.30, 0.50)
    DARK_BLUE = "336699"  # Section headers - RGB(0.20, 0.40, 0.60)
    MEDIUM_BLUE = "6699CC"  # Category headers - RGB(0.40, 0.60, 0.80)
    SECTION_A = "4D80B3"  # Section A category - RGB(0.30, 0.50, 0.70)
    LIGHT_BLUE = "D8EAF9"  # Zebra stripe rows - RGB(0.85, 0.92, 0.98)
    LIGHT_GRAY = "F2F2F2"  # Column headers - RGB(0.95, 0.95, 0.95)
    GREEN = "E5F8E5"  # Total/summary rows - RGB(0.90, 0.97, 0.90)
    WHITE = "FFFFFF"
    BLACK = "000000"
    URL_BLUE = "1A4CB3"  # URL text - RGB(0.10, 0.30, 0.70)


# Year headers for 11-year model (Year 0 through Year 10)
YEAR_HEADERS = [
    "Year 0",
    "Year 1",
    "Year 2",
    "Year 3",
    "Year 4",
    "Year 5",
    "Year 6",
    "Year 7",
    "Year 8",
    "Year 9",
    "Year 10",
]

# Sample configuration (can be overridden via --config)
SAMPLE_CONFIG = {
    "general": {
        "tax_rate": 0.25,
        "capex_y0": 150000,
        "capex_annual": 50000,
        "depreciation_years": 5,
        "debtor_days": 45,
        "creditor_days": 30,
        "interest_rate": 0.10,
        "equity_y0": 300000,
        "debt_y0": 0,
        "cost_inflation": 0.05,
    },
    "revenue_streams": [
        {
            "name": "Product A",
            "price": 10000,
            "volume": 10,
            "growth": 0.30,
            "cogs_pct": 0.25,
        },
        {
            "name": "Product B",
            "price": 5000,
            "volume": 20,
            "growth": 0.25,
            "cogs_pct": 0.40,
        },
    ],
    "fixed_costs": {
        "Salaries": 360000,
        "Rent & Utilities": 36000,
        "Marketing": 60000,
        "R&D": 72000,
        "Insurance": 12000,
    },
}


def apply_header_style(cell, bg=Colors.DARK_BLUE, fg=Colors.WHITE, size=12, bold=True):
    """Apply standard header styling to a cell."""
    cell.font = Font(name="Calibri", size=size, bold=bold, color=fg)
    cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = Border(
        left=Side(style="thin"),
        right=Side(style="thin"),
        top=Side(style="thin"),
        bottom=Side(style="thin"),
    )


def apply_data_style(cell, is_currency=False, is_percent=False, bold=False, bg=None):
    """Apply standard data cell styling."""
    if is_currency:
        cell.number_format = "#,##0"
    elif is_percent:
        cell.number_format = "0.0%"

    if bold:
        cell.font = Font(bold=True)

    if bg:
        cell.fill = PatternFill(start_color=bg, end_color=bg, fill_type="solid")


def validate_model(filepath: str) -> bool:
    """
    Validate the Excel model by computing all formulas.
    Returns True if no errors found, False otherwise.
    """
    try:
        from validate_excel_model import validate_excel_model

        success, report = validate_excel_model(filepath, verbose=True)
        print(report)
        return success
    except ImportError:
        # Fallback: try to import formulas directly
        try:
            import formulas
            import numpy as np

            print(f"\nValidating: {filepath}")
            print("=" * 60)

            xl_model = formulas.ExcelModel().loads(filepath).finish()
            solution = xl_model.calculate()

            errors = []
            for key, value in solution.items():
                try:
                    val = value.value[0][0] if hasattr(value, "value") else value
                except:
                    val = str(value)

                # Check for Excel errors
                val_str = str(val)
                if any(
                    err in val_str
                    for err in ["#REF!", "#VALUE!", "#DIV/0!", "#NAME?", "#N/A"]
                ):
                    errors.append(f"{key}: {val_str}")

                # Check for numeric errors
                if isinstance(val, (int, float)):
                    if np.isnan(val):
                        errors.append(f"{key}: NaN")
                    elif np.isinf(val):
                        errors.append(f"{key}: Infinity (division by zero?)")

            if errors:
                print(f"\n❌ VALIDATION FAILED - {len(errors)} errors found:")
                for err in errors[:20]:
                    print(f"   {err}")
                if len(errors) > 20:
                    print(f"   ... and {len(errors) - 20} more")
                return False
            else:
                print(f"\n✅ VALIDATION PASSED")
                print(f"   {len(solution)} cells computed successfully")
                return True

        except ImportError:
            print("\n⚠️  Validation skipped (formulas library not installed)")
            print("   Run: pip install formulas")
            print("   Then re-run to validate formulas")
            return True  # Don't block on missing validation


def create_financial_model_local(
    company_name: str, config: dict = None, use_sample: bool = False
) -> str:
    """
    Create a financial model Excel file locally using openpyxl.

    Args:
        company_name: Name of the company (used for filename)
        config: Configuration dict with revenue_streams, fixed_costs, general params
        use_sample: If True, use the built-in sample configuration

    Returns:
        Path to the created Excel file
    """
    if use_sample:
        config = SAMPLE_CONFIG

    if config is None:
        config = {
            "general": {
                "tax_rate": 0.25,
                "capex_y0": 500000,
                "capex_annual": 100000,
                "depreciation_years": 5,
                "debtor_days": 45,
                "creditor_days": 30,
                "interest_rate": 0.08,
                "equity_y0": 1000000,
                "debt_y0": 0,
                "cost_inflation": 0.03,
            },
            "revenue_streams": [
                {
                    "name": "Product A",
                    "price": 1000,
                    "volume": 100,
                    "growth": 0.25,
                    "cogs_pct": 0.30,
                },
                {
                    "name": "Product B",
                    "price": 2000,
                    "volume": 50,
                    "growth": 0.30,
                    "cogs_pct": 0.35,
                },
            ],
            "fixed_costs": {"Salaries": 500000, "Rent": 50000, "Marketing": 100000},
        }

    # Create workbook
    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    print(f"Creating financial model for {company_name}...")
    print("=" * 60)
    print("⚠️  Draft mode: reduced local model (non-production baseline)")
    print("   For full 14-sheet production models,")
    print("   use create_financial_model.py --from-template")

    # =========================================================================
    # SHEET 1: ASSUMPTIONS
    # =========================================================================
    print("  📊 Creating Assumptions sheet...")
    ws = wb.create_sheet("Assumptions", 0)
    ws.column_dimensions["A"].width = 35
    ws.column_dimensions["B"].width = 15
    ws.column_dimensions["C"].width = 10

    # Title
    ws["A1"] = "FINANCIAL MODEL ASSUMPTIONS"
    apply_header_style(ws["A1"], bg=Colors.TITLE_BLUE, size=14)
    ws.merge_cells("A1:M1")

    # General Parameters Section
    row = 3
    ws[f"A{row}"] = "GENERAL PARAMETERS"
    apply_header_style(ws[f"A{row}"])
    ws.merge_cells(f"A{row}:C{row}")
    row += 1

    general_params = [
        ("Tax Rate", config["general"].get("tax_rate", 0.25), "%"),
        ("CapEx Year 0", config["general"].get("capex_y0", 500000), "USD"),
        ("CapEx Annual", config["general"].get("capex_annual", 100000), "USD"),
        (
            "Depreciation Period",
            config["general"].get("depreciation_years", 5),
            "Years",
        ),
        ("Debtor Days", config["general"].get("debtor_days", 45), "Days"),
        ("Creditor Days", config["general"].get("creditor_days", 30), "Days"),
        ("Interest Rate", config["general"].get("interest_rate", 0.08), "%"),
        ("Cost Inflation", config["general"].get("cost_inflation", 0.03), "%"),
    ]

    for name, val, unit in general_params:
        ws[f"A{row}"] = name
        ws[f"B{row}"] = val
        ws[f"C{row}"] = unit
        if unit == "%":
            ws[f"B{row}"].number_format = "0.00%"
        elif unit == "USD":
            ws[f"B{row}"].number_format = "#,##0"
        row += 1

    # Revenue Streams Section
    row += 1
    ws[f"A{row}"] = "REVENUE STREAMS"
    apply_header_style(ws[f"A{row}"])
    ws.merge_cells(f"A{row}:M{row}")
    row += 1

    # Revenue headers
    headers = ["Stream Name", "Price (Y0)", "Unit"] + YEAR_HEADERS
    for col, header in enumerate(headers, 1):
        apply_header_style(ws.cell(row, col, header), bg=Colors.MEDIUM_BLUE, size=10)
    row += 1

    revenue_start_row = row
    for stream in config.get("revenue_streams", []):
        ws.cell(row, 1, stream["name"])
        ws.cell(row, 2, stream["price"])
        ws.cell(row, 2).number_format = "#,##0"
        ws.cell(row, 3, "USD")

        # Volume projections with growth formula
        for yr in range(11):
            col = yr + 4
            if yr == 0:
                ws.cell(row, col, stream["volume"])
                ws.cell(row, col).number_format = "0"
            else:
                prev_col = get_column_letter(col - 1)
                growth = stream.get("growth", 0.25)
                ws.cell(row, col).value = f"={prev_col}{row}*(1+{growth})"
                ws.cell(row, col).number_format = "0"
        row += 1

    # =========================================================================
    # SHEET 2: REVENUE
    # =========================================================================
    print("  📈 Creating Revenue sheet...")
    ws = wb.create_sheet("Revenue")
    ws.column_dimensions["A"].width = 30

    ws["A1"] = "REVENUE PROJECTIONS"
    apply_header_style(ws["A1"], bg=Colors.TITLE_BLUE, size=14)
    ws.merge_cells("A1:M1")

    row = 3
    headers = ["Revenue Stream", "Unit"] + YEAR_HEADERS
    for col, header in enumerate(headers, 1):
        apply_header_style(ws.cell(row, col, header))
    row += 1

    start_row = row
    for idx, stream in enumerate(config.get("revenue_streams", []), 1):
        ws.cell(row, 1, stream["name"])
        ws.cell(row, 2, "USD")

        for yr in range(11):
            col = yr + 3
            yr_col = get_column_letter(yr + 4)
            assumptions_row = revenue_start_row + idx - 1
            # Revenue = Price × Volume
            ws.cell(row, col).value = (
                f"=Assumptions!$B${assumptions_row}*Assumptions!{yr_col}${assumptions_row}"
            )
            ws.cell(row, col).number_format = "#,##0"
        row += 1

    # Total Revenue
    row += 1
    ws.cell(row, 1, "TOTAL REVENUE")
    ws.cell(row, 1).font = Font(bold=True, size=11)
    ws.cell(row, 2, "USD")
    for yr in range(11):
        col = yr + 3
        col_letter = get_column_letter(col)
        ws.cell(row, col).value = f"=SUM({col_letter}{start_row}:{col_letter}{row-2})"
        ws.cell(row, col).number_format = "#,##0"
        ws.cell(row, col).font = Font(bold=True)
        ws.cell(row, col).fill = PatternFill(
            start_color=Colors.GREEN, end_color=Colors.GREEN, fill_type="solid"
        )
    total_rev_row = row

    # =========================================================================
    # SHEET 3: P&L
    # =========================================================================
    print("  📉 Creating P&L sheet...")
    ws = wb.create_sheet("P&L")
    ws.column_dimensions["A"].width = 30

    ws["A1"] = "PROFIT & LOSS STATEMENT"
    apply_header_style(ws["A1"], bg=Colors.TITLE_BLUE, size=14)
    ws.merge_cells("A1:M1")

    row = 3
    headers = ["Line Item", "Unit"] + YEAR_HEADERS
    for col, header in enumerate(headers, 1):
        apply_header_style(ws.cell(row, col, header))
    row += 1

    # Revenue (linked from Revenue sheet)
    ws.cell(row, 1, "Revenue")
    ws.cell(row, 2, "USD")
    for yr in range(11):
        col = yr + 3
        col_letter = get_column_letter(col)
        ws.cell(row, col).value = f"=Revenue!{col_letter}{total_rev_row}"
        ws.cell(row, col).number_format = "#,##0"
    rev_row = row
    row += 1

    # COGS
    ws.cell(row, 1, "Cost of Goods Sold (COGS)")
    ws.cell(row, 2, "USD")
    avg_cogs = sum(
        s.get("cogs_pct", 0.30) for s in config.get("revenue_streams", [])
    ) / max(len(config.get("revenue_streams", [])), 1)
    for yr in range(11):
        col = yr + 3
        col_letter = get_column_letter(col)
        ws.cell(row, col).value = f"={col_letter}{rev_row}*{avg_cogs}"
        ws.cell(row, col).number_format = "#,##0"
    cogs_row = row
    row += 1

    # Gross Profit
    ws.cell(row, 1, "Gross Profit")
    ws.cell(row, 1).font = Font(bold=True)
    ws.cell(row, 2, "USD")
    for yr in range(11):
        col = yr + 3
        col_letter = get_column_letter(col)
        ws.cell(row, col).value = f"={col_letter}{rev_row}-{col_letter}{cogs_row}"
        ws.cell(row, col).number_format = "#,##0"
        ws.cell(row, col).font = Font(bold=True)
    gp_row = row
    row += 1

    # Gross Margin %
    ws.cell(row, 1, "Gross Margin %")
    ws.cell(row, 2, "%")
    for yr in range(11):
        col = yr + 3
        col_letter = get_column_letter(col)
        ws.cell(row, col).value = (
            f"=IF({col_letter}{rev_row}=0,0,{col_letter}{gp_row}/{col_letter}{rev_row})"
        )
        ws.cell(row, col).number_format = "0.0%"
    row += 2

    # Operating Expenses Section
    ws.cell(row, 1, "Operating Expenses")
    ws.cell(row, 1).font = Font(bold=True)
    opex_header_row = row
    row += 1

    opex_start_row = row
    fixed_costs = config.get("fixed_costs", {})
    if isinstance(fixed_costs, list):
        fixed_costs = {item["name"]: item["value"] for item in fixed_costs}

    for cost_name, cost_value in fixed_costs.items():
        ws.cell(row, 1, f"  {cost_name}")
        ws.cell(row, 2, "USD")
        for yr in range(11):
            col = yr + 3
            # Apply cost inflation after Year 0
            if yr == 0:
                ws.cell(row, col, cost_value)
            else:
                prev_col = get_column_letter(col - 1)
                inflation = config["general"].get("cost_inflation", 0.03)
                ws.cell(row, col).value = f"={prev_col}{row}*(1+{inflation})"
            ws.cell(row, col).number_format = "#,##0"
        row += 1
    opex_end_row = row - 1

    # Total OpEx
    row += 1
    ws.cell(row, 1, "Total Operating Expenses")
    ws.cell(row, 1).font = Font(bold=True)
    ws.cell(row, 2, "USD")
    for yr in range(11):
        col = yr + 3
        col_letter = get_column_letter(col)
        ws.cell(row, col).value = (
            f"=SUM({col_letter}{opex_start_row}:{col_letter}{opex_end_row})"
        )
        ws.cell(row, col).number_format = "#,##0"
        ws.cell(row, col).font = Font(bold=True)
    total_opex_row = row
    row += 2

    # EBITDA
    ws.cell(row, 1, "EBITDA")
    ws.cell(row, 1).font = Font(bold=True, size=11)
    ws.cell(row, 2, "USD")
    for yr in range(11):
        col = yr + 3
        col_letter = get_column_letter(col)
        ws.cell(row, col).value = f"={col_letter}{gp_row}-{col_letter}{total_opex_row}"
        ws.cell(row, col).number_format = "#,##0"
        ws.cell(row, col).font = Font(bold=True)
        ws.cell(row, col).fill = PatternFill(
            start_color=Colors.GREEN, end_color=Colors.GREEN, fill_type="solid"
        )
    ebitda_row = row
    row += 1

    # EBITDA Margin %
    ws.cell(row, 1, "EBITDA Margin %")
    ws.cell(row, 2, "%")
    for yr in range(11):
        col = yr + 3
        col_letter = get_column_letter(col)
        ws.cell(row, col).value = (
            f"=IF({col_letter}{rev_row}=0,0,{col_letter}{ebitda_row}/{col_letter}{rev_row})"
        )
        ws.cell(row, col).number_format = "0.0%"

    # =========================================================================
    # SHEET 4: SOURCES & REFERENCES
    # =========================================================================
    print("  📚 Creating Sources & References sheet...")
    ws = wb.create_sheet("Sources & References")
    ws.column_dimensions["A"].width = 40
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 50

    ws["A1"] = "SOURCES & REFERENCES"
    apply_header_style(ws["A1"], bg=Colors.TITLE_BLUE, size=14)
    ws.merge_cells("A1:C1")

    ws["A3"] = "Populate with market research data using serp_market_research.py"
    ws["A5"] = "TAM/SAM/SOM data, competitor research, and industry benchmarks go here."

    # =========================================================================
    # SAVE FILE
    # =========================================================================
    os.makedirs(".tmp", exist_ok=True)
    filename = f'{company_name.replace(" ", "_")}_financial_model.xlsx'
    filepath = os.path.join(".tmp", filename)
    wb.save(filepath)

    print()
    print("=" * 60)
    print(f"✅ Financial model created!")
    print(f"   File: {filepath}")
    print(f"   Sheets: {', '.join(wb.sheetnames)}")

    return filepath


def main():
    parser = argparse.ArgumentParser(
        description="Create local Excel financial model with formula validation",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
python create_financial_model_local.py --company "MyCompany" --sample
  python create_financial_model_local.py --company "MyCompany" --config config.json
  python create_financial_model_local.py --validate-only .tmp/existing_model.xlsx

Workflow:
  1. CREATE:   Builds Excel file with formulas using openpyxl
  2. VALIDATE: Computes all formulas using 'formulas' library
  3. UPLOAD:   Push to Google Sheets via sync_to_cloud.py
        """,
    )
    parser.add_argument("--company", help="Company name (for filename)")
    parser.add_argument("--config", help="Path to JSON config file")
    parser.add_argument(
        "--sample", action="store_true", help="Use built-in sample config"
    )
    parser.add_argument(
        "--validate-only", metavar="FILE", help="Only validate existing Excel file"
    )
    parser.add_argument(
        "--skip-validation", action="store_true", help="Skip formula validation"
    )

    args = parser.parse_args()

    # Validate-only mode
    if args.validate_only:
        success = validate_model(args.validate_only)
        sys.exit(0 if success else 1)

    # Create mode requires company name
    if not args.company:
        parser.error("--company is required (or use --validate-only)")

    # Load config if provided
    config = None
    if args.config and os.path.exists(args.config):
        with open(args.config, "r") as f:
            config = json.load(f)

    try:
        # Step 1: Create the Excel file
        filepath = create_financial_model_local(
            company_name=args.company, config=config, use_sample=args.sample
        )

        # Step 2: Validate formulas (unless skipped)
        if not args.skip_validation:
            print()
            print("Validating formulas...")
            success = validate_model(filepath)
            if not success:
                print()
                print("⚠️  Fix the errors above before uploading to Google Sheets")
                sys.exit(1)

        # Step 3: Print next steps
        print()
        print("Next steps:")
        print(f"  1. Review in Excel: start .tmp\\{os.path.basename(filepath)}")
        print(f"  2. Upload to Google Sheets:")
        print(f"     python execution/sync_to_cloud.py --file {filepath}")

        return filepath

    except Exception as e:
        print(f"\n❌ Error: {e}")
        import traceback

        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
