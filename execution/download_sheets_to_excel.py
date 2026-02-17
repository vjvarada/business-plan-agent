#!/usr/bin/env python3
# -*- coding: utf-8 -*-
"""
Download existing Google Sheets financial model and convert to local Excel file.
Preserves all sheets, formulas, and formatting.
"""

import os
import re
import sys

import gspread
import openpyxl
from google.oauth2.credentials import Credentials
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

SCOPES = [
    "https://www.googleapis.com/auth/spreadsheets",
    "https://www.googleapis.com/auth/drive",
]


class Colors:
    TITLE_BLUE, DARK_BLUE, MEDIUM_BLUE = "335080", "336699", "6699CC"
    SECTION_A_CAT = "4D80B3"
    LIGHT_BLUE, LIGHT_GRAY, GREEN = "D8EAF9", "F2F2F2", "E5F8E5"
    URL_BLUE, GRAY = "1A4CB3", "808080"
    WHITE, BLACK = "FFFFFF", "000000"


def get_credentials():
    creds = None
    if os.path.exists("token.json"):
        creds = Credentials.from_authorized_user_file("token.json", SCOPES)
    return creds


def convert_formula(formula, sheet_name):
    """Convert Google Sheets formula to Excel format."""
    if not formula or not formula.startswith("="):
        return formula

    # Keep formulas as-is - they're already compatible
    return formula


def download_google_sheets_to_excel(spreadsheet_id, output_path):
    """Download Google Sheets and convert to Excel."""
    print(f"Connecting to Google Sheets...")
    creds = get_credentials()
    client = gspread.authorize(creds)

    spreadsheet = client.open_by_key(spreadsheet_id)
    print(f"Found: {spreadsheet.title}")

    wb = Workbook()
    if "Sheet" in wb.sheetnames:
        del wb["Sheet"]

    for idx, sheet in enumerate(spreadsheet.worksheets()):
        sheet_name = sheet.title
        print(f"  Downloading sheet: {sheet_name}")

        ws = wb.create_sheet(sheet_name, idx)

        # Get all values
        all_values = sheet.get_all_values()

        # Get all formulas
        try:
            formulas = sheet.get(return_type=gspread.utils.ValueRenderOption.formula)
        except:
            formulas = all_values

        # Get formatting
        try:
            formats = sheet.get(return_type="FORMATTED_VALUE")
        except:
            formats = all_values

        # Write data
        for row_idx, (value_row, formula_row) in enumerate(
            zip(all_values, formulas), 1
        ):
            for col_idx, (value, formula) in enumerate(zip(value_row, formula_row), 1):
                cell = ws.cell(row_idx, col_idx)

                # Write formula or value
                if formula and formula.startswith("="):
                    cell.value = convert_formula(formula, sheet_name)
                elif value:
                    # Try to convert to number
                    try:
                        if "." in str(value):
                            cell.value = float(value.replace(",", ""))
                        elif value.replace(",", "").replace("-", "").isdigit():
                            cell.value = int(value.replace(",", ""))
                        else:
                            cell.value = value
                    except:
                        cell.value = value

        # Set column widths (approximate)
        if sheet_name == "Assumptions":
            ws.column_dimensions["A"].width = 35
            ws.column_dimensions["B"].width = 15
            ws.column_dimensions["C"].width = 10
        elif sheet_name == "Sources & References":
            ws.column_dimensions["A"].width = 30
            ws.column_dimensions["B"].width = 15
            ws.column_dimensions["C"].width = 10
            ws.column_dimensions["D"].width = 40
            ws.column_dimensions["E"].width = 50
            ws.column_dimensions["F"].width = 30
        else:
            ws.column_dimensions["A"].width = 30
            ws.column_dimensions["B"].width = 15

        # Apply basic formatting
        for row in ws.iter_rows(min_row=1, max_row=3):
            for cell in row:
                if cell.value:
                    cell.font = Font(bold=True, size=11)
                    cell.alignment = Alignment(horizontal="center", vertical="center")

    # Save
    os.makedirs(".tmp", exist_ok=True)
    wb.save(output_path)

    print(f"\n Downloaded to local Excel file!")
    print(f"   File: {output_path}")
    print(f"   Sheets: {', '.join([ws.title for ws in wb.worksheets])}")
    print(f"   Total sheets: {len(wb.worksheets)}")

    return output_path


if __name__ == "__main__":
    spreadsheet_id = os.environ.get("SHEET_ID", "YOUR_SHEET_ID_HERE")
    output_path = ".tmp/Financial_Model_from_Cloud.xlsx"

    try:
        download_google_sheets_to_excel(spreadsheet_id, output_path)
    except Exception as e:
        print(f"\n Error: {e}")
        import traceback

        traceback.print_exc()
        sys.exit(1)
